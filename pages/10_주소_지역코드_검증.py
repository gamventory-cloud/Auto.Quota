# -*- coding: utf-8 -*-
"""
주소 → 지역코드(areaM / areaD) 검증
VBA 애드인(우편번호_주소_체크.xlam) 이식판  v1.0

핵심 흐름
  1) 주소 원본(ID + 주소)에서 ID→주소 사전 생성
  2) 검증 대상(ID + areaM + areaD)의 각 행마다 주소를 찾아 코드 계산
  3) 입력값과 계산값을 대조해 일치 / 불일치 / 구군미확인 등으로 판정
"""
from __future__ import annotations

import io
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

# ──────────────────────────────────────────────────────────────
# utils.py 의존성 (없어도 단독 실행 가능)
# ──────────────────────────────────────────────────────────────
try:
    from utils import check_password  # type: ignore
except Exception:  # pragma: no cover
    def check_password() -> bool:
        return True


# ══════════════════════════════════════════════════════════════
# CORE : 순수 로직 (streamlit 호출 없음 — 단위 테스트 가능)
# ══════════════════════════════════════════════════════════════

# 기준표 원본 (VBA GenerateCodeDB 와 동일)
_RAW_CODE_DB: List[Tuple[str, int, str]] = [
    ("서울", 1, "강북구,1,광진구,2,노원구,3,도봉구,4,동대문구,5,성동구,6,성북구,7,중랑구,8,"
                "마포구,9,서대문구,10,은평구,11,용산구,12,종로구,13,중구,14,강남구,15,강동구,16,"
                "서초구,17,송파구,18,강서구,19,관악구,20,구로구,21,금천구,22,동작구,23,양천구,24,영등포구,25"),
    ("부산", 2, "강서구,1,금정구,2,기장군,3,남구,4,동구,5,동래구,6,부산진구,7,북구,8,사상구,9,"
                "사하구,10,서구,11,수영구,12,연제구,13,영도구,14,중구,15,해운대구,16"),
    ("대구", 3, "남구,1,달서구,2,달성군,3,동구,4,북구,5,서구,6,수성구,7,중구,8,군위군,9"),
    ("인천", 4, "강화군,1,계양구,2,남동구,4,동구,5,부평구,6,서구,7,연수구,8,옹진군,9,중구,10,미추홀구,11"),
    ("광주", 5, "동구,1,서구,2,남구,3,북구,4,광산구,5"),
    ("대전", 6, "대덕구,1,동구,2,서구,3,유성구,4,중구,5"),
    ("울산", 7, "남구,1,동구,2,북구,3,울주군,4,중구,5"),
    ("경기", 8, "가평군,1,고양시,2,과천시,3,광명시,4,광주시,5,구리시,6,군포시,7,김포시,8,남양주시,9,"
                "동두천시,10,부천시,11,성남시,12,수원시,13,시흥시,14,안산시,15,안성시,16,안양시,17,"
                "양주시,18,양평군,19,여주시,20,연천군,21,오산시,22,용인시,23,의왕시,24,의정부시,25,"
                "이천시,26,파주시,27,평택시,28,포천시,29,하남시,30,화성시,31"),
    ("강원", 9, "강릉시,1,고성군,2,동해시,3,삼척시,4,속초시,5,양구군,6,양양군,7,영월군,8,원주시,9,"
                "인제군,10,정선군,11,철원군,12,춘천시,13,태백시,14,평창군,15,홍천군,16,화천군,17,횡성군,18"),
    ("충북", 10, "괴산군,1,단양군,2,보은군,3,영동군,4,옥천군,5,음성군,6,제천시,7,증평군,8,진천군,9,"
                 "청원군,10,청주시,11,충주시,12"),
    ("충남", 11, "계룡시,1,공주시,2,금산군,3,논산시,4,당진시,5,보령시,6,부여군,7,서산시,8,서천군,9,"
                 "아산시,10,예산군,12,천안시,13,청양군,14,태안군,15,홍성군,16"),
    ("전북", 12, "고창군,1,군산시,2,김제시,3,남원시,4,무주군,5,부안군,6,순창군,7,완주군,8,익산시,9,"
                 "임실군,10,장수군,11,전주시,12,정읍시,13,진안군,14"),
    ("전남", 13, "강진군,1,고흥군,2,곡성군,3,광양시,4,구례군,5,나주시,6,담양군,7,목포시,8,무안군,9,"
                 "보성군,10,순천시,11,신안군,12,여수시,13,영광군,14,영암군,15,완도군,16,장성군,17,"
                 "장흥군,18,진도군,19,함평군,20,해남군,21,화순군,22"),
    ("경북", 14, "경산시,1,경주시,2,고령군,3,구미시,4,김천시,6,문경시,7,봉화군,8,상주시,9,성주군,10,"
                 "안동시,11,영덕군,12,영양군,13,영주시,14,영천시,15,예천군,16,울진군,17,의성군,18,"
                 "청도군,19,청송군,20,칠곡군,21,포항시,22,울릉군,23"),
    ("경남", 15, "거제시,1,거창군,2,고성군,3,김해시,4,남해군,5,밀양시,6,사천시,7,산청군,8,양산시,9,"
                 "의령군,10,진주시,11,창녕군,12,창원시,13,통영시,14,하동군,15,함안군,16,함양군,17,합천군,18"),
    ("제주", 16, "제주시,1,서귀포시,2"),
    ("세종", 17, "소정면,1,전의면,2,전동면,3,조치원읍,4,연서면,5,연동면,6,연기면,7,부강면,8,도담동,9,"
                 "장군면,10,한솔동,11,금남면,12,보람동,13,새롬동,14,아름동,15,고운동,16,종촌동,17,"
                 "소담동,18,대평동,19,가람동,20,다정동,21,해밀동,22,산울동,23,누리동,24,한별동,25,"
                 "반곡동,26,집현동,27,합강동,28,다솜동,29,용호동,30,나성동,31,세종동,32,어진동,33"),
]

# 시도 표기 변형 → 표준 명칭
SIDO_ALIASES: Dict[str, List[str]] = {
    "서울": ["서울"], "부산": ["부산"], "대구": ["대구"], "인천": ["인천"],
    "광주": ["광주"], "대전": ["대전"], "울산": ["울산"], "세종": ["세종"],
    "경기": ["경기"], "강원": ["강원"],
    "충북": ["충북", "충청북"], "충남": ["충남", "충청남"],
    "전북": ["전북", "전라북"], "전남": ["전남", "전라남"],
    "경북": ["경북", "경상북"], "경남": ["경상남", "경남"],
    "제주": ["제주"],
}

# 도로명/마을 이름 → 법정동 (VBA ConvertSejongRoadToDong)
_RAW_ROAD_RULES: List[Tuple[str, str, str]] = [
    # (시도, 주소에 포함된 문구, 대체할 시군구명)
    ("세종", "조치원", "조치원읍"), ("세종", "이화", "조치원읍"),
    ("세종", "전의", "전의면"), ("세종", "노곡", "전의면"),
    ("세종", "소정", "소정면"), ("세종", "전동", "전동면"),
    ("세종", "연서", "연서면"), ("세종", "연동", "연동면"),
    ("세종", "연기", "연기면"), ("세종", "부강", "부강면"),
    ("세종", "장군", "장군면"), ("세종", "금남", "금남면"),
    ("세종", "밝은뜰", "고운동"), ("세종", "마음로", "고운동"),
    ("세종", "도움3로", "종촌동"), ("세종", "도움", "어진동"),
    ("세종", "시청대로", "보람동"), ("세종", "남세종로", "보람동"),
    ("세종", "해밀", "해밀동"), ("세종", "새롬", "새롬동"),
    ("세종", "반곡", "반곡동"), ("세종", "보듬", "도담동"),
    ("세종", "가람", "가람동"), ("세종", "나성", "나성동"),
    ("세종", "한누리", "나성동"), ("세종", "다솜", "다솜동"),
    ("세종", "갈매", "집현동"), ("세종", "달빛", "종촌동"),
]

RESULT_MATCH = "일치"
RESULT_DIFF = "불일치"
RESULT_NO_SGG = "구군미확인"
RESULT_NO_SIDO = "주소해석불가"
RESULT_NO_ADDR = "주소없음"
RESULT_NO_ID = "ID없음"
RESULT_NO_INPUT = "입력값없음"


def build_default_code_df() -> pd.DataFrame:
    """기준표를 긴 형태 DataFrame으로 전개."""
    rows = []
    for sido, mcode, blob in _RAW_CODE_DB:
        parts = [p.strip() for p in blob.split(",")]
        for i in range(0, len(parts), 2):
            rows.append({
                "시도": sido,
                "areaM": int(mcode),
                "시군구": parts[i],
                "areaD": int(parts[i + 1]),
            })
    return pd.DataFrame(rows)


def build_default_road_df() -> pd.DataFrame:
    return pd.DataFrame(_RAW_ROAD_RULES, columns=["시도", "포함문구", "시군구"])


def nospace(value) -> str:
    """공백 제거 + 문자열화. NaN 안전."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return re.sub(r"\s+", "", str(value))


class AreaResolver:
    """기준표 + 도로명 규칙을 받아 주소 한 건을 코드로 변환."""

    def __init__(self, code_df: pd.DataFrame, road_df: Optional[pd.DataFrame] = None,
                 infer_sido: bool = True):
        self.infer_sido = infer_sido

        # 시도 → [(시군구키워드, 시군구명, areaM, areaD)] , 긴 이름 우선 정렬
        self.by_sido: Dict[str, List[Tuple[str, str, int, int]]] = {}
        self.sido_code: Dict[str, int] = {}
        for _, r in code_df.iterrows():
            sido = nospace(r["시도"])
            sgg = nospace(r["시군구"])
            if not sido or not sgg:
                continue
            try:
                m = int(r["areaM"])
                d = int(r["areaD"])
            except (TypeError, ValueError):
                continue
            self.by_sido.setdefault(sido, []).append((sgg, sgg, m, d))
            self.sido_code.setdefault(sido, m)
        for k in self.by_sido:
            self.by_sido[k].sort(key=lambda t: -len(t[0]))

        # 시도 없는 주소용: 전국에서 딱 하나뿐인 시군구 이름만 역추적에 사용
        counter: Dict[str, List[str]] = {}
        for sido, items in self.by_sido.items():
            for kw, *_ in items:
                counter.setdefault(kw, []).append(sido)
        self.unique_sgg = {kw: v[0] for kw, v in counter.items() if len(set(v)) == 1}

        # 도로명 규칙
        self.road: Dict[str, List[Tuple[str, str]]] = {}
        if road_df is not None:
            for _, r in road_df.iterrows():
                sido = nospace(r.get("시도"))
                token = nospace(r.get("포함문구"))
                target = nospace(r.get("시군구"))
                if sido and token and target:
                    self.road.setdefault(sido, []).append((token, target))

        # 시도 별칭 (긴 것부터)
        self.alias: List[Tuple[str, str]] = []
        for std, names in SIDO_ALIASES.items():
            for n in names:
                self.alias.append((n, std))
        self.alias.sort(key=lambda t: -len(t[0]))

    # ── 시도 판정 : 주소 전체에서 가장 앞에 등장하는 시도명 채택 ──
    def detect_sido(self, clean: str) -> Optional[str]:
        best_pos, best_len, best = 10**9, 0, None
        for name, std in self.alias:
            pos = clean.find(name)
            if pos < 0:
                continue
            if pos < best_pos or (pos == best_pos and len(name) > best_len):
                best_pos, best_len, best = pos, len(name), std
        return best

    # ── 시군구 판정 : 앞에 등장하는 것 우선, 같으면 긴 이름 우선 ──
    def _find_sgg(self, clean: str, sido: str):
        best = None
        best_key = (10**9, 0)
        for kw, name, m, d in self.by_sido.get(sido, []):
            pos = clean.find(kw)
            if pos < 0:
                continue
            key = (pos, -len(kw))
            if key < best_key:
                best_key, best = key, (name, m, d)
        return best

    def resolve(self, address) -> dict:
        """→ dict(status, areaM, areaD, 시도, 시군구, 근거)"""
        clean = nospace(address)
        if not clean:
            return {"status": RESULT_NO_ADDR, "areaM": None, "areaD": None,
                    "시도": "", "시군구": "", "근거": ""}

        sido = self.detect_sido(clean)
        basis = "시도명"

        if sido is None and self.infer_sido:
            for kw in sorted(self.unique_sgg, key=len, reverse=True):
                if kw in clean:
                    sido = self.unique_sgg[kw]
                    basis = "시군구 역추적"
                    break

        if sido is None:
            return {"status": RESULT_NO_SIDO, "areaM": None, "areaD": None,
                    "시도": "", "시군구": "", "근거": ""}

        hit = self._find_sgg(clean, sido)

        # 직접 매칭 실패 시 도로명 규칙 적용
        if hit is None:
            for token, target in self.road.get(sido, []):
                if token in clean:
                    for kw, name, m, d in self.by_sido.get(sido, []):
                        if kw == target:
                            hit = (name, m, d)
                            basis = f"도로명 규칙({token})"
                            break
                if hit:
                    break

        if hit is None:
            return {"status": RESULT_NO_SGG,
                    "areaM": self.sido_code.get(sido), "areaD": None,
                    "시도": sido, "시군구": "", "근거": basis}

        name, m, d = hit
        return {"status": "OK", "areaM": m, "areaD": d,
                "시도": sido, "시군구": name, "근거": basis}


def to_code(value) -> Optional[int]:
    """'01' → 1, 공백/문자 → None (VBA Val 과 유사하되 빈값 구분)."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    s = str(value).strip()
    if s == "":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def verify(
    target_df: pd.DataFrame,
    addr_map: Dict[str, str],
    col_id: str,
    col_m: str,
    col_d: str,
    resolver: AreaResolver,
) -> pd.DataFrame:
    """검증 대상 DataFrame에 결과 열을 붙여 반환 (원본 열 보존)."""
    cache: Dict[str, dict] = {}
    out_result, out_m, out_d = [], [], []
    out_sido, out_sgg, out_basis, out_addr = [], [], [], []

    for _, row in target_df.iterrows():
        key = nospace(row[col_id])
        user_m = to_code(row[col_m])
        user_d = to_code(row[col_d])

        if key == "" or key not in addr_map:
            out_result.append(RESULT_NO_ID)
            out_m.append(None); out_d.append(None)
            out_sido.append(""); out_sgg.append(""); out_basis.append(""); out_addr.append("")
            continue

        address = addr_map[key]
        if address not in cache:
            cache[address] = resolver.resolve(address)
        info = cache[address]

        out_addr.append(address)
        out_sido.append(info["시도"])
        out_sgg.append(info["시군구"])
        out_basis.append(info["근거"])

        if info["status"] in (RESULT_NO_ADDR, RESULT_NO_SIDO):
            out_result.append(info["status"])
            out_m.append(info["areaM"]); out_d.append(info["areaD"])
        elif info["status"] == RESULT_NO_SGG:
            out_result.append(RESULT_NO_SGG)
            out_m.append(info["areaM"]); out_d.append(None)
        elif user_m is None and user_d is None:
            out_result.append(RESULT_NO_INPUT)
            out_m.append(info["areaM"]); out_d.append(info["areaD"])
        elif user_m == info["areaM"] and user_d == info["areaD"]:
            out_result.append(RESULT_MATCH)
            out_m.append(None); out_d.append(None)
        else:
            out_result.append(RESULT_DIFF)
            out_m.append(info["areaM"]); out_d.append(info["areaD"])

    res = target_df.copy()
    res["검증결과"] = out_result
    res["계산_areaM"] = out_m
    res["계산_areaD"] = out_d
    res["매칭_시도"] = out_sido
    res["매칭_시군구"] = out_sgg
    res["판정근거"] = out_basis
    res["참조주소"] = out_addr
    return res


def make_addr_map(df: pd.DataFrame, col_id: str, col_addr: str) -> Tuple[Dict[str, str], int]:
    """ID→주소 사전. 중복 ID는 첫 값 우선(VBA 동작 유지). 중복 건수도 반환."""
    mapping: Dict[str, str] = {}
    dup = 0
    for _, row in df.iterrows():
        key = nospace(row[col_id])
        if key == "":
            continue
        if key in mapping:
            dup += 1
            continue
        val = row[col_addr]
        mapping[key] = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val).strip()
    return mapping, dup


def guess_col(columns: List[str], candidates: List[str]) -> Optional[str]:
    """헤더 자동 탐색 (대소문자·공백 무시, 완전일치 → 부분일치)."""
    norm = {c: nospace(c).upper() for c in columns}
    for cand in candidates:
        target = nospace(cand).upper()
        for c in columns:
            if norm[c] == target:
                return c
    for cand in candidates:
        target = nospace(cand).upper()
        for c in columns:
            if target and target in norm[c]:
                return c
    return None


# ══════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False, max_entries=5)
def read_sheets(data: bytes, filename: str) -> Dict[str, pd.DataFrame]:
    if filename.lower().endswith(".csv"):
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                return {"CSV": pd.read_csv(io.BytesIO(data), dtype=object, encoding=enc)}
            except UnicodeDecodeError:
                continue
        return {"CSV": pd.read_csv(io.BytesIO(data), dtype=object, encoding="utf-8", errors="replace")}
    book = pd.read_excel(io.BytesIO(data), sheet_name=None, dtype=object)
    return {k: v for k, v in book.items() if k != "CodeDB"}


def to_excel(result: pd.DataFrame, summary: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="검증결과", index=False)
        summary.to_excel(writer, sheet_name="요약", index=False)
        bad = result[result["검증결과"].isin([RESULT_DIFF, RESULT_NO_SGG, RESULT_NO_SIDO])]
        if not bad.empty:
            bad.to_excel(writer, sheet_name="확인필요", index=False)

        from openpyxl.styles import Font, PatternFill
        ws = writer.sheets["검증결과"]
        head_fill = PatternFill("solid", fgColor="E2EFDA")
        cols = list(result.columns)
        for name in ["검증결과", "계산_areaM", "계산_areaD", "매칭_시도", "매칭_시군구", "판정근거", "참조주소"]:
            if name in cols:
                cell = ws.cell(row=1, column=cols.index(name) + 1)
                cell.fill = head_fill
                cell.font = Font(bold=True)
        idx = cols.index("검증결과") + 1
        red = Font(color="FF0000", bold=True)
        for i, val in enumerate(result["검증결과"], start=2):
            if val in (RESULT_DIFF, RESULT_NO_SIDO):
                ws.cell(row=i, column=idx).font = red
    return buf.getvalue()


def main() -> None:
    st.set_page_config(page_title="주소 지역코드 검증", page_icon="📮", layout="wide")
    if not check_password():
        st.stop()

    st.title("📮 주소 지역코드 검증")
    st.caption("주소에서 areaM(시도) · areaD(시군구) 코드를 계산해 입력값과 대조합니다.")

    # 세션에 기준표 보관
    if "code_df" not in st.session_state:
        st.session_state.code_df = build_default_code_df()
    if "road_df" not in st.session_state:
        st.session_state.road_df = build_default_road_df()

    tab_run, tab_code, tab_road = st.tabs(["검증", "지역코드 기준표", "도로명 보정 규칙"])

    # ── 기준표 탭 ──────────────────────────────────────────────
    with tab_code:
        st.markdown("행정구역이 바뀌면 여기서 직접 고치면 됩니다. 고친 내용은 바로 검증에 반영됩니다.")
        st.session_state.code_df = st.data_editor(
            st.session_state.code_df, num_rows="dynamic", use_container_width=True,
            height=420, key="editor_code",
            column_config={
                "시도": st.column_config.TextColumn(width="small"),
                "areaM": st.column_config.NumberColumn("areaM (시도코드)", width="small"),
                "시군구": st.column_config.TextColumn(width="small"),
                "areaD": st.column_config.NumberColumn("areaD (시군구코드)", width="small"),
            },
        )
        c1, c2 = st.columns([1, 4])
        if c1.button("기본값으로 되돌리기", key="reset_code"):
            st.session_state.code_df = build_default_code_df()
            st.rerun()
        c2.download_button(
            "기준표 CSV 내려받기",
            st.session_state.code_df.to_csv(index=False).encode("utf-8-sig"),
            "지역코드_기준표.csv", "text/csv",
        )

    # ── 도로명 규칙 탭 ─────────────────────────────────────────
    with tab_road:
        st.markdown(
            "시군구명이 주소에 안 나오는 경우(주로 세종시 신도심 도로명 주소)에 쓰는 보정표입니다. "
            "**주소에 「포함문구」가 들어 있으면 해당 시군구로 간주**합니다. 위에 있는 규칙이 먼저 적용됩니다."
        )
        st.session_state.road_df = st.data_editor(
            st.session_state.road_df, num_rows="dynamic", use_container_width=True,
            height=380, key="editor_road",
            column_config={
                "시도": st.column_config.TextColumn(width="small"),
                "포함문구": st.column_config.TextColumn("포함문구 (예: 밝은뜰)", width="medium"),
                "시군구": st.column_config.TextColumn("→ 시군구", width="small"),
            },
        )
        if st.button("기본값으로 되돌리기", key="reset_road"):
            st.session_state.road_df = build_default_road_df()
            st.rerun()

    # ── 검증 탭 ────────────────────────────────────────────────
    with tab_run:
        up = st.file_uploader("엑셀 또는 CSV 파일", type=["xlsx", "xlsm", "xls", "csv"])
        if up is None:
            st.info("파일을 올리면 열 선택 화면이 나타납니다.")
            return

        try:
            sheets = read_sheets(up.getvalue(), up.name)
        except Exception as exc:
            st.error(f"파일을 읽지 못했습니다: {exc}")
            return
        if not sheets:
            st.error("읽을 수 있는 시트가 없습니다.")
            return

        names = list(sheets)
        st.subheader("1. 데이터 위치")
        same = st.checkbox("주소와 검증대상이 같은 시트에 있음", value=True)

        cc = st.columns(2)
        addr_sheet = cc[0].selectbox("주소 원본 시트", names, key="sheet_addr")
        tgt_sheet = addr_sheet if same else cc[1].selectbox(
            "검증 대상 시트", names, index=min(1, len(names) - 1), key="sheet_tgt")

        df_addr = sheets[addr_sheet]
        df_tgt = sheets[tgt_sheet]
        addr_cols = list(df_addr.columns)
        tgt_cols = list(df_tgt.columns)

        st.subheader("2. 열 지정")
        left, right = st.columns(2)
        with left:
            st.markdown("**주소 원본**")
            c_sid = st.selectbox("ID 열", addr_cols, key="c_sid",
                                 index=_idx(addr_cols, guess_col(addr_cols, ["ID", "id", "panel_id"])))
            c_addr = st.selectbox("주소 열", addr_cols, key="c_addr",
                                  index=_idx(addr_cols, guess_col(addr_cols, ["주소", "Address", "addr"])))
        with right:
            st.markdown("**검증 대상**")
            c_tid = st.selectbox("ID 열", tgt_cols, key="c_tid",
                                 index=_idx(tgt_cols, guess_col(tgt_cols, ["ID", "id", "panel_id"])))
            c_m = st.selectbox("areaM 열", tgt_cols, key="c_m",
                               index=_idx(tgt_cols, guess_col(tgt_cols, ["areaM", "area_M"])))
            c_d = st.selectbox("areaD 열", tgt_cols, key="c_d",
                               index=_idx(tgt_cols, guess_col(tgt_cols, ["areaD", "areaQ", "area_D"])))

        infer = st.checkbox("시도명이 없는 주소는 시군구 이름으로 시도를 추정", value=True,
                            help="예: 「성남시 분당구…」처럼 시도가 빠진 주소도 전국에서 이름이 하나뿐이면 찾아냅니다.")

        if not st.button("검증 실행", type="primary", use_container_width=True):
            return

        with st.spinner("주소를 해석하는 중…"):
            resolver = AreaResolver(st.session_state.code_df, st.session_state.road_df, infer_sido=infer)
            addr_map, dup = make_addr_map(df_addr, c_sid, c_addr)
            result = verify(df_tgt, addr_map, c_tid, c_m, c_d, resolver)

        st.subheader("3. 결과")
        if dup:
            st.warning(f"주소 원본에 중복 ID가 {dup:,}건 있습니다. 첫 번째 주소를 사용했습니다.")

        counts = result["검증결과"].value_counts()
        order = [RESULT_MATCH, RESULT_DIFF, RESULT_NO_SGG, RESULT_NO_SIDO,
                 RESULT_NO_ADDR, RESULT_NO_ID, RESULT_NO_INPUT]
        shown = [k for k in order if k in counts]
        cols = st.columns(max(len(shown), 1))
        for col, key in zip(cols, shown):
            col.metric(key, f"{int(counts[key]):,}")

        summary = pd.DataFrame({"검증결과": shown, "건수": [int(counts[k]) for k in shown]})

        need = result[result["검증결과"].isin([RESULT_DIFF, RESULT_NO_SGG, RESULT_NO_SIDO])]
        if need.empty:
            st.success("확인이 필요한 건이 없습니다.")
        else:
            st.markdown(f"**확인 필요 {len(need):,}건**")
            preview = [c_tid, c_m, c_d, "계산_areaM", "계산_areaD",
                       "검증결과", "매칭_시도", "매칭_시군구", "판정근거", "참조주소"]
            st.dataframe(need[[c for c in preview if c in need.columns]],
                         use_container_width=True, height=380)

        st.download_button(
            "결과 엑셀 내려받기",
            to_excel(result, summary),
            f"주소검증_{up.name.rsplit('.', 1)[0]}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


def _idx(options: List[str], value: Optional[str]) -> int:
    return options.index(value) if value in options else 0


if __name__ == "__main__":
    main()
