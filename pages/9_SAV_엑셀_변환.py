# -*- coding: utf-8 -*-
"""
SAV → 엑셀 변환 (v1.1)

SPSS .sav 파일을 업로드하면 여러 시트로 구성된 엑셀 파일을 내려받습니다.
  · Raw        : 숫자 코드 그대로
  · Label      : 값 레이블로 치환 (레이블 없는 변수는 원래 값 유지)
  · Code       : 문자형(주관식) 변수만 모아서 — 문자형 변수가 있을 때만 생성
  · 변수 가이드 : 변수명 + 변수 설명

이 페이지는 utils.py 없이도 단독으로 동작합니다.
"""

import io
import os
import tempfile

import numpy as np
import pandas as pd
import pyreadstat
import streamlit as st
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# 비밀번호 (utils.py 있으면 사용, 없으면 통과)
# ──────────────────────────────────────────────────────────────
try:
    from utils import check_password  # type: ignore
except Exception:  # pragma: no cover
    def check_password() -> bool:
        return True


st.set_page_config(page_title="SAV → 엑셀 변환", page_icon="📗", layout="wide")

if not check_password():
    st.stop()

st.title("📗 SAV → 엑셀 변환")
st.caption("SPSS .sav 파일을 Raw / Label / 변수 가이드 3개 시트의 엑셀로 바꿔 드립니다.")


# ──────────────────────────────────────────────────────────────
# 변환 로직
# ──────────────────────────────────────────────────────────────
def _clean(v):
    """NaN은 빈칸으로, 소수점 없는 실수는 정수로."""
    if v is None:
        return None
    if isinstance(v, float):
        if np.isnan(v):
            return None
        if float(v).is_integer():
            return int(v)
    return v


@st.cache_data(show_spinner=False, max_entries=5)
def read_sav_bytes(data: bytes, filename: str):
    """업로드된 바이트를 임시파일로 떨어뜨려 pyreadstat으로 읽는다."""
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".sav", delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        df, meta = pyreadstat.read_sav(tmp_path, apply_value_formats=False)
        return (
            df,
            dict(meta.column_names_to_labels or {}),
            dict(meta.variable_value_labels or {}),
            dict(getattr(meta, "readstat_variable_types", {}) or {}),
        )
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def build_raw(df: pd.DataFrame) -> pd.DataFrame:
    return df.map(_clean)


def build_label(df: pd.DataFrame, value_labels: dict) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if c in value_labels:
            m = {k: str(v).strip() for k, v in value_labels[c].items()}
            out[c] = out[c].map(lambda x, m=m: m.get(x, _clean(x)))
        else:
            out[c] = out[c].map(_clean)
    return out


def find_id_col(df: pd.DataFrame) -> str:
    """응답자를 되짚을 키 열. NO / id 계열을 우선 찾고 없으면 첫 열."""
    for cand in ("NO", "No", "no", "ID", "Id", "id", "panel_id"):
        if cand in df.columns:
            return cand
    return df.columns[0]


def find_text_cols(df: pd.DataFrame, var_types: dict) -> list:
    """SAV에서 문자형으로 선언된 변수 목록."""
    return [c for c in df.columns if str(var_types.get(c, "")).lower() == "string"]


def build_code(df: pd.DataFrame, text_cols: list, id_col: str) -> pd.DataFrame:
    """주관식 응답만 모은 시트. 행 순서는 Raw/Label과 동일하게 유지."""
    cols = ([id_col] if id_col not in text_cols else []) + text_cols
    out = df[cols].copy()
    for c in out.columns:
        if c in text_cols:
            out[c] = out[c].map(
                lambda x: None
                if x is None or (isinstance(x, float) and np.isnan(x)) or str(x).strip() == ""
                else str(x).strip()
            )
        else:
            out[c] = out[c].map(_clean)
    return out


def build_guide(df: pd.DataFrame, col_labels: dict) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "변수명": list(df.columns),
            "변수 내용": [str(col_labels.get(c) or "").strip() for c in df.columns],
        }
    )


def to_excel(sheets: dict) -> bytes:
    """{시트명: DataFrame} → 엑셀 바이트. 헤더 굵게 + 첫 행 고정."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"
            if name == "변수 가이드":
                ws.column_dimensions["A"].width = 18
                ws.column_dimensions["B"].width = 100
            elif name == "Code":
                ws.column_dimensions["A"].width = 10
                for j in range(2, min(ws.max_column, 200) + 1):
                    ws.column_dimensions[get_column_letter(j)].width = 45
            else:
                for j in range(1, min(ws.max_column, 200) + 1):
                    ws.column_dimensions[get_column_letter(j)].width = 14
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────────────────────
up = st.file_uploader("SAV 파일을 올려주세요", type=["sav"])

if not up:
    st.info("SPSS .sav 파일을 올리면 미리보기와 다운로드 버튼이 나타납니다.")
    st.stop()

try:
    df, col_labels, value_labels, var_types = read_sav_bytes(up.getvalue(), up.name)
except Exception as e:
    st.error(f"파일을 읽지 못했습니다: {e}")
    st.stop()

n_rows, n_cols = df.shape
n_labeled = sum(1 for c in df.columns if c in value_labels)
text_cols = find_text_cols(df, var_types)
id_col = find_id_col(df)

c1, c2, c3, c4 = st.columns(4)
c1.metric("응답자 수", f"{n_rows:,}")
c2.metric("변수 수", f"{n_cols:,}")
c3.metric("값 레이블이 있는 변수", f"{n_labeled:,}")
c4.metric("문자형 변수", f"{len(text_cols):,}")

if n_rows > 10_000:
    st.warning(
        f"행이 {n_rows:,}개입니다. 1만 행이 넘으면 변환이 느리거나 "
        "메모리 한도에 걸릴 수 있습니다."
    )

st.divider()

st.subheader("담을 시트 고르기")
s1, s2, s3, s4 = st.columns(4)
want_raw = s1.checkbox("Raw (숫자 코드)", value=True)
want_label = s2.checkbox("Label (값 레이블)", value=True)
want_code = s3.checkbox(
    "Code (문자형 변수)",
    value=bool(text_cols),
    disabled=not text_cols,
    help="SAV에 문자형으로 저장된 변수만 모읍니다." if text_cols else "이 파일에는 문자형 변수가 없습니다.",
)
want_guide = s4.checkbox("변수 가이드", value=True)

if not text_cols:
    st.info(
        "이 파일에는 문자형 변수가 없어 Code 시트를 만들지 않습니다. "
        "주관식 응답이 있어야 하는데 비어 있다면, SAV로 내보낼 때 해당 문항이 "
        "숫자형으로 선언돼 내용이 빠졌을 수 있습니다."
    )

if not (want_raw or want_label or want_code or want_guide):
    st.warning("시트를 하나 이상 선택해주세요.")
    st.stop()

# ── 시트 구성 (Raw → Label → Code → 변수 가이드) ──
sheets = {}
if want_raw:
    sheets["Raw"] = build_raw(df)
if want_label:
    sheets["Label"] = build_label(df, value_labels)
if want_code and text_cols:
    sheets["Code"] = build_code(df, text_cols, id_col)
if want_guide:
    sheets["변수 가이드"] = build_guide(df, col_labels)

st.subheader("미리보기")
tabs = st.tabs(list(sheets.keys()))
for tab, (name, frame) in zip(tabs, sheets.items()):
    with tab:
        st.dataframe(
            frame.head(20).astype(str).replace("None", ""),
            use_container_width=True,
            hide_index=True,
        )
        if len(frame) > 20:
            st.caption(f"위 20행만 표시 · 전체 {len(frame):,}행")

st.divider()

if st.button("엑셀로 변환하기", type="primary", use_container_width=True):
    with st.spinner("엑셀 파일을 만드는 중입니다…"):
        try:
            xlsx_bytes = to_excel(sheets)
        except Exception as e:
            st.error(f"엑셀 생성에 실패했습니다: {e}")
            st.stop()
    st.session_state["sav2xlsx_bytes"] = xlsx_bytes
    st.session_state["sav2xlsx_name"] = os.path.splitext(up.name)[0] + ".xlsx"

if st.session_state.get("sav2xlsx_bytes"):
    st.success("변환이 끝났습니다.")
    st.download_button(
        "엑셀 파일 내려받기",
        data=st.session_state["sav2xlsx_bytes"],
        file_name=st.session_state["sav2xlsx_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
