# -*- coding: utf-8 -*-
"""
SAV → 엑셀 변환 (v2.4)

SPSS .sav 파일을 업로드하면 여러 시트로 구성된 엑셀 파일을 내려받습니다.
  · Raw        : 숫자 코드 그대로
  · Label      : 값 레이블로 치환 (레이블 없는 변수는 원래 값 유지)
  · Open       : 키 변수(NO, id) + 문자형(주관식) 변수
                 문자형 변수가 없어도 키 변수만으로 만든다
  · Code       : DP 코드북 형식 (변수마다 문항 + 코드값/보기 블록)
  · 변수 가이드 : 변수명 + 변수 설명

'엑셀 값 반영하기' 탭에서는 반대로, 코딩·수정을 마친 엑셀을 올려
ID 로 짝을 맞춰 SAV 값을 덮어씁니다. 결과는 SAV 로도, 위 시트 구성의
엑셀로도 바로 받을 수 있습니다. (SAV 를 다시 올릴 필요 없음)

이 페이지는 utils.py 없이도 단독으로 동작합니다.
"""

import io
import os
import re
import tempfile

import numpy as np
import pandas as pd
import pyreadstat
import streamlit as st
from openpyxl.styles import Font, PatternFill

# ──────────────────────────────────────────────────────────────
# 비밀번호 (utils.py 있으면 사용, 없으면 통과)
# ──────────────────────────────────────────────────────────────
try:
    from utils import check_password  # type: ignore
except Exception:  # pragma: no cover
    def check_password() -> bool:
        return True


st.set_page_config(page_title="SAV → 엑셀 변환", page_icon="📗", layout="wide")

if not check_password():
    st.stop()

st.title("📗 SAV → 엑셀 변환")
st.caption("SPSS .sav 파일을 Raw / Label / Open / Code / 변수 가이드 시트의 엑셀로 바꿔 드립니다.")


# ──────────────────────────────────────────────────────────────
# 변환 로직
# ──────────────────────────────────────────────────────────────
def _clean(v):
    """NaN은 빈칸으로, 소수점 없는 실수는 정수로."""
    if v is None:
        return None
    if isinstance(v, float):
        if np.isnan(v):
            return None
        if float(v).is_integer():
            return int(v)
    return v


@st.cache_data(show_spinner=False, max_entries=5)
def read_sav_bytes(data: bytes, filename: str):
    """업로드된 바이트를 임시파일로 떨어뜨려 pyreadstat으로 읽는다."""
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".sav", delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name
        df, meta = pyreadstat.read_sav(tmp_path, apply_value_formats=False)
        return (
            df,
            dict(meta.column_names_to_labels or {}),
            dict(meta.variable_value_labels or {}),
            dict(getattr(meta, "readstat_variable_types", {}) or {}),
        )
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def build_raw(df: pd.DataFrame) -> pd.DataFrame:
    return df.map(_clean)


def build_label(df: pd.DataFrame, value_labels: dict) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if c in value_labels:
            m = {k: str(v).strip() for k, v in value_labels[c].items()}
            out[c] = out[c].map(lambda x, m=m: m.get(x, _clean(x)))
        else:
            out[c] = out[c].map(_clean)
    return out


def find_key_cols(df: pd.DataFrame) -> list:
    """응답자를 되짚을 키 열. NO / id 를 원래 순서대로 모으고, 없으면 첫 열."""
    keys = [c for c in df.columns if str(c).strip().lower() in ("no", "id")]
    return keys if keys else [df.columns[0]]


def find_text_cols(df: pd.DataFrame, var_types: dict) -> list:
    """SAV에서 문자형으로 선언된 변수 목록."""
    return [c for c in df.columns if str(var_types.get(c, "")).lower() == "string"]


def build_open(df: pd.DataFrame, text_cols: list, key_cols: list) -> pd.DataFrame:
    """키 변수 + 주관식 응답. 행 순서는 Raw/Label과 동일하게 유지.

    문자형 변수가 없으면 키 변수만 담긴 시트가 된다. 주관식 코딩을 할 때
    이 시트에 열을 직접 추가해 쓸 수 있도록 빈 채로라도 만들어 둔다.
    """
    cols = list(key_cols) + [c for c in text_cols if c not in key_cols]
    out = df[cols].copy()
    for c in out.columns:
        if c in text_cols:
            out[c] = out[c].map(
                lambda x: None
                if x is None or (isinstance(x, float) and np.isnan(x)) or str(x).strip() == ""
                else str(x).strip()
            )
        else:
            out[c] = out[c].map(_clean)
    return out


# 머리글에 칠할 색. 엑셀 기본 팔레트의 연한 색들로, 검은 글씨가 잘 보인다.
# 순서가 화면에 나오는 순서다. 기본값은 맨 앞 항목.
HEAD_COLORS = {
    "주황": "#FCE4D6",
    "파랑": "#D9E1F2",
    "초록": "#E2EFDA",
    "노랑": "#FFF2CC",
    "회색": "#EDEDED",
    "자주": "#E4DFEC",
}
DEFAULT_HEAD_COLOR = next(iter(HEAD_COLORS.values()))

# 머리글 띠를 몇 열까지 칠할지. 모든 시트에 같이 적용된다.
# 데이터가 있는 데까지만 칠하면 오른쪽이 끊겨 보이므로,
# 최소 이 수까지 칠하고 데이터가 더 많으면 그 끝에서 여유분만큼 더 칠한다.
#
# openpyxl 은 RowDimension 에 fill 을 걸어도 스타일 번호만 붙이고 색은
# 넣지 않으므로(fillId 가 0 으로 남는다) 셀을 하나씩 칠해야 한다.
HEAD_FILL_COLS = 40
HEAD_FILL_MARGIN = 8

# 엑셀 열 상한. 이걸 넘기면 엑셀이 파일을 못 여는데 openpyxl 은 막지 않는다.
EXCEL_MAX_COLS = 16384


def _fill(color: str):
    """'#FCE4D6' 또는 'FCE4D6' → PatternFill. 빈 값이면 None."""
    if not color:
        return None
    return PatternFill("solid", fgColor=str(color).lstrip("#").upper())


def _code_str(v) -> str:
    """값 라벨의 코드를 표시용 문자열로. 1.0 -> '1'"""
    if isinstance(v, float) and float(v).is_integer():
        return str(int(v))
    return str(v)


def _option_text(code_str: str, label) -> str:
    """보기 문구에서 앞에 붙은 코드값을 뗀다.

    SPSS 값 라벨은 '  1) 남성' 처럼 코드가 앞에 붙어 저장되는 경우가 많다.
    코드값은 A열에 따로 들어가므로 중복이라 뗀다.
    떼고 나면 아무것도 안 남는 경우(척도 중간값처럼 '  5)' 만 있는 경우)는
    원문을 그대로 둔다. 빈 칸으로 보이면 누락처럼 읽히기 때문이다.
    """
    s = str(label).strip()
    m = re.match(r"^" + re.escape(code_str) + r"\s*[)\.]\s*", s)
    if m and s[m.end():].strip():
        return s[m.end():].strip()
    return s


def build_codebook(df: pd.DataFrame, col_labels: dict,
                   value_labels: dict, key_cols: list) -> pd.DataFrame:
    """DP 코드북 형식. 변수마다 블록 하나.

        q1        SQ1. 귀하의 성별은 무엇입니까?
        코드값     보기
        1         남성
        2         여성
        (빈 줄)
        (빈 줄)

    값 라벨이 없는 변수도 머리글까지는 넣고 코드 부분만 비운다.
    키 변수(no, id)는 문항이 아니므로 제외한다.
    """
    rows = []
    for c in df.columns:
        if c in key_cols:
            continue
        rows.append([str(c), str(col_labels.get(c) or "").strip()])
        rows.append(["코드값", "보기"])
        for code, lab in sorted((value_labels.get(c) or {}).items()):
            cs = _code_str(code)
            rows.append([cs, _option_text(cs, lab)])
        rows.append([None, None])
        rows.append([None, None])
    return pd.DataFrame(rows, columns=["변수", "내용"])


def build_guide(df: pd.DataFrame, col_labels: dict) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "변수명": list(df.columns),
            "변수 내용": [str(col_labels.get(c) or "").strip() for c in df.columns],
        }
    )


@st.cache_data(show_spinner=False, max_entries=5)
def list_sheets(data: bytes, filename: str):
    """엑셀 시트 이름 목록. CSV 면 None."""
    if os.path.splitext(filename)[1].lower() in (".csv", ".txt"):
        return None
    return pd.ExcelFile(io.BytesIO(data)).sheet_names


@st.cache_data(show_spinner=False, max_entries=5)
def read_table_bytes(data: bytes, filename: str, sheet=0) -> pd.DataFrame:
    """엑셀/CSV 업로드를 DataFrame 으로. 값은 손대지 않고 그대로 읽는다.

    sheet: 엑셀 시트 이름. CSV 면 무시한다.
    """
    ext = os.path.splitext(filename)[1].lower()
    bio = io.BytesIO(data)
    if ext in (".csv", ".txt"):
        for enc in ("utf-8-sig", "cp949", "utf-8"):
            try:
                bio.seek(0)
                return pd.read_csv(bio, dtype=object, encoding=enc)
            except UnicodeDecodeError:
                continue
        bio.seek(0)
        return pd.read_csv(bio, dtype=object, encoding="latin1")
    return pd.read_excel(bio, sheet_name=sheet, dtype=object)


def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and np.isnan(v):
        return True
    return str(v).strip() == ""


def _as_number(v):
    """숫자로 읽히면 float, 아니면 None."""
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def _to_text(v) -> str:
    """문자형으로 바꿀 때의 표시. 1.0 -> '1'"""
    if _is_blank(v):
        return ""
    if isinstance(v, float) and float(v).is_integer():
        return str(int(v))
    return str(v).strip()


SKIP_LABEL = "(넘기기)"


def guess_mapping(df: pd.DataFrame, patch: pd.DataFrame,
                  sav_key: str, patch_key: str) -> pd.DataFrame:
    """엑셀 열마다 어느 SAV 변수에 넣을지 짐작한 표.

    이름이 같으면(대소문자 무시) 그 변수를, 없으면 '(넘기기)'.
    화면에서 사람이 고쳐 쓸 수 있게 값 예시와 채워진 칸 수를 함께 담는다.
    """
    sav_by_lower = {str(c).lower(): c for c in df.columns}
    rows = []
    for pc in patch.columns:
        if str(pc) == str(patch_key):
            continue
        vals = [v for v in patch[pc] if not _is_blank(v)]
        hit = sav_by_lower.get(str(pc).lower())
        if hit is None or hit == sav_key:
            hit = SKIP_LABEL
        rows.append({
            "엑셀 열": str(pc),
            "값 예시": " / ".join(_to_text(v) for v in vals[:3]),
            "채워진 칸": len(vals),
            "SAV 변수": hit,
        })
    return pd.DataFrame(rows)


def apply_patch(df: pd.DataFrame, value_labels: dict, var_types: dict,
                patch: pd.DataFrame, sav_key: str, patch_key: str,
                mapping: dict) -> tuple:
    """엑셀 값을 SAV 데이터에 덮어쓴다.

    · ID 로 행을 짝지은 뒤, mapping 에 적힌 대로 열을 덮어쓴다.
      mapping: {엑셀 열 이름: SAV 변수 이름}. '(넘기기)' 는 건너뛴다.
    · 엑셀의 빈칸은 건드리지 않는다. 기존 값이 그대로 남는다.
    · 숫자 변수에 문자 값이 하나라도 들어오면 그 변수 전체를 문자형으로
      바꾼다. SPSS 는 한 변수에 숫자와 문자를 섞을 수 없기 때문이다.
      이때 값 라벨은 숫자 코드에 붙는 것이라 쓸 수 없게 되므로 버린다.

    반환: (새 df, 새 value_labels, 리포트 dict)
    """
    out = df.copy()
    new_labels = {k: dict(v) for k, v in value_labels.items()}

    pairs, only_in_patch = [], []
    for pc in patch.columns:
        if str(pc) == str(patch_key):
            continue
        sc = mapping.get(str(pc), SKIP_LABEL)
        if sc in (SKIP_LABEL, None, "") or sc not in out.columns:
            only_in_patch.append(str(pc))
        else:
            pairs.append((sc, pc))

    # ── ID 짝 맞추기 ──
    key_map = {_to_text(v): i for i, v in enumerate(out[sav_key])}
    changes, to_text_cols, unmatched_ids = {}, [], []
    matched_ids = set()

    for _, prow in patch.iterrows():
        kid = _to_text(prow[patch_key])
        if kid == "":
            continue
        idx = key_map.get(kid)
        if idx is None:
            unmatched_ids.append(kid)
            continue
        matched_ids.add(kid)
        for sc, pc in pairs:
            v = prow[pc]
            if _is_blank(v):
                continue                      # 빈칸은 건드리지 않는다
            changes.setdefault(sc, []).append((idx, v))

    # ── 열마다 반영 ──
    for sc, items in changes.items():
        numeric_col = str(var_types.get(sc, "")).lower() != "string"
        has_text = any(_as_number(v) is None for _, v in items)

        if numeric_col and has_text:
            # 열 전체를 문자형으로 바꾼다
            out[sc] = out[sc].map(_to_text)
            to_text_cols.append(sc)
            new_labels.pop(sc, None)
            for idx, v in items:
                out.iat[idx, out.columns.get_loc(sc)] = _to_text(v)
        elif numeric_col:
            col = out.columns.get_loc(sc)
            for idx, v in items:
                out.iat[idx, col] = _as_number(v)
        else:
            out[sc] = out[sc].map(_to_text)
            col = out.columns.get_loc(sc)
            for idx, v in items:
                out.iat[idx, col] = _to_text(v)

    report = {
        "pairs": [sc for sc, _ in pairs],
        "only_in_patch": only_in_patch,
        "changed": {k: len(v) for k, v in changes.items()},
        "to_text": to_text_cols,
        "matched": len(matched_ids),
        "unmatched_ids": unmatched_ids,
    }
    return out, new_labels, report


def write_sav(df: pd.DataFrame, col_labels: dict, value_labels: dict) -> bytes:
    """DataFrame 을 .sav 바이트로."""
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "out.sav")
        pyreadstat.write_sav(
            df, path,
            column_labels=[col_labels.get(c) or "" for c in df.columns],
            variable_value_labels={k: v for k, v in value_labels.items()
                                   if k in df.columns and v} or None,
        )
        with open(path, "rb") as f:
            return f.read()


def to_excel(sheets: dict, head_color: str = DEFAULT_HEAD_COLOR) -> bytes:
    """{시트명: DataFrame} → 엑셀 바이트.

    head_color: 머리글에 칠할 색. 빈 문자열이면 칠하지 않는다.
                Code 시트는 변수명·문항 줄, 나머지는 첫 행에 적용된다.
    """
    head_fill = _fill(head_color)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            # Code 는 변수 블록이 이어지는 형태라 표 머리글이 없다.
            is_code = name == "Code"
            frame.to_excel(writer, sheet_name=name, index=False,
                           header=not is_code)
            ws = writer.sheets[name]
            wide = max(ws.max_column + HEAD_FILL_MARGIN, HEAD_FILL_COLS)
            wide = min(wide, EXCEL_MAX_COLS)

            if is_code:
                ws.column_dimensions["A"].width = 14
                ws.column_dimensions["B"].width = 90
                # 변수명·문항 줄은 색을 채우고, '코드값' 줄은 굵게만.
                # 블록이 이어지는 시트라 눈으로 경계를 찾을 수 있어야 한다.
                #
                # 행 전체를 칠하려면 RowDimension 에 서식을 건다.
                # customFormat="1" 로 저장되는데, 엑셀에서 행 머리글을 눌러
                # 색을 칠했을 때와 같은 방식이라 오른쪽 끝까지 칠해진다.
                # 값이 든 칸(A·B)에도 따로 지정한다. 일부 뷰어가
                # 행 서식을 무시하고 셀 서식만 보기 때문이다.
                for r in range(1, ws.max_row + 1):
                    if ws.cell(row=r, column=1).value != "코드값":
                        continue
                    for c in (1, 2):
                        ws.cell(row=r, column=c).font = Font(bold=True)
                    if r > 1:
                        for c in range(1, wide + 1):
                            head = ws.cell(row=r - 1, column=c)
                            if c <= 2:
                                head.font = Font(bold=True)
                            if head_fill:
                                head.fill = head_fill
                continue

            for c in range(1, wide + 1):
                cell = ws.cell(row=1, column=c)
                if c <= ws.max_column:      # 굵게는 이름이 있는 칸만
                    cell.font = Font(bold=True)
                if head_fill:
                    cell.fill = head_fill
            ws.freeze_panes = "A2"
            if name == "변수 가이드":
                ws.column_dimensions["A"].width = 18
                ws.column_dimensions["B"].width = 100
            # Raw / Label / Open 은 너비를 지정하지 않는다.
            # 엑셀 기본 너비로 두면 사용자가 전체 선택 후 한 번에 조절할 수 있다.
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────
# UI
#   흐름은 위에서 아래로 한 줄이다.
#     SAV 올리기 → (선택) 엑셀 값 반영 → 시트 고르기 → 내려받기
#   반영을 켜면 그 아래 단계가 모두 '반영된 데이터' 기준으로 돌아간다.
# ──────────────────────────────────────────────────────────────
up = st.file_uploader("SAV 파일을 올려주세요", type=["sav"])

if not up:
    st.info("SPSS .sav 파일을 올리면 다음 단계가 나타납니다.")
    st.stop()

try:
    df, col_labels, value_labels, var_types = read_sav_bytes(up.getvalue(), up.name)
except Exception as e:
    st.error(f"파일을 읽지 못했습니다: {e}")
    st.stop()

c1, c2, c3, c4 = st.columns(4)
c1.metric("응답자 수", f"{len(df):,}")
c2.metric("변수 수", f"{len(df.columns):,}")
c3.metric("값 레이블이 있는 변수", f"{sum(1 for c in df.columns if c in value_labels):,}")
c4.metric("문자형 변수", f"{len(find_text_cols(df, var_types)):,}")

if len(df) > 10_000:
    st.warning(
        f"행이 {len(df):,}개입니다. 1만 행이 넘으면 변환이 느리거나 "
        "메모리 한도에 걸릴 수 있습니다."
    )

st.divider()

# ══════════════════════════════════════════════════════════════
#  1. 엑셀 값 반영 (선택)
# ══════════════════════════════════════════════════════════════
st.subheader("1. 엑셀 값 반영 " + "(선택)")

do_patch = st.checkbox(
    "엑셀 파일의 값으로 덮어쓰기",
    value=False,
    help="코딩·수정을 마친 엑셀을 올리면 ID 로 짝을 맞춰 같은 이름의 변수를 "
         "덮어씁니다. 엑셀의 빈칸은 건드리지 않습니다.",
)

# 아래 단계에서 쓸 '작업 데이터'. 반영을 안 하면 원본 그대로다.
work_df, work_labels, work_types = df, value_labels, var_types
patched, rep = False, None

if do_patch:
    pf = st.file_uploader("수정 값이 든 엑셀 또는 CSV",
                          type=["xlsx", "xls", "csv"], key="SX_patch")
    if not pf:
        st.info("엑셀 파일을 올리면 짝을 맞춰 보여드립니다.")
        st.stop()

    # ── 시트 고르기 (엑셀이 여러 시트일 때) ──
    try:
        sheet_names = list_sheets(pf.getvalue(), pf.name)
    except Exception as e:
        st.error(
            f"시트 목록을 읽지 못했습니다: {e}\n\n"
            ".xls 파일이라면 requirements.txt 에 xlrd 가 있는지 확인해 주세요."
        )
        st.stop()

    sheet = 0
    if sheet_names:
        if len(sheet_names) == 1:
            sheet = sheet_names[0]
            st.caption(f"시트: {sheet}")
        else:
            HINTS = ("코딩", "수정", "반영", "결과", "data", "raw")
            guess = next((i for i, s in enumerate(sheet_names)
                          if any(h in str(s).lower() for h in HINTS)), 0)
            sheet = st.selectbox(
                f"시트 고르기 (총 {len(sheet_names)}개)",
                sheet_names, index=guess,
                help="값이 든 시트를 고르세요. 첫 시트가 표지인 경우가 많습니다.",
            )

    try:
        patch = read_table_bytes(pf.getvalue(), pf.name, sheet)
    except Exception as e:
        st.error(f"파일을 읽지 못했습니다: {e}")
        st.stop()

    if patch.empty or not len(patch.columns):
        st.warning("고른 시트가 비어 있습니다. 다른 시트를 골라 주세요.")
        st.stop()

    st.write(f"올리신 파일: {len(patch):,}행 × {len(patch.columns)}열")
    with st.expander("고른 시트 미리보기"):
        st.dataframe(
            patch.head(10).astype(str).replace("None", "").replace("nan", ""),
            hide_index=True, use_container_width=True,
        )

    kc = find_key_cols(df)
    k1, k2 = st.columns(2)
    with k1:
        sav_key = st.selectbox(
            "SAV 의 ID 변수", list(df.columns),
            index=list(df.columns).index(kc[0]) if kc else 0,
        )
    with k2:
        pcols = [str(c) for c in patch.columns]
        guess = next((i for i, c in enumerate(pcols)
                      if c.lower() == str(sav_key).lower()), 0)
        patch_key = st.selectbox("엑셀의 ID 열", pcols, index=guess)

    # ── 변수 짝 맞추기 (표에서 직접 고칠 수 있다) ──
    st.markdown("**변수 짝 맞추기**")
    st.caption(
        "이름이 같은 변수는 미리 채워 뒀습니다. 'SAV 변수' 칸을 눌러 바꾸거나, "
        "넣지 않을 열은 " + SKIP_LABEL + " 로 두세요."
    )

    guess_df = guess_mapping(df, patch, sav_key, patch_key)
    if guess_df.empty:
        st.warning("ID 열 말고는 열이 없습니다. 다른 시트를 골라 주세요.")
        st.stop()

    edited = st.data_editor(
        guess_df,
        hide_index=True,
        use_container_width=True,
        # 파일·시트·ID 가 바뀌면 표를 새로 그린다
        key=f"SX_map_{pf.name}_{sheet}_{sav_key}_{patch_key}",
        column_config={
            "엑셀 열": st.column_config.TextColumn("엑셀 열", disabled=True),
            "값 예시": st.column_config.TextColumn("값 예시", disabled=True,
                                                width="medium"),
            "채워진 칸": st.column_config.NumberColumn("채워진 칸", disabled=True,
                                                   width="small"),
            "SAV 변수": st.column_config.SelectboxColumn(
                "SAV 변수", options=[SKIP_LABEL] + list(df.columns),
                required=True,
            ),
        },
    )

    mapping = dict(zip(edited["엑셀 열"], edited["SAV 변수"]))

    # 같은 SAV 변수에 두 열을 넣으면 뒤엣것이 앞엣것을 덮는다
    used = [v for v in mapping.values() if v != SKIP_LABEL]
    dups = sorted({v for v in used if used.count(v) > 1})
    if dups:
        st.warning(
            "같은 SAV 변수에 엑셀 열이 둘 이상 연결됐습니다. "
            "표 아래쪽 열이 위쪽을 덮어씁니다 — " + ", ".join(dups)
        )

    with st.spinner("맞춰 보는 중입니다…"):
        new_df, new_labels, rep = apply_patch(
            df, value_labels, var_types, patch, sav_key, patch_key, mapping)

    m1, m2, m3 = st.columns(3)
    m1.metric("짝이 맞은 응답자", f"{rep['matched']:,}")
    m2.metric("덮어쓸 변수", f"{len(rep['changed']):,}")
    m3.metric("바뀌는 셀", f"{sum(rep['changed'].values()):,}")

    if rep["unmatched_ids"]:
        st.warning(
            f"SAV 에 없는 ID {len(rep['unmatched_ids'])}개는 넘겼습니다 — "
            + ", ".join(rep["unmatched_ids"][:10])
            + (" …" if len(rep["unmatched_ids"]) > 10 else "")
        )

    if rep["only_in_patch"]:
        st.info(
            f"연결하지 않아 넘긴 열 {len(rep['only_in_patch'])}개 — "
            + ", ".join(rep["only_in_patch"][:10])
            + (" …" if len(rep["only_in_patch"]) > 10 else "")
        )

    if rep["to_text"]:
        st.warning(
            "문자 값이 섞여 아래 변수는 **문자형으로 바뀝니다**. "
            "SPSS 는 한 변수에 숫자와 문자를 섞을 수 없어서, "
            "다른 응답자의 숫자도 글자가 되고 값 라벨은 버려집니다.\n\n"
            + ", ".join(rep["to_text"])
        )

    if not rep["changed"]:
        st.warning(
            "덮어쓸 값이 없습니다. ID 열이 맞는지, 위 표에서 SAV 변수를 "
            "연결했는지 확인해 주세요."
        )
        st.stop()

    with st.expander(f"변수별 변경 셀 수 ({len(rep['changed'])}개 변수)"):
        st.dataframe(
            pd.DataFrame({
                "변수": list(rep["changed"].keys()),
                "바뀌는 셀": list(rep["changed"].values()),
                "문자형으로 바뀜": ["예" if k in rep["to_text"] else ""
                                for k in rep["changed"]],
            }),
            hide_index=True, use_container_width=True,
        )

    # 이 아래는 모두 반영된 데이터로 돈다.
    # 유형이 바뀐 변수를 반영하지 않으면 Open 시트가 옛 기준으로 만들어진다.
    work_df, work_labels = new_df, new_labels
    work_types = dict(var_types)
    for c in rep["to_text"]:
        work_types[c] = "string"
    patched = True

st.divider()

# ══════════════════════════════════════════════════════════════
#  2. 담을 시트 고르기
# ══════════════════════════════════════════════════════════════
st.subheader("2. 담을 시트 고르기")

key_cols = find_key_cols(work_df)
text_cols = find_text_cols(work_df, work_types)

s1, s2, s3, s4, s5 = st.columns(5)
want_raw = s1.checkbox("Raw (숫자 코드)", value=True)
want_label = s2.checkbox("Label (값 레이블)", value=True)
want_open = s3.checkbox(
    "Open (주관식)",
    value=True,
    help="키 변수(" + ", ".join(key_cols) + ")와 문자형 변수를 담습니다.",
)
want_code = s4.checkbox(
    "Code (코드북)",
    value=True,
    help="변수마다 문항과 코드값/보기를 블록으로 정리합니다. "
         "키 변수(" + ", ".join(key_cols) + ")는 제외합니다.",
)
want_guide = s5.checkbox("변수 가이드", value=True)

# ── 머리글 색 ──
PICK_OWN, NO_FILL = "직접 고르기", "색 없음"
choice = st.radio(
    "머리글 색",
    list(HEAD_COLORS) + [PICK_OWN, NO_FILL],
    horizontal=True,
    help="Code 시트는 변수명·문항 줄, 나머지 시트는 첫 행에 칠합니다.",
)
if choice == NO_FILL:
    head_color = ""
elif choice == PICK_OWN:
    head_color = st.color_picker("색 고르기", DEFAULT_HEAD_COLOR)
else:
    head_color = HEAD_COLORS[choice]

if head_color:
    st.markdown(
        f'<div style="display:flex;align-items:center;gap:10px;'
        f'font-size:13px;opacity:.75;margin:2px 0 6px;">'
        f'<span style="display:inline-block;width:74px;height:20px;'
        f'background:{head_color};border:1px solid rgba(128,128,128,.4);'
        f'border-radius:3px;"></span>{head_color.upper()}</div>',
        unsafe_allow_html=True,
    )
else:
    st.caption("색 없이 굵게만 표시됩니다.")

if want_open and not text_cols:
    st.info(
        "문자형 변수가 없어 Open 시트에 키 변수("
        + ", ".join(key_cols)
        + ")만 담깁니다. 주관식 응답을 옆에 붙여 코딩하실 때 쓰시면 됩니다."
    )

if not (want_raw or want_label or want_open or want_code or want_guide):
    st.warning("시트를 하나 이상 선택해주세요.")
    st.stop()

# ── 시트 구성 (Raw → Label → Open → Code → 변수 가이드) ──
sheets = {}
if want_raw:
    sheets["Raw"] = build_raw(work_df)
if want_label:
    sheets["Label"] = build_label(work_df, work_labels)
if want_open:
    sheets["Open"] = build_open(work_df, text_cols, key_cols)
if want_code:
    sheets["Code"] = build_codebook(work_df, col_labels, work_labels, key_cols)
if want_guide:
    sheets["변수 가이드"] = build_guide(work_df, col_labels)

with st.expander("미리보기", expanded=not patched):
    tabs = st.tabs(list(sheets.keys()))
    for tab, (name, frame) in zip(tabs, sheets.items()):
        with tab:
            st.dataframe(
                frame.head(20).astype(str).replace("None", ""),
                use_container_width=True,
                hide_index=True,
            )
            if len(frame) > 20:
                st.caption(f"위 20행만 표시 · 전체 {len(frame):,}행")

st.divider()

# ══════════════════════════════════════════════════════════════
#  3. 파일 만들기
# ══════════════════════════════════════════════════════════════
st.subheader("3. 파일 만들기")

stem = os.path.splitext(up.name)[0] + ("_반영" if patched else "")

if st.button("만들기", type="primary", use_container_width=True):
    with st.spinner("파일을 만드는 중입니다…"):
        try:
            st.session_state["SX_xlsx"] = to_excel(sheets, head_color)
            st.session_state["SX_sav"] = (
                write_sav(work_df, col_labels, work_labels) if patched else None)
        except Exception as e:
            st.error(f"파일 생성에 실패했습니다: {e}")
            st.stop()
    st.session_state["SX_stem"] = stem

if st.session_state.get("SX_xlsx"):
    st.success("다 됐습니다.")
    stem = st.session_state.get("SX_stem", stem)
    d1, d2 = st.columns(2)
    with d1:
        st.download_button(
            "엑셀 내려받기",
            data=st.session_state["SX_xlsx"],
            file_name=stem + ".xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    if st.session_state.get("SX_sav"):
        with d2:
            st.download_button(
                "SAV 내려받기",
                data=st.session_state["SX_sav"],
                file_name=stem + ".sav",
                mime="application/octet-stream",
                use_container_width=True,
            )
