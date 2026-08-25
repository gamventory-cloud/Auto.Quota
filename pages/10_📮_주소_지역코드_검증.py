# -*- coding: utf-8 -*-
"""
주소 → 지역코드(areaM / areaD) 검증
VBA 애드인(우편번호_주소_체크.xlam) 이식판  v3.0

핵심 흐름
  1) 주소 파일(ID + 주소 + 우편번호)에서 ID→주소 사전 생성
  2) 주소 텍스트로 시군구가 확정된 건에서 「우편번호 → 시군구」 대조표를 자동 학습
  3) 데이터 파일(ID + areaM + areaD)의 각 행마다 코드 계산
       판정 순서 : 주소 텍스트 → 우편번호 대조표 → 도로명 보정 규칙
  4) 입력값과 계산값을 대조해 일치 / 불일치 / 구군미확인 등으로 판정
  5) 결과는 원본을 건드리지 않고 별도 엑셀 파일로 저장
"""
from __future__ import annotations

import io
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

# ──────────────────────────────────────────────────────────────
# utils.py 의존성 (없어도 단독 실행 가능)
# ──────────────────────────────────────────────────────────────
try:
    from utils import check_password  # type: ignore
except Exception:  # pragma: no cover
    def check_password() -> bool:
        return True


# ══════════════════════════════════════════════════════════════
# CORE : 순수 로직 (streamlit 호출 없음 — 단위 테스트 가능)
# ══════════════════════════════════════════════════════════════

# 기준표 원본 (VBA GenerateCodeDB 와 동일)
_RAW_CODE_DB: List[Tuple[str, int, str]] = [
    ("서울", 1, "강북구,1,광진구,2,노원구,3,도봉구,4,동대문구,5,성동구,6,성북구,7,중랑구,8,"
                "마포구,9,서대문구,10,은평구,11,용산구,12,종로구,13,중구,14,강남구,15,강동구,16,"
                "서초구,17,송파구,18,강서구,19,관악구,20,구로구,21,금천구,22,동작구,23,양천구,24,영등포구,25"),
    ("부산", 2, "강서구,1,금정구,2,기장군,3,남구,4,동구,5,동래구,6,부산진구,7,북구,8,사상구,9,"
                "사하구,10,서구,11,수영구,12,연제구,13,영도구,14,중구,15,해운대구,16"),
    ("대구", 3, "남구,1,달서구,2,달성군,3,동구,4,북구,5,서구,6,수성구,7,중구,8,군위군,9"),
    ("인천", 4, "강화군,1,계양구,2,남동구,4,동구,5,부평구,6,서구,7,연수구,8,옹진군,9,중구,10,미추홀구,11"),
    ("광주", 5, "동구,1,서구,2,남구,3,북구,4,광산구,5"),
    ("대전", 6, "대덕구,1,동구,2,서구,3,유성구,4,중구,5"),
    ("울산", 7, "남구,1,동구,2,북구,3,울주군,4,중구,5"),
    ("경기", 8, "가평군,1,고양시,2,과천시,3,광명시,4,광주시,5,구리시,6,군포시,7,김포시,8,남양주시,9,"
                "동두천시,10,부천시,11,성남시,12,수원시,13,시흥시,14,안산시,15,안성시,16,안양시,17,"
                "양주시,18,양평군,19,여주시,20,연천군,21,오산시,22,용인시,23,의왕시,24,의정부시,25,"
                "이천시,26,파주시,27,평택시,28,포천시,29,하남시,30,화성시,31"),
    ("강원", 9, "강릉시,1,고성군,2,동해시,3,삼척시,4,속초시,5,양구군,6,양양군,7,영월군,8,원주시,9,"
                "인제군,10,정선군,11,철원군,12,춘천시,13,태백시,14,평창군,15,홍천군,16,화천군,17,횡성군,18"),
    ("충북", 10, "괴산군,1,단양군,2,보은군,3,영동군,4,옥천군,5,음성군,6,제천시,7,증평군,8,진천군,9,"
                 "청원군,10,청주시,11,충주시,12"),
    ("충남", 11, "계룡시,1,공주시,2,금산군,3,논산시,4,당진시,5,보령시,6,부여군,7,서산시,8,서천군,9,"
                 "아산시,10,예산군,12,천안시,13,청양군,14,태안군,15,홍성군,16"),
    ("전북", 12, "고창군,1,군산시,2,김제시,3,남원시,4,무주군,5,부안군,6,순창군,7,완주군,8,익산시,9,"
                 "임실군,10,장수군,11,전주시,12,정읍시,13,진안군,14"),
    ("전남", 13, "강진군,1,고흥군,2,곡성군,3,광양시,4,구례군,5,나주시,6,담양군,7,목포시,8,무안군,9,"
                 "보성군,10,순천시,11,신안군,12,여수시,13,영광군,14,영암군,15,완도군,16,장성군,17,"
                 "장흥군,18,진도군,19,함평군,20,해남군,21,화순군,22"),
    ("경북", 14, "경산시,1,경주시,2,고령군,3,구미시,4,김천시,6,문경시,7,봉화군,8,상주시,9,성주군,10,"
                 "안동시,11,영덕군,12,영양군,13,영주시,14,영천시,15,예천군,16,울진군,17,의성군,18,"
                 "청도군,19,청송군,20,칠곡군,21,포항시,22,울릉군,23"),
    ("경남", 15, "거제시,1,거창군,2,고성군,3,김해시,4,남해군,5,밀양시,6,사천시,7,산청군,8,양산시,9,"
                 "의령군,10,진주시,11,창녕군,12,창원시,13,통영시,14,하동군,15,함안군,16,함양군,17,합천군,18"),
    ("제주", 16, "제주시,1,서귀포시,2"),
    ("세종", 17, "소정면,1,전의면,2,전동면,3,조치원읍,4,연서면,5,연동면,6,연기면,7,부강면,8,도담동,9,"
                 "장군면,10,한솔동,11,금남면,12,보람동,13,새롬동,14,아름동,15,고운동,16,종촌동,17,"
                 "소담동,18,대평동,19,가람동,20,다정동,21,해밀동,22,산울동,23,누리동,24,한별동,25,"
                 "반곡동,26,집현동,27,합강동,28,다솜동,29,용호동,30,나성동,31,세종동,32,어진동,33"),
]

# 시도 표기 변형 → 표준 명칭
SIDO_ALIASES: Dict[str, List[str]] = {
    "서울": ["서울"], "부산": ["부산"], "대구": ["대구"], "인천": ["인천"],
    "광주": ["광주"], "대전": ["대전"], "울산": ["울산"], "세종": ["세종"],
    "경기": ["경기"], "강원": ["강원"],
    "충북": ["충북", "충청북"], "충남": ["충남", "충청남"],
    "전북": ["전북", "전라북"], "전남": ["전남", "전라남"],
    "경북": ["경북", "경상북"], "경남": ["경상남", "경남"],
    "제주": ["제주"],
}

# 도로명/마을 이름 → 법정동 (VBA ConvertSejongRoadToDong)
_RAW_ROAD_RULES: List[Tuple[str, str, str]] = [
    # (시도, 주소에 포함된 문구, 대체할 시군구명)  ※ 위에서부터 먼저 적용
    # ── 읍·면 (최우선) ──
    ("세종", "조치원", "조치원읍"), ("세종", "이화", "조치원읍"),
    ("세종", "전의", "전의면"), ("세종", "노곡", "전의면"),
    ("세종", "소정", "소정면"), ("세종", "전동", "전동면"),
    ("세종", "연서", "연서면"), ("세종", "연동", "연동면"),
    ("세종", "연기", "연기면"), ("세종", "부강", "부강면"),
    ("세종", "장군", "장군면"), ("세종", "금남", "금남면"),
    # ── 도로명 : 이름에 동명이 그대로 들어간 것 ──
    ("세종", "다정", "다정동"), ("세종", "대평", "대평동"),
    ("세종", "산울", "산울동"), ("세종", "소담", "소담동"),
    ("세종", "보람", "보람동"), ("세종", "새롬", "새롬동"),
    ("세종", "반곡", "반곡동"), ("세종", "가람", "가람동"),
    ("세종", "나성", "나성동"), ("세종", "다솜", "다솜동"),
    ("세종", "해밀", "해밀동"), ("세종", "아름", "아름동"),
    ("세종", "종촌", "종촌동"), ("세종", "집현", "집현동"),
    ("세종", "합강", "합강동"), ("세종", "용호", "용호동"),
    ("세종", "한별", "한별동"), ("세종", "고운", "고운동"),
    # ── 도로명 : 동명과 다른 것 (※ 확인 후 수정하세요) ──
    ("세종", "밝은뜰", "고운동"), ("세종", "마음", "고운동"),
    ("세종", "만남로", "고운동"), ("세종", "가온로", "다정동"),
    ("세종", "도움3로", "종촌동"), ("세종", "도움", "어진동"),
    ("세종", "달빛", "종촌동"), ("세종", "보듬", "도담동"),
    ("세종", "금송로", "도담동"), ("세종", "한누리", "나성동"),
    ("세종", "시청대로", "보람동"), ("세종", "남세종로", "보람동"),
    ("세종", "국책연구원", "반곡동"), ("세종", "누리로", "반곡동"),
    ("세종", "절재로", "소담동"), ("세종", "중앙공원서로", "세종동"),
    ("세종", "갈매", "집현동"),
]

RESULT_MATCH = "일치"
RESULT_DIFF = "불일치"
RESULT_NO_SGG = "구군미확인"
RESULT_NO_SIDO = "주소해석불가"
RESULT_NO_ADDR = "주소없음"
RESULT_NO_ID = "ID없음"
RESULT_NO_INPUT = "입력값없음"


def build_default_code_df() -> pd.DataFrame:
    """기준표를 긴 형태 DataFrame으로 전개."""
    rows = []
    for sido, mcode, blob in _RAW_CODE_DB:
        parts = [p.strip() for p in blob.split(",")]
        for i in range(0, len(parts), 2):
            rows.append({
                "시도": sido,
                "areaM": int(mcode),
                "시군구": parts[i],
                "areaD": int(parts[i + 1]),
            })
    return pd.DataFrame(rows)


def build_default_road_df() -> pd.DataFrame:
    return pd.DataFrame(_RAW_ROAD_RULES, columns=["시도", "포함문구", "시군구"])


def nospace(value) -> str:
    """공백 제거 + 문자열화. NaN 안전."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return re.sub(r"\s+", "", str(value))


def norm_zip(value) -> str:
    """우편번호를 숫자 5자리로 정규화. 구 6자리(135-080) 등은 빈 문자열."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    digits = re.sub(r"\D", "", str(value))
    return digits.zfill(5) if 0 < len(digits) <= 5 else ""


class AreaResolver:
    """기준표 + 도로명 규칙을 받아 주소 한 건을 코드로 변환."""

    def __init__(self, code_df: pd.DataFrame, road_df: Optional[pd.DataFrame] = None,
                 zip_df: Optional[pd.DataFrame] = None, infer_sido: bool = True):
        self.infer_sido = infer_sido

        # 시도 → [(시군구키워드, 시군구명, areaM, areaD)] , 긴 이름 우선 정렬
        self.by_sido: Dict[str, List[Tuple[str, str, int, int]]] = {}
        self.sido_code: Dict[str, int] = {}
        for _, r in code_df.iterrows():
            sido = nospace(r["시도"])
            sgg = nospace(r["시군구"])
            if not sido or not sgg:
                continue
            try:
                m = int(r["areaM"])
                d = int(r["areaD"])
            except (TypeError, ValueError):
                continue
            self.by_sido.setdefault(sido, []).append((sgg, sgg, m, d))
            self.sido_code.setdefault(sido, m)
        for k in self.by_sido:
            self.by_sido[k].sort(key=lambda t: -len(t[0]))

        # 시도 없는 주소용: 전국에서 딱 하나뿐인 시군구 이름만 역추적에 사용
        counter: Dict[str, List[str]] = {}
        for sido, items in self.by_sido.items():
            for kw, *_ in items:
                counter.setdefault(kw, []).append(sido)
        self.unique_sgg = {kw: v[0] for kw, v in counter.items() if len(set(v)) == 1}

        # 도로명 규칙
        self.road: Dict[str, List[Tuple[str, str]]] = {}
        if road_df is not None:
            for _, r in road_df.iterrows():
                sido = nospace(r.get("시도"))
                token = nospace(r.get("포함문구"))
                target = nospace(r.get("시군구"))
                if sido and token and target:
                    self.road.setdefault(sido, []).append((token, target))

        # 우편번호 → (시도, 시군구)
        self.zip_map: Dict[str, Tuple[str, str]] = {}
        if zip_df is not None:
            for _, r in zip_df.iterrows():
                z = norm_zip(r.get("우편번호"))
                sido = nospace(r.get("시도"))
                sgg = nospace(r.get("시군구"))
                if z and sido and sgg:
                    self.zip_map[z] = (sido, sgg)

        # 시도 별칭 (긴 것부터)
        self.alias: List[Tuple[str, str]] = []
        for std, names in SIDO_ALIASES.items():
            for n in names:
                self.alias.append((n, std))
        self.alias.sort(key=lambda t: -len(t[0]))

    # ── 시도 판정 : 주소 전체에서 가장 앞에 등장하는 시도명 채택 ──
    def detect_sido(self, clean: str) -> Optional[str]:
        best_pos, best_len, best = 10**9, 0, None
        for name, std in self.alias:
            pos = clean.find(name)
            if pos < 0:
                continue
            if pos < best_pos or (pos == best_pos and len(name) > best_len):
                best_pos, best_len, best = pos, len(name), std
        return best

    # ── 시군구 판정 : 앞에 등장하는 것 우선, 같으면 긴 이름 우선 ──
    def _find_sgg(self, clean: str, sido: str):
        best = None
        best_key = (10**9, 0)
        for kw, name, m, d in self.by_sido.get(sido, []):
            pos = clean.find(kw)
            if pos < 0:
                continue
            key = (pos, -len(kw))
            if key < best_key:
                best_key, best = key, (name, m, d)
        return best

    def _code_of(self, sido: str, sgg: str):
        for kw, name, m, d in self.by_sido.get(sido, []):
            if kw == sgg:
                return (name, m, d)
        return None

    def resolve_text(self, address) -> Optional[Tuple[str, str]]:
        """주소 텍스트만으로 (시도, 시군구) 확정. 대조표 학습에 사용."""
        clean = nospace(address)
        if not clean:
            return None
        sido = self.detect_sido(clean)
        if sido is None:
            return None
        hit = self._find_sgg(clean, sido)
        return (sido, hit[0]) if hit else None

    def resolve(self, address, zipcode="") -> dict:
        """판정 순서: 주소 텍스트 → 우편번호 대조표 → 도로명 규칙."""
        clean = nospace(address)
        z = norm_zip(zipcode)
        zip_hit = self.zip_map.get(z) if z else None

        if not clean and not zip_hit:
            return {"status": RESULT_NO_ADDR, "areaM": None, "areaD": None,
                    "시도": "", "시군구": "", "근거": ""}

        # ① 주소 텍스트
        sido = self.detect_sido(clean) if clean else None
        if sido is None and clean and self.infer_sido:
            for kw in sorted(self.unique_sgg, key=len, reverse=True):
                if kw in clean:
                    sido = self.unique_sgg[kw]
                    break
        text_hit = self._find_sgg(clean, sido) if sido else None

        if text_hit:
            name, m, d = text_hit
            basis = "주소"
            if zip_hit and zip_hit != (sido, name):
                basis = f"주소(우편번호는 {zip_hit[1]})"
            return {"status": "OK", "areaM": m, "areaD": d,
                    "시도": sido, "시군구": name, "근거": basis}

        # ② 우편번호 대조표
        if zip_hit:
            code = self._code_of(*zip_hit)
            if code:
                name, m, d = code
                return {"status": "OK", "areaM": m, "areaD": d,
                        "시도": zip_hit[0], "시군구": name, "근거": "우편번호"}

        # ③ 도로명 보정 규칙
        if sido:
            for token, target in self.road.get(sido, []):
                if token in clean:
                    code = self._code_of(sido, target)
                    if code:
                        name, m, d = code
                        return {"status": "OK", "areaM": m, "areaD": d,
                                "시도": sido, "시군구": name,
                                "근거": f"도로명 규칙({token})"}
            return {"status": RESULT_NO_SGG,
                    "areaM": self.sido_code.get(sido), "areaD": None,
                    "시도": sido, "시군구": "", "근거": "시도만 확인"}

        return {"status": RESULT_NO_SIDO, "areaM": None, "areaD": None,
                "시도": "", "시군구": "", "근거": ""}


def to_code(value) -> Optional[int]:
    """'01' → 1, 공백/문자 → None (VBA Val 과 유사하되 빈값 구분)."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    s = str(value).strip()
    if s == "":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def verify(
    target_df: pd.DataFrame,
    addr_map: Dict[str, Tuple[str, str]],
    col_id: str,
    col_m: str,
    col_d: str,
    resolver: AreaResolver,
) -> pd.DataFrame:
    """검증 대상 DataFrame에 결과 열을 붙여 반환 (원본 열 보존)."""
    cache: Dict[Tuple[str, str], dict] = {}
    out_result, out_m, out_d = [], [], []
    out_sido, out_sgg, out_basis, out_addr, out_zip = [], [], [], [], []

    for _, row in target_df.iterrows():
        key = nospace(row[col_id])
        user_m = to_code(row[col_m])
        user_d = to_code(row[col_d])

        if key == "" or key not in addr_map:
            out_result.append(RESULT_NO_ID)
            out_m.append(None); out_d.append(None)
            out_sido.append(""); out_sgg.append(""); out_basis.append("")
            out_addr.append(""); out_zip.append("")
            continue

        address, zipcode = addr_map[key]
        ckey = (address, zipcode)
        if ckey not in cache:
            cache[ckey] = resolver.resolve(address, zipcode)
        info = cache[ckey]

        out_addr.append(address)
        out_zip.append(zipcode)
        out_sido.append(info["시도"])
        out_sgg.append(info["시군구"])
        out_basis.append(info["근거"])

        if info["status"] in (RESULT_NO_ADDR, RESULT_NO_SIDO):
            out_result.append(info["status"])
            out_m.append(info["areaM"]); out_d.append(info["areaD"])
        elif info["status"] == RESULT_NO_SGG:
            out_result.append(RESULT_NO_SGG)
            out_m.append(info["areaM"]); out_d.append(None)
        elif user_m is None and user_d is None:
            out_result.append(RESULT_NO_INPUT)
            out_m.append(info["areaM"]); out_d.append(info["areaD"])
        elif user_m == info["areaM"] and user_d == info["areaD"]:
            out_result.append(RESULT_MATCH)
            out_m.append(None); out_d.append(None)
        else:
            out_result.append(RESULT_DIFF)
            out_m.append(info["areaM"]); out_d.append(info["areaD"])

    res = target_df.copy()
    res["검증결과"] = out_result
    res["계산_areaM"] = out_m
    res["계산_areaD"] = out_d
    res["매칭_시도"] = out_sido
    res["매칭_시군구"] = out_sgg
    res["판정근거"] = out_basis
    res["참조주소"] = out_addr
    res["참조우편번호"] = out_zip
    return res


def make_addr_map(df: pd.DataFrame, col_id: str, col_addr: str,
                  col_zip: Optional[str] = None) -> Tuple[Dict[str, Tuple[str, str]], int]:
    """ID→(주소, 우편번호) 사전. 중복 ID는 첫 값 우선. 중복 건수도 반환."""
    mapping: Dict[str, Tuple[str, str]] = {}
    dup = 0
    for _, row in df.iterrows():
        key = nospace(row[col_id])
        if key == "":
            continue
        if key in mapping:
            dup += 1
            continue
        val = row[col_addr]
        addr = "" if (val is None or (isinstance(val, float) and pd.isna(val))) else str(val).strip()
        z = norm_zip(row[col_zip]) if col_zip else ""
        mapping[key] = (addr, z)
    return mapping, dup


def learn_zip_map(df: pd.DataFrame, col_addr: str, col_zip: str,
                  resolver: "AreaResolver") -> Tuple[pd.DataFrame, pd.DataFrame]:
    """주소 텍스트로 시군구가 확정된 건에서 우편번호 대조표를 학습.

    반환: (학습표, 충돌표)
      학습표 — 우편번호 하나가 시군구 하나로만 대응하는 것
      충돌표 — 같은 우편번호에 시군구가 여럿 붙은 것 (주소 오기재 의심)
    """
    seen: Dict[str, Dict[Tuple[str, str], int]] = {}
    for _, row in df.iterrows():
        z = norm_zip(row[col_zip])
        if not z:
            continue
        hit = resolver.resolve_text(row[col_addr])
        if hit:
            seen.setdefault(z, {})
            seen[z][hit] = seen[z].get(hit, 0) + 1

    learned, clash = [], []
    for z, counts in sorted(seen.items()):
        if len(counts) == 1:
            (sido, sgg), n = next(iter(counts.items()))
            learned.append({"우편번호": z, "시도": sido, "시군구": sgg, "출처": "학습", "건수": n})
        else:
            for (sido, sgg), n in sorted(counts.items(), key=lambda kv: -kv[1]):
                clash.append({"우편번호": z, "시도": sido, "시군구": sgg, "건수": n})
    return (pd.DataFrame(learned, columns=["우편번호", "시도", "시군구", "출처", "건수"]),
            pd.DataFrame(clash, columns=["우편번호", "시도", "시군구", "건수"]))


def merge_zip_tables(*tables: Optional[pd.DataFrame]) -> pd.DataFrame:
    """여러 대조표를 병합. 뒤에 오는 표가 우선(수동 입력이 학습분을 덮어씀)."""
    frames = [t for t in tables if t is not None and len(t)]
    if not frames:
        return pd.DataFrame(columns=["우편번호", "시도", "시군구", "출처"])
    out = pd.concat(frames, ignore_index=True)
    out["우편번호"] = out["우편번호"].map(norm_zip)
    out = out[out["우편번호"] != ""]
    out = out.drop_duplicates(subset=["우편번호"], keep="last").sort_values("우편번호")
    for c in ("시도", "시군구", "출처"):
        if c not in out.columns:
            out[c] = ""
    return out[["우편번호", "시도", "시군구", "출처"]].reset_index(drop=True)


def guess_col(columns: List[str], candidates: List[str]) -> Optional[str]:
    """헤더 자동 탐색 (대소문자·공백 무시, 완전일치 → 부분일치)."""
    norm = {c: nospace(c).upper() for c in columns}
    for cand in candidates:
        target = nospace(cand).upper()
        for c in columns:
            if norm[c] == target:
                return c
    for cand in candidates:
        target = nospace(cand).upper()
        for c in columns:
            if target and target in norm[c]:
                return c
    return None


# ══════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════

ADDR_ID_CANDS = ["ID", "id", "panel_id", "패널ID"]
ADDR_CANDS = ["HO_ADD1", "주소", "Address", "addr", "HO_ADD"]
ZIP_CANDS = ["HO_ZIP", "우편번호", "zipcode", "zip", "post"]
DATA_ID_CANDS = ["id", "ID", "panel_id", "f_id"]
M_CANDS = ["areaM", "area_M", "AREAM"]
D_CANDS = ["areaD", "areaQ", "area_D", "AREAD"]
NONE_LABEL = "(사용 안 함)"


@st.cache_data(show_spinner=False, max_entries=6)
def sheet_names(data: bytes, filename: str) -> List[str]:
    """파일 전체를 읽지 않고 시트 이름만 가져온다."""
    low = filename.lower()
    if low.endswith(".csv"):
        return ["(CSV)"]
    if low.endswith(".xls"):
        import xlrd
        return xlrd.open_workbook(file_contents=data, on_demand=True).sheet_names()
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


@st.cache_data(show_spinner=False, max_entries=4)
def read_sheet(data: bytes, filename: str, sheet: str) -> pd.DataFrame:
    """선택한 시트 하나만 읽는다 (큰 파일 메모리 절약)."""
    if filename.lower().endswith(".csv"):
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                return pd.read_csv(io.BytesIO(data), dtype=object, encoding=enc)
            except UnicodeDecodeError:
                continue
        return pd.read_csv(io.BytesIO(data), dtype=object, encoding="utf-8",
                           encoding_errors="replace")
    return pd.read_excel(io.BytesIO(data), sheet_name=sheet, dtype=object)


@st.cache_data(show_spinner=False, max_entries=3)
def read_zip_csv(data: bytes) -> pd.DataFrame:
    for enc in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            return pd.read_csv(io.BytesIO(data), dtype=object, encoding=enc)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(io.BytesIO(data), dtype=object, encoding_errors="replace")


def pick_sheet(label: str, data: bytes, filename: str, key: str) -> str:
    names = [n for n in sheet_names(data, filename) if n != "CodeDB"]
    if not names:
        st.error("읽을 수 있는 시트가 없습니다.")
        st.stop()
    if len(names) == 1:
        return names[0]
    return st.selectbox(label, names, key=key)


def _idx(options: List[str], value: Optional[str]) -> int:
    return options.index(value) if value in options else 0


def build_output(result: pd.DataFrame, keep_cols: List[str], full: bool) -> pd.DataFrame:
    added = ["검증결과", "계산_areaM", "계산_areaD", "매칭_시도", "매칭_시군구",
             "판정근거", "참조주소", "참조우편번호"]
    if full:
        return result
    cols = [c for c in keep_cols if c in result.columns] + added
    seen, ordered = set(), []
    for c in cols:
        if c not in seen:
            seen.add(c); ordered.append(c)
    return result[ordered]


def to_excel(result: pd.DataFrame, summary: pd.DataFrame,
             zip_df: Optional[pd.DataFrame] = None,
             clash: Optional[pd.DataFrame] = None) -> bytes:
    from openpyxl.styles import Font, PatternFill
    buf = io.BytesIO()
    need = result[result["검증결과"].isin([RESULT_DIFF, RESULT_NO_SGG, RESULT_NO_SIDO, RESULT_NO_ID])]
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="요약", index=False)
        if not need.empty:
            need.to_excel(writer, sheet_name="확인필요", index=False)
        result.to_excel(writer, sheet_name="전체결과", index=False)
        if clash is not None and len(clash):
            clash.to_excel(writer, sheet_name="우편번호충돌", index=False)
        if zip_df is not None and len(zip_df):
            zip_df.to_excel(writer, sheet_name="우편번호대조표", index=False)

        head_fill = PatternFill("solid", fgColor="E2EFDA")
        red = Font(color="FF0000", bold=True)
        cols = list(result.columns)
        for sheet_name, frame in (("확인필요", need), ("전체결과", result)):
            if sheet_name not in writer.sheets:
                continue
            ws = writer.sheets[sheet_name]
            for name in ["검증결과", "계산_areaM", "계산_areaD", "매칭_시도",
                         "매칭_시군구", "판정근거", "참조주소", "참조우편번호"]:
                if name in cols:
                    cell = ws.cell(row=1, column=cols.index(name) + 1)
                    cell.fill = head_fill
                    cell.font = Font(bold=True)
            idx = cols.index("검증결과") + 1
            for i, val in enumerate(frame["검증결과"], start=2):
                if val in (RESULT_DIFF, RESULT_NO_SIDO):
                    ws.cell(row=i, column=idx).font = red
            ws.freeze_panes = "A2"
    return buf.getvalue()


def unresolved_zip_table(result: pd.DataFrame) -> pd.DataFrame:
    """확정도가 낮은 건을 우편번호 단위로 묶어, 한 번만 확인하면 되는 목록으로.

    대상 ① 구군미확인  ② 도로명 보정 규칙으로 판정된 건(추정이므로 확인 대상)
    """
    mask = ((result["검증결과"] == RESULT_NO_SGG) |
            (result["판정근거"].astype(str).str.startswith("도로명 규칙")))
    bad = result[mask & (result["참조우편번호"].astype(str) != "")]
    if bad.empty:
        return pd.DataFrame(columns=["우편번호", "시도", "현재판정", "시군구", "건수", "주소예시"])
    rows = []
    for z, grp in bad.groupby("참조우편번호"):
        samples = sorted({str(a) for a in grp["참조주소"] if str(a).strip()})[:3]
        guesses = sorted({str(g) for g in grp["매칭_시군구"] if str(g).strip()})
        rows.append({
            "우편번호": z,
            "시도": grp["매칭_시도"].iloc[0],
            "현재판정": " / ".join(guesses) if guesses else "(미확인)",
            "시군구": "",
            "건수": len(grp),
            "주소예시": "  /  ".join(samples),
        })
    return pd.DataFrame(rows).sort_values("건수", ascending=False).reset_index(drop=True)


def main() -> None:
    st.set_page_config(page_title="주소 지역코드 검증", page_icon="📮", layout="wide")
    if not check_password():
        st.stop()

    st.title("📮 주소 지역코드 검증")
    st.caption("주소와 우편번호로 areaM · areaD 를 계산해, 데이터 파일에 입력된 값과 ID로 맞춰 대조합니다.")

    ss = st.session_state
    ss.setdefault("code_df", build_default_code_df())
    ss.setdefault("road_df", build_default_road_df())
    ss.setdefault("zip_manual", pd.DataFrame(columns=["우편번호", "시도", "시군구", "출처"]))
    ss.setdefault("zip_learned", pd.DataFrame(columns=["우편번호", "시도", "시군구", "출처"]))
    ss.setdefault("result", None)

    tab_run, tab_zip, tab_code, tab_road = st.tabs(
        ["검증", "우편번호 대조표", "지역코드 기준표", "도로명 보정 규칙"])

    # ── 우편번호 대조표 ───────────────────────────────────────
    with tab_zip:
        st.markdown(
            "우편번호 5자리는 시군구와 일대일로 대응합니다. 주소 텍스트에 시군구명이 없는 경우"
            "(세종시 신도심 도로명 주소 등)에 이 표로 판정합니다.\n\n"
            "검증을 실행하면 **주소로 시군구가 확정된 건에서 대조표를 자동으로 학습**합니다. "
            "그래도 남는 우편번호는 아래에서 직접 채우고, CSV로 내려받아 다음 프로젝트에 다시 올리면 계속 쌓입니다."
        )
        up_zip = st.file_uploader("저장해둔 대조표 CSV 불러오기", type=["csv"], key="up_zip")
        if up_zip is not None and st.button("불러오기", key="btn_load_zip"):
            try:
                loaded = read_zip_csv(up_zip.getvalue())
                loaded["출처"] = "불러옴"
                ss.zip_manual = merge_zip_tables(loaded, ss.zip_manual)
                st.success(f"{len(loaded):,}건을 불러왔습니다.")
            except Exception as exc:
                st.error(f"읽지 못했습니다: {exc}")

        st.markdown("**직접 입력 · 수정한 항목** (학습분보다 우선 적용됩니다)")
        ss.zip_manual = st.data_editor(
            ss.zip_manual, num_rows="dynamic", use_container_width=True,
            height=260, key="editor_zip",
            column_config={
                "우편번호": st.column_config.TextColumn(width="small"),
                "시도": st.column_config.TextColumn(width="small"),
                "시군구": st.column_config.TextColumn(width="small"),
                "출처": st.column_config.TextColumn(width="small"),
            },
        )
        merged = merge_zip_tables(ss.zip_learned, ss.zip_manual)
        c1, c2 = st.columns([1, 3])
        c1.metric("대조표 총 건수", f"{len(merged):,}")
        c2.download_button(
            "대조표 CSV 내려받기 (학습분 + 직접입력)",
            merged.to_csv(index=False).encode("utf-8-sig"),
            "우편번호_대조표.csv", "text/csv", use_container_width=True,
        )

    # ── 기준표 ────────────────────────────────────────────────
    with tab_code:
        st.markdown("행정구역이 바뀌면 여기서 직접 고치면 됩니다. 고친 내용은 바로 검증에 반영됩니다.")
        ss.code_df = st.data_editor(
            ss.code_df, num_rows="dynamic", use_container_width=True,
            height=420, key="editor_code",
            column_config={
                "시도": st.column_config.TextColumn(width="small"),
                "areaM": st.column_config.NumberColumn("areaM (시도코드)", width="small"),
                "시군구": st.column_config.TextColumn(width="small"),
                "areaD": st.column_config.NumberColumn("areaD (시군구코드)", width="small"),
            },
        )
        c1, c2 = st.columns([1, 4])
        if c1.button("기본값으로 되돌리기", key="reset_code"):
            ss.code_df = build_default_code_df()
            st.rerun()
        c2.download_button("기준표 CSV 내려받기",
                           ss.code_df.to_csv(index=False).encode("utf-8-sig"),
                           "지역코드_기준표.csv", "text/csv")

    # ── 도로명 규칙 ───────────────────────────────────────────
    with tab_road:
        st.markdown(
            "주소에 시군구명이 없고 우편번호로도 못 찾을 때 마지막으로 쓰는 보정표입니다. "
            "**주소에 「포함문구」가 있으면 해당 시군구로 간주**하며, 위에 있는 규칙이 먼저 적용됩니다. "
            "이 규칙으로 판정된 건은 결과의 `판정근거` 열에 `도로명 규칙(...)` 으로 표시되니, "
            "정확도가 걱정되면 그 건만 따로 확인하면 됩니다."
        )
        ss.road_df = st.data_editor(
            ss.road_df, num_rows="dynamic", use_container_width=True,
            height=420, key="editor_road",
            column_config={
                "시도": st.column_config.TextColumn(width="small"),
                "포함문구": st.column_config.TextColumn("포함문구 (예: 밝은뜰)", width="medium"),
                "시군구": st.column_config.TextColumn("→ 시군구", width="small"),
            },
        )
        if st.button("기본값으로 되돌리기", key="reset_road"):
            ss.road_df = build_default_road_df()
            st.rerun()

    # ── 검증 ──────────────────────────────────────────────────
    with tab_run:
        st.subheader("1. 파일")
        one_file = st.checkbox("주소와 데이터가 한 파일에 있음", value=False)

        u1, u2 = st.columns(2)
        with u1:
            f_addr = st.file_uploader("주소 파일 (ID + 주소 + 우편번호)",
                                      type=["xlsx", "xlsm", "xls", "csv"], key="up_addr")
        with u2:
            f_data = (f_addr if one_file else
                      st.file_uploader("데이터 파일 (ID + areaM + areaD)",
                                       type=["xlsx", "xlsm", "xls", "csv"], key="up_data"))
            if one_file:
                st.info("같은 파일에서 시트·열을 따로 고릅니다.")

        if f_addr is None or f_data is None:
            st.info("파일을 올리면 열 선택 화면이 나타납니다.")
            return

        b_addr, b_data = f_addr.getvalue(), f_data.getvalue()
        try:
            with st.spinner("파일을 읽는 중…"):
                sh_addr = pick_sheet("주소 파일 시트", b_addr, f_addr.name, "sh_addr")
                df_addr = read_sheet(b_addr, f_addr.name, sh_addr)
                sh_data = pick_sheet("데이터 파일 시트", b_data, f_data.name, "sh_data")
                df_data = read_sheet(b_data, f_data.name, sh_data)
        except Exception as exc:
            st.error(f"파일을 읽지 못했습니다: {exc}")
            return

        st.caption(f"주소 {len(df_addr):,}행 · {len(df_addr.columns)}열   |   "
                   f"데이터 {len(df_data):,}행 · {len(df_data.columns)}열")

        st.subheader("2. 열 지정")
        left, right = st.columns(2)
        ac, dc = list(df_addr.columns), list(df_data.columns)
        with left:
            st.markdown("**주소 파일**")
            c_sid = st.selectbox("ID 열", ac, key="c_sid", index=_idx(ac, guess_col(ac, ADDR_ID_CANDS)))
            c_addr = st.selectbox("주소 열", ac, key="c_addr", index=_idx(ac, guess_col(ac, ADDR_CANDS)))
            zopts = [NONE_LABEL] + ac
            c_zip = st.selectbox("우편번호 열", zopts, key="c_zip",
                                 index=_idx(zopts, guess_col(ac, ZIP_CANDS)))
        with right:
            st.markdown("**데이터 파일**")
            c_tid = st.selectbox("ID 열", dc, key="c_tid", index=_idx(dc, guess_col(dc, DATA_ID_CANDS)))
            cm1, cm2 = st.columns(2)
            c_m = cm1.selectbox("areaM 열", dc, key="c_m", index=_idx(dc, guess_col(dc, M_CANDS)))
            c_d = cm2.selectbox("areaD 열", dc, key="c_d", index=_idx(dc, guess_col(dc, D_CANDS)))

        o1, o2 = st.columns(2)
        infer = o1.checkbox("시도명 없는 주소는 시군구 이름으로 추정", value=True,
                            help="예: 「성남시 분당구…」처럼 시도가 빠진 주소도 전국에서 이름이 하나뿐이면 찾아냅니다.")
        full = o2.checkbox("결과 파일에 데이터 파일의 모든 열 포함", value=False,
                           help="끄면 ID · areaM · areaD · 결과 열만 담깁니다. 열이 많은 파일은 꺼두는 쪽이 빠릅니다.")

        zip_col = None if c_zip == NONE_LABEL else c_zip

        if st.button("검증 실행", type="primary", use_container_width=True):
            with st.spinner("주소를 해석하는 중…"):
                base = AreaResolver(ss.code_df, ss.road_df, infer_sido=infer)
                clash = pd.DataFrame()
                if zip_col:
                    learned, clash = learn_zip_map(df_addr, c_addr, zip_col, base)
                    ss.zip_learned = merge_zip_tables(
                        learned[["우편번호", "시도", "시군구", "출처"]])
                zip_table = merge_zip_tables(ss.zip_learned, ss.zip_manual)
                resolver = AreaResolver(ss.code_df, ss.road_df, zip_table, infer_sido=infer)
                addr_map, dup = make_addr_map(df_addr, c_sid, c_addr, zip_col)
                result = verify(df_data, addr_map, c_tid, c_m, c_d, resolver)
            ss.result = dict(result=result, dup=dup, clash=clash, zip_table=zip_table,
                             keep=[c for c in ("No", "no") if c in df_data.columns] + [c_tid, c_m, c_d],
                             full=full, fname=f_data.name)

        if ss.result is None:
            return

        r = ss.result
        result, output = r["result"], build_output(r["result"], r["keep"], r["full"])

        st.subheader("3. 결과")
        if r["dup"]:
            st.warning(f"주소 파일에 중복 ID가 {r['dup']:,}건 있습니다. 첫 번째 주소를 사용했습니다.")

        counts = result["검증결과"].value_counts()
        order = [RESULT_MATCH, RESULT_DIFF, RESULT_NO_SGG, RESULT_NO_SIDO,
                 RESULT_NO_ADDR, RESULT_NO_ID, RESULT_NO_INPUT]
        shown = [k for k in order if k in counts]
        for col, key in zip(st.columns(max(len(shown), 1)), shown):
            col.metric(key, f"{int(counts[key]):,}")
        summary = pd.DataFrame({"검증결과": shown, "건수": [int(counts[k]) for k in shown]})

        basis = result[result["판정근거"] != ""]["판정근거"].str.replace(r"\(.*\)", "", regex=True)
        if len(basis):
            st.caption("판정근거별 — " + " · ".join(f"{k} {v:,}건" for k, v in basis.value_counts().items()))

        clash = r["clash"]
        if clash is not None and len(clash):
            st.error(f"같은 우편번호에 서로 다른 시군구가 붙은 건이 {clash['우편번호'].nunique()}개 있습니다. "
                     "주소가 잘못 적힌 응답자일 수 있습니다.")
            st.dataframe(clash, use_container_width=True, height=180)

        # 미확인 우편번호 채우기
        todo = unresolved_zip_table(result)
        if len(todo):
            st.markdown(f"#### 우편번호 {len(todo)}개를 확인하면 {int(todo['건수'].sum()):,}건이 확정됩니다")
            st.caption("주소에 시군구명이 없어 도로명 규칙으로 추정했거나, 아예 못 찾은 건들입니다. "
                       "`현재판정`이 맞으면 그대로 두고, 틀렸으면 `시군구`에 올바른 값을 적은 뒤 "
                       "「대조표에 반영」을 누르세요. 반영분은 대조표 탭에서 CSV로 저장해 다음 프로젝트에 재사용할 수 있습니다.")
            edited = st.data_editor(
                todo, use_container_width=True, height=300, key="editor_todo",
                disabled=["우편번호", "시도", "현재판정", "건수", "주소예시"],
                column_config={
                    "우편번호": st.column_config.TextColumn(width="small"),
                    "시도": st.column_config.TextColumn(width="small"),
                    "현재판정": st.column_config.TextColumn("현재 판정", width="small"),
                    "시군구": st.column_config.TextColumn("올바른 시군구 (입력)", width="small"),
                    "건수": st.column_config.NumberColumn(width="small"),
                    "주소예시": st.column_config.TextColumn(width="large"),
                },
            )
            if st.button("대조표에 반영하고 다시 검증", key="btn_apply_todo"):
                filled = edited[edited["시군구"].astype(str).str.strip() != ""].copy()
                if filled.empty:
                    st.warning("입력된 시군구가 없습니다.")
                else:
                    filled["출처"] = "수동"
                    ss.zip_manual = merge_zip_tables(ss.zip_manual, filled[["우편번호", "시도", "시군구", "출처"]])
                    st.success(f"{len(filled)}건을 반영했습니다. 「검증 실행」을 다시 눌러주세요.")

        need = output[output["검증결과"].isin([RESULT_DIFF, RESULT_NO_SGG, RESULT_NO_SIDO, RESULT_NO_ID])]
        if need.empty:
            st.success("확인이 필요한 건이 없습니다.")
        else:
            st.markdown(f"**확인 필요 {len(need):,}건**")
            st.dataframe(need, use_container_width=True, height=400)

        with st.spinner("결과 파일을 만드는 중…"):
            blob = to_excel(output, summary, r["zip_table"], clash)
        st.download_button(
            "결과 엑셀 내려받기",
            blob,
            f"주소검증결과_{r['fname'].rsplit('.', 1)[0]}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


if __name__ == "__main__":
    main()
