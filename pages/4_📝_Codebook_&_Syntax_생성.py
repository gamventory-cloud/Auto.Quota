import streamlit as st
import pandas as pd
import sys
import os
import re
import io
import textwrap
import collections
from collections import Counter

# 워드/엑셀 관련 라이브러리
try:
    from docx import Document
    from docx.document import Document as _Document
    from docx.oxml.text.paragraph import CT_P
    from docx.oxml.table import CT_Tbl
    from docx.table import _Cell, Table
    from docx.text.paragraph import Paragraph
    from docx.oxml.ns import qn 
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    st.error("필수 라이브러리가 설치되지 않았습니다. requirements.txt에 'python-docx'와 'openpyxl'을 추가하세요.")
    st.stop()

# 1. 상위 폴더의 utils.py를 불러오기 위한 경로 설정
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import utils

# 2. 페이지 기본 설정
st.set_page_config(page_title="설문지 코드북 생성", layout="wide")

# 3. 비밀번호 잠금
if not utils.check_password():
    st.stop()

st.title("📝 설문지 읽기 & 코드북/신텍스 자동 생성 (AHP & Full Logic)")

# ==============================================================================
# [Part 0] 동그라미 숫자 매핑 (추가됨)
# ==============================================================================
CIRCLE_MAP = {'①':'1','②':'2','③':'3','④':'4','⑤':'5','⑥':'6','⑦':'7','⑧':'8','⑨':'9','⑩':'10'}

# ==============================================================================
# [Part 1] 핵심 파싱 함수
# ==============================================================================

def iter_block_items(parent):
    if isinstance(parent, _Document):
        parent_elm = parent.element.body
    elif isinstance(parent, _Cell):
        parent_elm = parent._tc
    else:
        raise ValueError("iter_block_items: 지원하지 않는 부모 객체입니다.")

    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)

# ==============================================================================
# [Part 2] 유틸리티 및 텍스트 처리 함수
# ==============================================================================

def clean_empty_parentheses(text):
    if not text: return text
    return re.sub(r"\(\s*\)", "", text).strip()

def clean_header_text(text):
    text = text.strip()
    # [수정] 동그라미 숫자 대응
    match = re.search(r"([①-⑩]|\d+)", text)
    if match:
        raw_code = match.group(1)
        code = CIRCLE_MAP.get(raw_code, raw_code)
        label = re.sub(r"[\(\[\{\<]?\s*" + re.escape(raw_code) + r"\s*[\)\]\}\>]?[\.]?", "", text).strip()
        if not label: label = f"{code}점"
        return f"{code}={label}"
    return f"{text}={text}"

def extract_options_from_line(text):
    # [수정] 동그라미 숫자 포함 패턴
    pattern = re.compile(r"([①-⑩]|(?:\d+|[a-zA-Z])[\)\.])")
    matches = list(pattern.finditer(text))
    if not matches:
        return []
    results = []
    for i in range(len(matches)):
        start = matches[i].start()
        end = matches[i+1].start() if i + 1 < len(matches) else len(text)
        item = text[start:end].strip()
        item = clean_empty_parentheses(item)
        if item:
            results.append(item)
    return results

def summarize_label_regex(text):
    if not text: return ""
    text = re.sub(r"\(PROG.*?\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\[PROG.*?\]", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\(.*?(입력|기입|범위|선택).*?\)", "", text)
    text = re.sub(r"\[.*?(선택|기입|응답).*?\]", "", text)
    text = re.sub(r"^다음은.*?질문입니다\.?", "", text).strip()
    text = re.sub(r"^다음.*?대해.*?(선택|응답).*?주십시오\.?", "", text).strip()
    text = text.replace("귀하의 ", "").replace("귀하께서는 ", "").replace("귀 댁의 ", "")
    text = text.replace("응답자 본인의 ", "").replace("평소 ", "")
    patterns = [
        r"은 무엇입니까\?*$", r"는 무엇입니까\?*$", r"는 무엇인가요\?*$",
        r"을 선택해 주십시오\.?$", r"를 선택해 주십시오\.?$",
        r"을 선택해 주세요\.?$", r"를 선택해 주세요\.?$",
        r"을 기입해 주십시오\.?$", r"를 기입해 주십시오\.?$",
        r"을 입력하여 주십시오\.?$", r"를 입력하여 주십시오\.?$",
        r"에 대해 어떻게 생각하십니까\?*$",
        r"정도입니까\?*$", r"되십니까\?*$", r"인가요\?*$", r"있습니까\?*$"
    ]
    for pat in patterns: text = re.sub(pat, "", text)
    replacements = { "만족하는 정도": "만족도", "얼마나 만족하십니까": "만족도", "얼마나 자주": "빈도", "이유는 무엇": "이유", "생각나는 이미지": "이미지", "구입한 적이": "구입 경험", "이용한 경험": "이용 경험", "어디입니까": "장소", "누구입니까": "대상" }
    for old, new in replacements.items(): 
        if old in text: text = text.replace(old, new)
    text = text.strip(); text = re.sub(r"\?+$", "", text); text = re.sub(r"\.$", "", text)
    return text.strip()

def check_section_header(text, current_prefix):
    clean_text = text.strip()
    new_prefix = current_prefix
    if re.search(r"Screening", clean_text, re.IGNORECASE) or "스크리닝" in clean_text:
        new_prefix = "SQ"
    elif re.search(r"Part\s*([A-Z])", clean_text, re.IGNORECASE):
        match = re.search(r"Part\s*([A-Z])", clean_text, re.IGNORECASE)
        new_prefix = match.group(1).upper()
    elif re.search(r"^DQ", clean_text, re.IGNORECASE) or "인구 통계" in clean_text:
        new_prefix = "DQ"
    return new_prefix

# ==============================================================================
# [Part 3] 테이블 추출기 (Extractors)
# ==============================================================================

# [NEW] AHP 이원비교 테이블 전용 추출기 (Q11 등 대응)
def extract_ahp_table(table, current_var):
    rows = table.rows
    if len(rows) < 2: return None
    
    # 헤더 분석: A와 B가 있고 척도(9, 7, 5...)가 있는지 확인
    header_text = " ".join([c.text for c in rows[0].cells])
    if not ("A" in header_text and "B" in header_text and ("중요" in header_text or "9" in header_text)):
        return None

    # AHP 9점 척도 정의
    ahp_scale_pairs = [
        "1=A 절대 중요(9)", "2=A 매우 중요(7)", "3=A 상당히 중요(5)", "4=A 약간 중요(3)", 
        "5=A와 B 동등(1)", 
        "6=B 약간 중요(3)", "7=B 상당히 중요(5)", "8=B 매우 중요(7)", "9=B 절대 중요(9)"
    ]
    scale_str = "\n".join(ahp_scale_pairs)
    
    extracted_entries = []
    
    for i, row in enumerate(rows[1:]):
        cells = row.cells
        if len(cells) < 3: continue
        
        # 좌측 항목(A)과 우측 항목(B) 추출
        # 병합된 셀이나 빈 셀을 건너뛰고 텍스트가 있는 첫/마지막 셀 찾기
        item_a = cells[0].text.strip()
        item_b = cells[-1].text.strip()
        
        # 중간에 숫자가 없거나 A, B가 비어있으면 유효한 행이 아님
        if not item_a or not item_b or item_a == item_b: 
            continue
            
        var_name = f"{current_var['변수명']}_{i+1}"
        label = f"[{current_var['변수명']}] {item_a} vs {item_b}"
        
        extracted_entries.append({
            "변수명": var_name,
            "질문 내용": label,
            "보기 값": scale_str,
            "유형": "Scale"
        })
        
    return extracted_entries

def check_mixed_text_input(entry):
    if entry["유형"] != "Single" and entry["유형"] != "Open": return [entry]
    full_text = entry["질문 내용"]
    if "보기_list" in entry: full_text += " " + " ".join(entry["보기_list"])
    pattern = re.compile(r"\([^)]*?입력[^)]*?\)\s*([가-힣a-zA-Z]+)")
    matches = list(pattern.finditer(full_text))
    if len(matches) < 2: return [entry]
    new_entries = []
    base_var = entry["변수명"]; base_label = entry["질문 내용"]
    clean_base = re.sub(r"\([^)]*?입력[^)]*?\)\s*[가-힣a-zA-Z]*", "", base_label).strip()
    for i, match in enumerate(matches):
        unit = match.group(1)
        new_entries.append({ "변수명": f"{base_var}_{i+1}", "질문 내용": f"[{base_var}] {clean_base} ({unit})", "보기 값": "(숫자입력)", "유형": "Open" })
    return new_entries

def extract_embedded_open_entry(entry):
    if entry["유형"] not in ["Single", "Multi"]: return []
    vals_str = entry.get("보기 값", "")
    if not vals_str: return []
    new_entries = []
    lines = vals_str.split('\n')
    normalized_lines = [line.replace("（", "(").replace("）", ")").replace("[", "(").replace("]", ")") for line in lines]
    for line in normalized_lines:
        if "=" not in line: continue
        parts = line.split("=", 1)
        code = parts[0].strip(); label = parts[1].strip()
        if "(" in label and ")" in label:
            paren_content_match = re.search(r"\(([^)]+)\)", label)
            if paren_content_match:
                content = paren_content_match.group(1)
                if any(k in content for k in ["입력", "기입", "범위", "구체적", "작성"]):
                    unit = ""
                    suffix_match = re.search(r"\)[^)]*$", label)
                    if suffix_match:
                        suffix = suffix_match.group(0).replace(")", "").strip()
                        if suffix: unit = f" ({suffix})"
                    new_entries.append({
                        "변수명": f"{entry['변수명']}_{code}",
                        "질문 내용": f"[{entry['변수명']}] {code}번 선택 시 구체적 내용{unit}",
                        "보기 값": "(숫자입력)" if "범위" in content or "수" in content or "명" in suffix else "(주관식)",
                        "유형": "Open"
                    })
    return new_entries

# 변수 매핑 테이블 (SQ8, SQ8-1, SQ10-1 등)
def extract_mapped_option_table(table, extracted_data, variable_map, current_entry):
    rows = table.rows
    if len(rows) < 2: return None
    header_cells = [c.text.strip() for c in rows[0].cells]
    
    option_col_idx = -1
    for i, h in enumerate(header_cells):
        if "보기" in h: option_col_idx = i; break
    if option_col_idx == -1: return None
    
    multi_keywords = ["복수응답", "모두 선택", "중복선택", "중복 응답", "모두 골라", "중복 선택", "복수 선택", "모두 체크"]

    target_vars = {} 
    existing_vars = list(variable_map.keys())
    current_var_name = current_entry["변수명"] if current_entry else None
    if current_var_name: existing_vars.append(current_var_name)
    
    def normalize_name(n): return re.sub(r"[^a-zA-Z0-9]", "", n).upper()

    for i, h in enumerate(header_cells):
        if i == option_col_idx: continue
        norm_h = normalize_name(h)
        if not norm_h: continue
        for var_name in existing_vars:
            norm_v = normalize_name(var_name)
            if norm_h == norm_v or (len(norm_h) > 2 and norm_h in norm_v):
                target_vars[i] = var_name
                break
                
    if not target_vars: return None
    
    var_options_map = {v: [] for v in target_vars.values()} 
    
    for row in rows[1:]:
        if len(row.cells) <= option_col_idx: continue
        opt_text = row.cells[option_col_idx].text.strip()
        if not opt_text: continue
        
        code = ""; val = ""
        # [수정] 동그라미 숫자 대응
        match = re.match(r"^([①-⑩]|\d+|[a-zA-Z])[\)\.]?\s*(.*)", opt_text)
        if match:
            raw = match.group(1).replace(')','').replace('.','') 
            code = CIRCLE_MAP.get(raw, raw)
            val = match.group(2).strip()
        else: val = opt_text
            
        for col_idx, var_name in target_vars.items():
            if len(row.cells) > col_idx:
                check_val = row.cells[col_idx].text.strip()
                if check_val:
                    final_code = check_val if check_val.isdigit() else code
                    if final_code: var_options_map[var_name].append((final_code, val))

    updates = 0
    vars_to_process = [v for v in var_options_map.keys() if v in variable_map]
    vars_to_process.sort(key=lambda x: variable_map[x], reverse=True) 
    
    for var_name in vars_to_process:
        opts_tuples = var_options_map[var_name] 
        if not opts_tuples: continue
        idx = variable_map[var_name]
        original_item = extracted_data[idx]
        is_multi = any(k in original_item["질문 내용"] for k in multi_keywords)
        
        if is_multi:
            new_items = []
            full_opts_str = "\n".join([f"{c}={l}" for c, l in opts_tuples])
            for c, l in opts_tuples:
                new_items.append({ "변수명": f"{var_name}_{c}", "질문 내용": f"{original_item['질문 내용']} ({l})", "보기 값": full_opts_str, "유형": "Multi" })
            del extracted_data[idx]
            for item in reversed(new_items): extracted_data.insert(idx, item)
            updates += 1
        else:
            opts_str = "\n".join([f"{c}={l}" for c, l in opts_tuples])
            extracted_data[idx]["보기 값"] = opts_str
            updates += 1

    if current_entry and current_entry["변수명"] in var_options_map:
        opts_tuples = var_options_map[current_entry["변수명"]]
        if opts_tuples:
            if "보기_list" not in current_entry: current_entry["보기_list"] = []
            opts_str = "\n".join([f"{c}={l}" for c, l in opts_tuples])
            current_entry["보기 값"] = opts_str
            for c, l in opts_tuples: current_entry["보기_list"].append(f"{c}) {l}")
            updates += 1
                
    if updates > 0:
        new_map = {}
        for i, item in enumerate(extracted_data): new_map[item['변수명']] = i
        variable_map.clear(); variable_map.update(new_map)
    return updates > 0

def extract_unit_input_table(table, current_var):
    extracted = []
    unit_keywords = ["명", "세", "개", "원", "년", "월"]
    unit_col_idx = -1
    for i, cell in enumerate(table.rows[0].cells):
        if any(u in cell.text for u in unit_keywords): unit_col_idx = i; break
    if unit_col_idx == -1 and len(table.rows) > 1:
         for i, cell in enumerate(table.rows[-1].cells):
            if any(u in cell.text for u in unit_keywords): unit_col_idx = i; break
    label_col_idx = 0
    if len(table.columns) > 1:
        if unit_col_idx == 1: label_col_idx = 0
        else: label_col_idx = 1
    for i, row in enumerate(table.rows):
        cells = row.cells
        if len(cells) <= label_col_idx: continue
        row_label = cells[label_col_idx].text.strip()
        if row_label.isdigit() and len(cells) > label_col_idx + 1: row_label = cells[label_col_idx + 1].text.strip()
        if not row_label or "입력" in row_label: continue
        unit = ""
        if unit_col_idx != -1 and len(cells) > unit_col_idx:
            unit_text = cells[unit_col_idx].text.strip()
            if unit_text in unit_keywords: unit = f" ({unit_text})"
        extracted.append({ "변수명": f"{current_var['변수명']}_{i+1}", "질문 내용": f"[{current_var['변수명']}] {row_label}{unit}", "보기 값": "(숫자입력)", "유형": "Open" })
    return extracted

def extract_child_demographics_table(table, current_var):
    headers = [c.text.strip() for c in table.rows[0].cells]
    gender_col_idx = -1; birth_col_idx = -1
    for idx, h in enumerate(headers):
        if "성별" in h: gender_col_idx = idx
        if "생년" in h or "생일" in h or "생월" in h: birth_col_idx = idx
    if gender_col_idx == -1 or birth_col_idx == -1: return None 
    extracted_entries = []
    for i, row in enumerate(table.rows[1:]): 
        cells = row.cells
        if len(cells) <= max(gender_col_idx, birth_col_idx): continue
        row_label = cells[0].text.strip(); gender_text = cells[gender_col_idx].text.strip(); birth_text = cells[birth_col_idx].text.strip()
        if not row_label: continue 
        gender_opts = extract_options_from_line(gender_text); gender_vals_str = ""
        if gender_opts:
            g_lines = []
            for opt in gender_opts:
                # [수정] 동그라미 숫자 대응
                m = re.match(r"^([①-⑩]|\d+|[a-zA-Z])[\)\.]?\s*(.*)", opt)
                if m: 
                    code = CIRCLE_MAP.get(m.group(1), m.group(1).replace(')','').replace('.',''))
                    g_lines.append(f"{code}={m.group(2).strip()}")
                else: g_lines.append(opt)
            gender_vals_str = "\n".join(g_lines)
        extracted_entries.append({ "변수명": f"{current_var['변수명']}_{i+1}_1", "질문 내용": f"[{current_var['변수명']}] {row_label} - 성별", "보기 값": gender_vals_str, "유형": "Single" })
        has_year = "년" in birth_text; has_month = "월" in birth_text
        if has_year: extracted_entries.append({ "변수명": f"{current_var['변수명']}_{i+1}_2", "질문 내용": f"[{current_var['변수명']}] {row_label} - 생년 (년)", "보기 값": "(숫자입력)", "유형": "Open" })
        if has_month: extracted_entries.append({ "변수명": f"{current_var['변수명']}_{i+1}_3", "질문 내용": f"[{current_var['변수명']}] {row_label} - 생월 (월)", "보기 값": "(숫자입력)", "유형": "Open" })
    return extracted_entries

def extract_time_split_table(table, current_var):
    extracted = []
    for i, row in enumerate(table.rows):
        cells_text = [c.text.strip() for c in row.cells if c.text.strip()]
        if not cells_text: continue
        row_full_text = " ".join(cells_text)
        is_header_row = ("시간" in row_full_text and "분" in row_full_text and "입력" not in row_full_text and "범위" not in row_full_text and "(" not in row_full_text)
        if is_header_row: continue
        row_label = cells_text[0]
        clean_label = re.sub(r"※.*", "", row_label).strip().replace(":", "").strip()
        if len(clean_label) > 40 or not clean_label: continue
        extracted.append({ "변수명": f"{current_var['변수명']}_{i+1}_H", "질문 내용": f"[{current_var['변수명']}] {clean_label} (시간)", "보기 값": "(숫자입력)", "유형": "Open" })
        extracted.append({ "변수명": f"{current_var['변수명']}_{i+1}_M", "질문 내용": f"[{current_var['변수명']}] {clean_label} (분)", "보기 값": "(숫자입력)", "유형": "Open" })
    return extracted

def extract_horizontal_scale_table(table, current_var):
    rows = table.rows
    if len(rows) < 2: return None
    
    numeric_row_idx = -1
    label_row_idx = -1
    
    for i, row in enumerate(rows):
        cells_text = [c.text.strip() for c in row.cells if c.text.strip()]
        if not cells_text: continue
        # [수정] 동그라미 숫자 감지 강화
        numeric_count = sum(1 for t in cells_text if t.isdigit() or t in CIRCLE_MAP)
        if len(cells_text) > 0 and (numeric_count / len(cells_text)) > 0.7:
            numeric_row_idx = i
        elif len(cells_text) > 0:
            label_row_idx = i
            
    if numeric_row_idx == -1: return None
    
    codes = []
    for c in rows[numeric_row_idx].cells:
        t = c.text.strip()
        if not t: continue
        # [수정] 동그라미 숫자를 아라비아 숫자로 변환
        codes.append(CIRCLE_MAP.get(t, t))

    labels = []
    if label_row_idx != -1:
        labels = [c.text.strip() for c in rows[label_row_idx].cells if c.text.strip()]
    
    scale_pairs = []
    
    # 모든 코드를 살리되, 라벨이 부족하면 양극단 매핑
    if codes:
        if len(labels) == 2: # 양극단
            scale_pairs.append(f"{codes[0]}={labels[0]}")
            for c in codes[1:-1]: scale_pairs.append(f"{c}={c}점")
            scale_pairs.append(f"{codes[-1]}={labels[1]}")
        elif len(labels) == len(codes): # 1:1 매핑
             for i in range(len(codes)): scale_pairs.append(f"{codes[i]}={labels[i]}")
        else: # 매핑 애매하면 그냥 순서대로 넣고 나머진 점수
             for i, c in enumerate(codes):
                 if i < len(labels): scale_pairs.append(f"{c}={labels[i]}")
                 else: scale_pairs.append(f"{c}={c}점")

    if scale_pairs:
        current_var["보기 값"] = "\n".join(scale_pairs)
        return [current_var]
    return None

def extract_horizontal_input_table(table, current_var):
    rows = table.rows
    if len(rows) < 2: return None
    extracted = []
    headers = rows[0].cells
    values = rows[1].cells
    for i in range(len(headers)):
        header_text = headers[i].text.strip()
        value_text = values[i].text.strip()
        if not header_text: continue
        clean_label = clean_empty_parentheses(header_text)
        if "시간" in value_text and "분" in value_text and ("입력" in value_text or "(" in value_text):
             extracted.append({ "변수명": f"{current_var['변수명']}_{i+1}_H", "질문 내용": f"[{current_var['변수명']}] {clean_label} (시간)", "보기 값": "(숫자입력)", "유형": "Open" })
             extracted.append({ "변수명": f"{current_var['변수명']}_{i+1}_M", "질문 내용": f"[{current_var['변수명']}] {clean_label} (분)", "보기 값": "(숫자입력)", "유형": "Open" })
        else:
            extracted.append({ "변수명": f"{current_var['변수명']}_{i+1}", "질문 내용": f"[{current_var['변수명']}] {clean_label}", "보기 값": "(숫자입력)", "유형": "Open" })
    return extracted

def extract_plain_input_table(table, current_var):
    extracted = []
    for i, row in enumerate(table.rows):
        cells_text = [c.text.strip() for c in row.cells if c.text.strip()]
        if not cells_text: continue
        row_full_text = " ".join(cells_text)
        # [수정] 동그라미 숫자 패턴 회피
        if re.search(r"([①-⑩]|\d+|[a-zA-Z])[\)\.]", row_full_text): continue
        clean_label = re.sub(r"\(\s*입력.*?\)", "", row_full_text).replace(":", "").strip()
        clean_label = re.sub(r"[a-zA-Z]+$", "", clean_label).strip()
        if not clean_label: continue
        extracted.append({ "변수명": f"{current_var['변수명']}_{i+1}", "질문 내용": f"[{current_var['변수명']}] {clean_label}", "보기 값": "(숫자입력)", "유형": "Open" })
    return extracted

def extract_constant_sum_table(table, current_var):
    extracted_entries = []
    q_text = current_var.get("질문 내용", "")
    for i, row in enumerate(table.rows):
        cells = row.cells
        if len(cells) < 2: continue
        label_cell = cells[0].text.strip(); input_cell = cells[1].text.strip()
        if not label_cell: continue
        if "합계" in label_cell or "Total" in label_cell or "TOTAL" in label_cell: continue
        sub_var_name = f"{current_var['변수명']}_{i+1}"
        final_label = f"[{current_var['변수명']}] {label_cell}"
        if "%" in input_cell or "퍼센트" in q_text: final_label += " (%)"
        extracted_entries.append({ "변수명": sub_var_name, "질문 내용": final_label, "보기 값": "(숫자입력)", "유형": "Open" })
    return extracted_entries

def is_multiple_choice(entry):
    vals = str(entry.get("보기 값", "")); q_text = str(entry.get("질문 내용", ""))
    if re.search(r"([①-⑩]|\d+[\)\.])", vals) or "=" in vals: return True
    if "선택]" in q_text: return True
    return False

def check_and_split_time(entry):
    if is_multiple_choice(entry): return [entry]
    val = str(entry.get("보기 값", "")) + str(entry.get("질문 내용", ""))
    is_time_related = ("시간" in val or "시" in val or "분" in val) and ("입력" in val or "기입" in val)
    if not is_time_related: return [entry]
    has_hour_unit = bool(re.search(r"(\)|\]|\}|_)\s*시간", val) or re.search(r"시간\s*(\(|\[|\{|_)", val))
    has_minute_unit = bool(re.search(r"(\)|\]|\}|_)\s*분", val) or re.search(r"분\s*(\(|\[|\{|_)", val))
    if has_hour_unit and has_minute_unit:
        entry_h = entry.copy(); entry_h["변수명"] += "_H"; entry_h["질문 내용"] += " (시간)"; entry_h["유형"] = "Open"
        entry_m = entry.copy(); entry_m["변수명"] += "_M"; entry_m["질문 내용"] += " (분)"; entry_m["유형"] = "Open"
        return [entry_h, entry_m]
    elif has_hour_unit:
        entry_h = entry.copy(); entry_h["변수명"] += "_H"; entry_h["질문 내용"] += " (시간)"; entry_h["유형"] = "Open"
        return [entry_h]
    elif has_minute_unit:
        entry_m = entry.copy(); entry_m["변수명"] += "_M"; entry_m["질문 내용"] += " (분)"; entry_m["유형"] = "Open"
        return [entry_m]
    if "분" in val and "시간" in val:
        entry_m = entry.copy(); entry_m["변수명"] += "_M"; entry_m["질문 내용"] += " (분)"; entry_m["유형"] = "Open"
        return [entry_m]
    return [entry]

def check_and_split_date(entry):
    if is_multiple_choice(entry): return [entry]
    val = str(entry.get("보기 값", "")) + str(entry.get("질문 내용", ""))
    if "억" in val: return [entry]
    if re.search(r"(몇\s*명|명\s*수|인원|\(\s*\)\s*명|\[\s*\]\s*명)", val): return [entry]
    def has_unit(text, unit):
        p1 = re.search(r"(\)|\]|\}|_)\s*" + unit, text); p2 = re.search(unit + r"\s*(\(|\[|\{|_)", text)
        p3 = (unit in text) and ("입력" in text or "기입" in text); return bool(p1 or p2 or p3)
    has_year = has_unit(val, "년"); has_month = has_unit(val, "월") or has_unit(val, "개월"); has_day = has_unit(val, "일")
    if not (has_year or has_month or has_day): return [entry]
    new_entries = []
    if has_year: y = entry.copy(); y["변수명"] += "_Y"; y["질문 내용"] += " (년)"; y["유형"] = "Open"; new_entries.append(y)
    if has_month: m = entry.copy(); m["변수명"] += "_M"; m["질문 내용"] += " (월)"; m["유형"] = "Open"; new_entries.append(m)
    if has_day: d = entry.copy(); d["변수명"] += "_D"; d["질문 내용"] += " (일)"; d["유형"] = "Open"; new_entries.append(d)
    if new_entries: return new_entries
    return [entry]

def check_and_split_money(entry):
    if is_multiple_choice(entry): return [entry]
    val = str(entry.get("보기 값", "")) + str(entry.get("질문 내용", ""))
    val_clean = val.replace(" ", "")
    if "만원" not in val_clean and "만 원" not in val: return [entry]
    new_entries = []
    if "억" in val_clean: e = entry.copy(); e["변수명"] += "_E"; e["질문 내용"] += " (억)"; e["유형"] = "Open"; new_entries.append(e)
    if "천" in val_clean: c = entry.copy(); c["변수명"] += "_C"; c["질문 내용"] += " (천)"; c["유형"] = "Open"; new_entries.append(c)
    if "백" in val_clean: b = entry.copy(); b["변수명"] += "_B"; b["질문 내용"] += " (백)"; b["유형"] = "Open"; new_entries.append(b)
    if new_entries: return new_entries
    return [entry]

def check_and_split_percent(entry):
    val = str(entry.get("보기 값", "")) + str(entry.get("질문 내용", ""))
    if "나" in val and "배우자" in val and ("%" in val or "100" in val):
        entry_me = entry.copy(); entry_me["변수명"] += "_1"; entry_me["질문 내용"] += " (나)"; entry_me["유형"] = "Open"
        entry_sp = entry.copy(); entry_sp["변수명"] += "_2"; entry_sp["질문 내용"] += " (배우자)"; entry_sp["유형"] = "Open"
        entry_sum = entry.copy(); entry_sum["변수명"] += "_3"; entry_sum["질문 내용"] += " (합계)"; entry_sum["유형"] = "Open"
        return [entry_me, entry_sp, entry_sum]
    return [entry]

def collapse_consecutive_duplicates(item_list):
    if not item_list: return []
    collapsed = [item_list[0]]
    for item in item_list[1:]:
        if item != collapsed[-1]: collapsed.append(item)
    return collapsed

def extract_double_scale_table(table, current_var):
    rows = table.rows
    if len(rows) < 3: return None
    raw_cat_cells = [c.text.strip() for c in rows[0].cells]; non_empty_cats = [c for c in raw_cat_cells if c]
    if len(non_empty_cats) < 2: return None 
    categories = collapse_consecutive_duplicates(non_empty_cats)
    if len(categories) != 2: return None
    scale_row_cells = [c.text.strip() for c in rows[1].cells]; scales = scale_row_cells[1:]
    if len(scales) % 2 != 0: return None
    mid = len(scales) // 2
    left_scale = scales[:mid]; right_scale = scales[mid:]
    left_norm = "".join(left_scale).replace(" ", ""); right_norm = "".join(right_scale).replace(" ", "")
    if left_norm != right_norm: return None
    scale_pairs = []
    for idx, txt in enumerate(left_scale):
        if txt: scale_pairs.append(f"{idx+1}={txt}")
    scale_str = "\n".join(scale_pairs)
    cat1_label = categories[0]; cat2_label = categories[1]
    extracted_entries = []
    for r_idx, row in enumerate(rows[2:]):
        cells = row.cells
        if not cells: continue
        q_text = cells[0].text.strip()
        if not q_text: continue
        q_text_clean = re.sub(r"^[\d\w]+[\)\.]\s*", "", q_text)
        var_base = f"{current_var['변수명']}_{r_idx+1}"
        entry1 = { "변수명": f"{var_base}_1", "질문 내용": f"[{cat1_label}] {q_text_clean}", "보기 값": scale_str, "유형": "Scale" }
        entry2 = { "변수명": f"{var_base}_2", "질문 내용": f"[{cat2_label}] {q_text_clean}", "보기 값": scale_str, "유형": "Scale" }
        extracted_entries.append(entry1); extracted_entries.append(entry2)
    return extracted_entries

def extract_table_scale(table):
    rows = table.rows
    if len(rows) < 2: return None, False
    headers = [cell.text.strip().replace('\n', ' ') for cell in rows[0].cells]
    first_data_row = [cell.text.strip() for cell in rows[1].cells]
    
    numeric_cells = []
    for cell_text in first_data_row:
        if "입력" in cell_text or "범위" in cell_text or "%" in cell_text: numeric_cells.append(None); continue
        
        # [수정] 동그라미 숫자 대응 (B1-B4 매트릭스 등)
        match = re.search(r"([①-⑩]|\d+)", cell_text)
        if match: 
            raw_code = match.group(1)
            numeric_cells.append(CIRCLE_MAP.get(raw_code, raw_code))
        else: numeric_cells.append(None)
            
    body_numeric_count = sum(1 for x in numeric_cells if x is not None)
    if len(first_data_row) > 0 and (body_numeric_count / len(first_data_row)) >= 0.3:
        scale_pairs = []
        for i, val in enumerate(numeric_cells):
            if i >= len(headers): break
            if val is not None and headers[i]: 
                scale_pairs.append(f"{val}={headers[i].strip()}")
        if scale_pairs: return "\n".join(scale_pairs), True

    # 헤더에 숫자가 있는 경우 (기존 로직)
    potential_values = []
    header_numeric_count = sum(1 for h in headers if re.search(r"(\d)", h))
    if len(headers) > 0 and (header_numeric_count / len(headers)) >= 0.3:
        for idx, h_text in enumerate(headers):
            if not h_text: continue
            if idx == 0 and not re.search(r"\d", h_text): continue
            potential_values.append(clean_header_text(h_text))
        if potential_values: return "\n".join(potential_values), False
    return None, False

def is_input_table(table):
    if len(table.rows) < 1: return False
    target_count = 0; total_rows = len(table.rows)
    for row in table.rows:
        if len(row.cells) > 1:
            cell_text = row.cells[1].text
            if "입력" in cell_text or "(" in cell_text or "%" in cell_text or "_" in cell_text: target_count += 1
    if total_rows > 0 and (target_count / total_rows) >= 0.3: return True
    return False

def extract_multi_column_input_table(table, current_var, force_row_count=None):
    rows = table.rows
    if len(rows) < 2: return None
    headers = [cell.text.strip() for cell in rows[0].cells]
    non_empty_headers = [h for h in headers if h]
    if len(non_empty_headers) < 1: return None
    first_data_row_cells = [c.text.strip() for c in rows[1].cells[1:]] 
    # [수정] 동그라미 숫자 대응
    digit_count = sum(1 for c in first_data_row_cells if (c.isdigit() or c in CIRCLE_MAP) and len(c) == 1)
    if len(first_data_row_cells) > 0 and (digit_count / len(first_data_row_cells)) > 0.5: return None
    extracted_entries = []
    actual_data_rows = len(rows) - 1
    target_loop_count = actual_data_rows
    if force_row_count and force_row_count > actual_data_rows: target_loop_count = force_row_count
    sub_item_count = 0
    for i in range(target_loop_count):
        sub_item_count += 1
        if i < actual_data_rows:
            curr_row = rows[i+1]
            first_cell = curr_row.cells[0].text.strip()
            row_label = first_cell if first_cell else f"{sub_item_count}순위"
        else: row_label = f"{sub_item_count}순위"
        for c_idx in range(len(headers)):
            if c_idx == 0: continue
            raw_header = headers[c_idx] if c_idx < len(headers) else ""
            col_header = raw_header if raw_header else f"Col{c_idx}"
            var_name = f"{current_var['변수명']}_{sub_item_count}_{c_idx}"
            var_label = f"[{current_var['변수명']}] {row_label} - {col_header}"
            extracted_entries.append({ "변수명": var_name, "질문 내용": var_label, "보기 값": "(주관식)", "유형": "Open" })
    return extracted_entries

def check_and_split_max_n_text(entry):
    if entry["유형"] != "Single" and entry["유형"] != "Open": return None
    q_text = entry["질문 내용"]
    if "보기_list" in entry: q_text += " " + " ".join(entry["보기_list"])
    q_text_norm = q_text.replace("［", "[").replace("］", "]").replace("（", "(").replace("）", ")")
    count = 0
    patterns = [ r"\[\s*최대\s*(\d+)", r"최대\s*(\d+)\s*(?:개|대|곳|명|순위)", r"최대.*?(\d+)", r"(\d+)개.*?기입" ]
    for pat in patterns:
        match = re.search(pat, q_text_norm)
        if match: count = int(match.group(1)); break
    if count == 0 and "3" in q_text_norm and ("기입" in q_text_norm or "작성" in q_text_norm): count = 3
    if count < 1: return None
    has_manufacturer = "제조사" in q_text_norm; has_brand = "브랜드" in q_text_norm
    new_entries = []
    for i in range(1, count + 1):
        if has_manufacturer and has_brand:
            v1 = entry.copy(); v1["변수명"] = f"{entry['변수명']}_{i}_1"; v1["질문 내용"] = f"[{entry['변수명']}] {i}순위 - 제조사"; v1["유형"] = "Open"
            if "보기_list" in v1: del v1["보기_list"]
            v2 = entry.copy(); v2["변수명"] = f"{entry['변수명']}_{i}_2"; v2["질문 내용"] = f"[{entry['변수명']}] {i}순위 - 브랜드"; v2["유형"] = "Open"
            if "보기_list" in v2: del v2["보기_list"]
            new_entries.append(v1); new_entries.append(v2)
        else:
            v = entry.copy(); v["변수명"] = f"{entry['변수명']}_{i}"; v["질문 내용"] = f"[{entry['변수명']}] {i}순위"; v["유형"] = "Open"
            if "보기_list" in v: del v["보기_list"]
            new_entries.append(v)
    return new_entries

def is_option_description_table(table):
    if len(table.rows) < 1: return False
    # [수정] 동그라미 숫자 대응
    pattern = re.compile(r"^([①-⑩]|\d+|[a-zA-Z])[\)\.]")
    match_count = 0
    for row in table.rows:
        if not row.cells: continue
        text = row.cells[0].text.strip()
        if pattern.match(text): match_count += 1
    return (match_count / len(table.rows)) >= 0.5

def extract_single_choice_options(table):
    options = []
    for row in table.rows:
        cells_text = [c.text.strip() for c in row.cells if c.text.strip()]
        if not cells_text: continue
        first_cell_text = cells_text[0]
        # [수정] 동그라미 숫자 대응
        match = re.match(r"^([①-⑩]|\d+|[a-zA-Z])[\)\.]", first_cell_text)
        if match:
            raw_code = match.group(1).replace(')','').replace('.','')
            code = CIRCLE_MAP.get(raw_code, raw_code)
            clean_first = first_cell_text[len(match.group(0)):].strip()
            label_parts = []
            if clean_first: label_parts.append(clean_first)
            if len(cells_text) > 1: label_parts.extend(cells_text[1:])
            final_label = " - ".join(label_parts); final_label = clean_empty_parentheses(final_label) 
            options.append(f"{code}={final_label}")
        else:
            row_text = " - ".join(cells_text); row_text = clean_empty_parentheses(row_text) 
            options.append(row_text)
    return "\n".join(options)

def extract_options_from_table(table):
    options = []
    idx = 1
    for row in table.rows:
        for cell in row.cells:
            text = cell.text.strip(); text = clean_empty_parentheses(text)
            if text: options.append(f"{idx}={text}"); idx += 1
    return "\n".join(options)

def check_ranking_selection_question(entry):
    q_text = entry["질문 내용"]
    if ("순서" in q_text or "순위" in q_text) and "선택" in q_text:
        match_rank = re.search(r"~\s*(\d+)\s*순위", q_text)
        if match_rank: return int(match_rank.group(1))
        match_count = re.search(r"(\d+)개", q_text)
        if match_count: return int(match_count.group(1))
    return None

# ==============================================================================
# [Part 4] 지능형 테이블 분석 (Scanning)
# ==============================================================================

def analyze_table_structure(table):
    rows = table.rows
    if len(rows) < 1: return "UNKNOWN"
    all_text = ""; first_row_text = ""; second_row_text = ""; has_input_pattern = False
    input_keywords = ["입력", "범위", "cm", "kg", "명", "개", "회", "( )", "()"]
    
    row0_digits = 0; row0_len = 0
    row1_digits = 0; row1_len = 0
    
    # [NEW] 보기 목록형 테이블 감지 (SQ8)
    if "보기" in [c.text.strip() for c in rows[0].cells]:
        return "MAPPED_OPTION"
    
    # [NEW] 단위 입력형 테이블 감지 (SQ6)
    unit_keywords = ["명", "세", "개", "원", "년"]
    has_unit_col = False
    for row in rows:
        if any(cell.text.strip() in unit_keywords for cell in row.cells):
            has_unit_col = True
            break
    if has_unit_col: return "UNIT_INPUT"
    
    # [FIX] 수평 척도형 테이블 감지 로직 강화 (B1-1)
    has_numeric_row = False
    for row in rows:
        cells = [c.text.strip() for c in row.cells if c.text.strip()]
        if len(cells) >= 5: # 최소 5점 척도 이상
            # [수정] 동그라미 숫자 감지 강화
            digit_count = sum(1 for c in cells if c.isdigit() or c in CIRCLE_MAP)
            if digit_count / len(cells) > 0.8: 
                has_numeric_row = True
                break
    if has_numeric_row: return "HORIZONTAL_SCALE"

    for i, row in enumerate(rows):
        row_txt = " ".join([c.text.strip() for c in row.cells])
        all_text += row_txt + " "; 
        if i == 0: 
            first_row_text = row_txt
            row0_len = len(row.cells)
            # [수정] 동그라미 숫자 감지
            row0_digits = sum(1 for c in row.cells if re.search(r"^([①-⑩]|\d+)[\)\.]?$", c.text.strip()))
        if i == 1: 
            second_row_text = row_txt
            row1_len = len(row.cells)
            # [수정] 동그라미 숫자 감지
            row1_digits = sum(1 for c in row.cells if (c.text.strip().isdigit() or c.text.strip() in CIRCLE_MAP))
            
        if any(k in row_txt for k in input_keywords): has_input_pattern = True

    # 1. [최우선] 매트릭스 척도형 (E1-1 방어용)
    if len(table.columns) >= 4 and row0_digits >= 3 and not has_input_pattern:
        return "STANDARD"

    # 2. 자녀 정보 (SQ6)
    if "성별" in all_text and ("생년" in all_text or "생일" in all_text): return "CHILD_DEMO"
    
    # 3. 시간 분할 (세로형 - A2, A4)
    if "시간" in all_text and "분" in all_text and has_input_pattern:
        if len(table.columns) <= 4:
            return "TIME_SPLIT"

    # 4. 가로형 척도 (B2, A10-1)
    if len(rows) == 2 and not has_input_pattern:
        row0_is_numeric = row0_len > 0 and (row0_digits / row0_len) > 0.5
        row1_is_numeric = row1_len > 0 and (row1_digits / row1_len) > 0.5
        if (row0_is_numeric and not row1_is_numeric) or (not row0_is_numeric and row1_is_numeric):
            return "HORIZONTAL_SCALE"

    # 5. 가로형 입력 (B3, B4)
    is_row1_input = any(k in second_row_text for k in input_keywords)
    if len(rows) >= 2 and len(table.columns) >= 2 and is_row1_input:
        return "HORIZONTAL_INPUT"
    
    # 6. 고정 합계
    if ("합계" in all_text or "Total" in all_text) and ("%" in all_text or "100" in all_text):
        if len(table.columns) == 2: return "CONSTANT_SUM"
        
    # 7. 단순 입력 (A1)
    is_option_table = bool(re.search(r"([①-⑩]|\d+|[a-zA-Z])[\)\.]", first_row_text))
    if has_input_pattern and not is_option_table and len(table.columns) <= 2: return "PLAIN_INPUT"
    
    return "STANDARD"

# ==============================================================================
# [Part 5] 메인 파서
# ==============================================================================

def parse_word_to_df(docx_file):
    doc = Document(docx_file)
    extracted_data = []
    var_pattern = re.compile(r"^([a-zA-Z가-힣0-9\-\_]+)(?:[\.\s]|\s+)(.*)")
    # [SQ10 해결] 띄어쓰기 포함된 키워드 추가
    multi_keywords = ["복수응답", "모두 선택", "중복선택", "중복 응답", "모두 골라", "중복 선택", "복수 선택", "중복가능", "모두 체크", "모두 응답"]
    current_entry = None
    is_parent_added = False 
    
    # [NEW] 섹션 인식 변수
    current_prefix = "Q"
    prefix_counters = collections.defaultdict(int)
    
    # [NEW] 워드 자동번호 인식용 카운터
    auto_num_counters = collections.defaultdict(int)
    
    variable_map = {} 
    
    pending_ranking_count = None
    ranking_options_buffer = []
    pending_max_n_count = None
    
    allowed_starts = ['Q', 'A', 'S', 'D', 'M', 'P', 'R', 'I', 'B', 'C', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'N', 'O', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', '문', '설문']

    def flush_entry(entry):
        nonlocal is_parent_added, pending_max_n_count
        if "질문 내용" in entry: entry["질문 내용"] = clean_empty_parentheses(entry["질문 내용"])
        
        if pending_ranking_count is not None and ranking_options_buffer:
            final_opts_str = "\n".join(ranking_options_buffer)
            results = []
            for i in range(1, pending_ranking_count + 1):
                results.append({ "변수명": f"{entry['변수명']}_{i}", "질문 내용": f"{entry['질문 내용']} ({i}순위)", "보기 값": final_opts_str, "유형": "Ranking_Sel" })
            return results
        if pending_max_n_count is not None:
            # [FIX] 보기가 있는 경우(Mapped Table 등에서 유입), Open이 아니라 Selection으로 처리
            has_options = bool(entry.get("보기 값") or entry.get("보기_list"))
            opts_str = entry.get("보기 값", "")
            if not opts_str and entry.get("보기_list"):
                opts_str = "\n".join(entry["보기_list"])

            new_entries = []
            for i in range(1, pending_max_n_count + 1):
                if has_options:
                    # 보기가 있으면 Ranking_Sel로 변경
                    v = entry.copy()
                    v["변수명"] = f"{entry['변수명']}_{i}"
                    v["질문 내용"] = f"[{entry['변수명']}] {i}순위"
                    v["보기 값"] = opts_str
                    v["유형"] = "Ranking_Sel"
                    if "보기_list" in v: del v["보기_list"]
                    new_entries.append(v)
                else:
                    # 기존 주관식 처리
                    has_manufacturer = "제조사" in entry["질문 내용"]; has_brand = "브랜드" in entry["질문 내용"]
                    if has_manufacturer and has_brand:
                        v1 = entry.copy(); v1["변수명"] = f"{entry['변수명']}_{i}_1"; v1["질문 내용"] = f"[{entry['변수명']}] {i}순위 - 제조사"; v1["유형"] = "Open"
                        if "보기_list" in v1: del v1["보기_list"]
                        v2 = entry.copy(); v2["변수명"] = f"{entry['변수명']}_{i}_2"; v2["질문 내용"] = f"[{entry['변수명']}] {i}순위 - 브랜드"; v2["유형"] = "Open"
                        if "보기_list" in v2: del v2["보기_list"]
                        new_entries.append(v1); new_entries.append(v2)
                    else:
                        v = entry.copy(); v["변수명"] = f"{entry['변수명']}_{i}"; v["질문 내용"] = f"[{entry['변수명']}] {i}순위"; v["유형"] = "Open"
                        if "보기_list" in v: del v["보기_list"]
                        new_entries.append(v)
            pending_max_n_count = None
            return new_entries
        raw_options = entry.get("보기_list", [])
        
        is_multi = any(k in entry["질문 내용"] for k in multi_keywords)
        if "D6_2" in entry["변수명"].replace("-", "_"): is_multi = True
        
        if is_multi and raw_options:
            full_options_str_list = []
            for opt in raw_options:
                # [수정] 동그라미 숫자 대응
                opt_match = re.match(r"^\s*([①-⑩]|\d+|[a-zA-Z])[\)\.]\s*(.*)", opt)
                if opt_match:
                    raw_code = opt_match.group(1).replace(')','').replace('.','')
                    code = CIRCLE_MAP.get(raw_code, raw_code)
                    label = clean_empty_parentheses(opt_match.group(2))
                    full_options_str_list.append(f"{code}={label}")
            full_options_str = "\n".join(full_options_str_list)
            results = []
            for opt in raw_options:
                opt_match = re.match(r"^\s*([①-⑩]|\d+|[a-zA-Z])[\)\.]\s*(.*)", opt)
                if opt_match:
                    raw_code = opt_match.group(1).replace(')','').replace('.','')
                    code = CIRCLE_MAP.get(raw_code, raw_code)
                    label = clean_empty_parentheses(opt_match.group(2))
                    results.append({ "변수명": f"{entry['변수명']}_{code}", "질문 내용": f"{entry['질문 내용']} ({label})", "보기 값": full_options_str, "유형": "Multi" })
            return results
        else:
            # 단일 선택 보기 변환
            clean_opts = []
            for opt in raw_options:
                opt_match = re.match(r"^\s*([①-⑩]|\d+|[a-zA-Z])[\)\.]\s*(.*)", opt)
                if opt_match:
                    raw_code = opt_match.group(1).replace(')','').replace('.','')
                    code = CIRCLE_MAP.get(raw_code, raw_code)
                    clean_opts.append(f"{code}={clean_empty_parentheses(opt_match.group(2))}")
                else: clean_opts.append(opt)

            entry["보기 값"] = "\n".join(clean_opts)
            if "보기_list" in entry: del entry["보기_list"]
            
            mixed_input = check_mixed_text_input(entry)
            if len(mixed_input) > 1: return mixed_input
            
            split_entries = check_and_split_time(entry)
            if len(split_entries) == 1: split_entries = check_and_split_date(split_entries[0])
            if len(split_entries) == 1: split_entries = check_and_split_money(split_entries[0])
            if len(split_entries) == 1: split_entries = check_and_split_percent(split_entries[0])
            
            embedded_opens = extract_embedded_open_entry(split_entries[0])
            if embedded_opens:
                split_entries.extend(embedded_opens)
                
            return split_entries

    for block in iter_block_items(doc):
        # 표 내부 섹션 헤더 감지
        if isinstance(block, Table):
            if len(block.rows) > 0 and len(block.rows[0].cells) > 0:
                first_cell_text = block.rows[0].cells[0].text
                current_prefix = check_section_header(first_cell_text, current_prefix)
        
        if isinstance(block, Paragraph):
            text = block.text.strip()
            current_prefix = check_section_header(text, current_prefix)

            if block._p.pPr is not None and block._p.pPr.numPr is not None:
                try:
                    num_id = block._p.pPr.numPr.numId.val
                    ilvl = block._p.pPr.numPr.ilvl.val if block._p.pPr.numPr.ilvl is not None else 0
                    auto_num_counters[(num_id, ilvl)] += 1
                    num_val = auto_num_counters[(num_id, ilvl)]
                    
                    if not re.match(r"^(\d+|[①-⑩]|[a-zA-Z])[\)\.]", text):
                        if "?" in text or "다." in text or "시오" in text or len(text) > 40:
                            prefix_counters[current_prefix] += 1
                            q_num = prefix_counters[current_prefix]
                            text = f"{current_prefix}{q_num}. {text}"
                        else:
                            text = f"{num_val}) {text}"
                except:
                    pass

            if not text: continue
            if re.match(r"^\[PROG", text, re.IGNORECASE) or re.match(r"^\(PROG", text, re.IGNORECASE): continue
            text = re.sub(r"\[PROG.*?\]", "", text, flags=re.IGNORECASE)
            text = re.sub(r"\(PROG.*?\)", "", text, flags=re.IGNORECASE)
            text = text.strip()
            if not text: continue
            
            match_var = var_pattern.match(text)
            is_new_q = False
            if match_var:
                temp_var = match_var.group(1)
                is_valid_start = False
                for start_char in allowed_starts:
                    if temp_var.upper().startswith(start_char):
                        is_valid_start = True
                        break
                
                if temp_var.replace(".", "").isdigit():
                    if current_entry is None: is_new_code = True
                elif is_valid_start:
                    if temp_var not in ["보기", "다음", "참고", "주"]: is_new_q = True
            
            if is_new_q:
                if current_entry and not is_parent_added:
                    flushed_data = flush_entry(current_entry)
                    if flushed_data: 
                        for item in flushed_data:
                            variable_map[item['변수명']] = len(extracted_data)
                            extracted_data.append(item)
                            
                var_name = match_var.group(1).replace("-", "_"); label = match_var.group(2)
                inline_opts = extract_options_from_line(label)
                if inline_opts:
                    first_opt = inline_opts[0]; split_idx = label.find(first_opt)
                    if split_idx != -1: q_text = label[:split_idx].strip(); current_entry = { "변수명": var_name, "질문 내용": q_text, "보기 값": "", "보기_list": inline_opts, "유형": "Single" }
                    else: current_entry = { "변수명": var_name, "질문 내용": label.strip(), "보기 값": "", "보기_list": [], "유형": "Single" }
                else: current_entry = { "변수명": var_name, "질문 내용": label.strip(), "보기 값": "", "보기_list": [], "유형": "Single" }
                is_parent_added = False
                rank_count = check_ranking_selection_question(current_entry)
                if rank_count: pending_ranking_count = rank_count; ranking_options_buffer = [] 
                else: pending_ranking_count = None; ranking_options_buffer = []
                
                # [FIX] Force Max N check based on text pattern (regardless of function return)
                q_norm = current_entry["질문 내용"].replace("［", "[").replace("］", "]").replace("（", "(").replace("）", ")")
                max_n_match = re.search(r"최대\s*(\d+)", q_norm)
                if max_n_match:
                    pending_max_n_count = int(max_n_match.group(1))
                else:
                    pending_max_n_count = None
                
                if "1개 선택" in current_entry["질문 내용"]: current_entry["유형"] = "Single"
            elif current_entry:
                if not is_parent_added:
                    # [FIX] S5 등 옵션 강제 인식
                    # 숫자로 시작하는 문단은 무조건 보기로 간주 (1) S 2WD 같은 경우)
                    opts_in_line = extract_options_from_line(text)
                    # [수정] 동그라미 숫자 포함 패턴 대응
                    if not opts_in_line and re.match(r"^([①-⑩]|\d+)[\)\.]", text): opts_in_line = [text]

                    if opts_in_line:
                        if pending_ranking_count:
                            for opt in opts_in_line:
                                opt_match = re.match(r"^(\d+|[①-⑩]|[a-zA-Z])[\)\.]\s*(.*)", opt)
                                if opt_match: 
                                    raw_code = opt_match.group(1).replace(')','').replace('.','')
                                    code = CIRCLE_MAP.get(raw_code, raw_code)
                                    val = opt_match.group(2)
                                    ranking_options_buffer.append(f"{code}={val}")
                        else:
                            if "보기_list" in current_entry: current_entry["보기_list"].extend(opts_in_line)
                    elif "=" in text or "점" in text:
                         if "보기_list" in current_entry: current_entry["보기_list"].append(text)
                    elif "[주관식]" in text or "직접 기입" in text:
                        current_entry["유형"] = "Open"
                        if "보기_list" in current_entry: current_entry["보기_list"].append("(주관식)")
                    else:
                        if "보기_list" in current_entry and not current_entry["보기_list"]: current_entry["질문 내용"] += " " + text

        elif isinstance(block, Table):
            rows = block.rows
            if len(rows) < 1: continue

            # 지능형 테이블 분석 (Scanning)
            table_type = analyze_table_structure(block)
            
            new_entries = []
            
            # [NEW] AHP 이원비교 우선 처리
            ahp_entries = extract_ahp_table(block, current_entry)
            if ahp_entries:
                new_entries = ahp_entries
                
            elif table_type == "MAPPED_OPTION":
                is_updated = extract_mapped_option_table(block, extracted_data, variable_map, current_entry)
            
            elif table_type == "UNIT_INPUT":
                if current_entry and not is_parent_added:
                    new_entries = extract_unit_input_table(block, current_entry)

            elif table_type == "CHILD_DEMO":
                if current_entry and not is_parent_added:
                    new_entries = extract_child_demographics_table(block, current_entry)
            
            elif table_type == "HORIZONTAL_SCALE":
                if current_entry and not is_parent_added:
                    new_entries = extract_horizontal_scale_table(block, current_entry)

            elif table_type == "HORIZONTAL_INPUT":
                if current_entry and not is_parent_added:
                    new_entries = extract_horizontal_input_table(block, current_entry)

            elif table_type == "TIME_SPLIT":
                if current_entry and not is_parent_added:
                    new_entries = extract_time_split_table(block, current_entry)
            
            elif table_type == "CONSTANT_SUM":
                if current_entry and not is_parent_added:
                    new_entries = extract_constant_sum_table(block, current_entry)
            
            elif table_type == "PLAIN_INPUT":
                if current_entry and not is_parent_added:
                    new_entries = extract_plain_input_table(block, current_entry)
            
            elif table_type == "STANDARD":
                if current_entry and not is_parent_added:
                    ds = extract_double_scale_table(block, current_entry)
                    if ds: new_entries = ds
                    else:
                        q_type = current_entry.get("유형")
                        if any(k in current_entry["질문 내용"] for k in multi_keywords): q_type = "Multi"
                        if q_type in ["Single", "Multi"]:
                            is_opt = False
                            fc = rows[0].cells[0].text.strip()
                            if re.match(r"^(\d+|[①-⑩]|[a-zA-Z])[\)\.]", fc): is_opt = True
                            if is_opt:
                                opt_str = extract_single_choice_options(block)
                                if q_type == "Single": current_entry["보기 값"] = opt_str; extracted_data.append(current_entry)
                                else:
                                    parsed_opts = []
                                    for line in opt_str.split('\n'):
                                        if '=' in line: c, l = line.split('=', 1); parsed_opts.append(f"{c}) {l}")
                                        else: parsed_opts.append(line)
                                    if "보기_list" not in current_entry: current_entry["보기_list"] = []
                                    current_entry["보기_list"].extend(parsed_opts)
                                    is_parent_added = True
                                    continue
                        
                        if pending_ranking_count and not new_entries:
                            opts = extract_options_from_table(block)
                            if opts: ranking_options_buffer.append(opts)
                            continue
                        
                        if not new_entries:
                            mc = extract_multi_column_input_table(block, current_entry, force_row_count=pending_max_n_count)
                            if mc: new_entries = mc; pending_max_n_count = None
                        
                        if not new_entries and current_entry.get("유형") in ["Single", "Multi"]:
                            if is_option_description_table(block):
                                opt_str = extract_single_choice_options(block)
                                current_entry["보기 값"] = opt_str
                                extracted_data.append(current_entry)
                                is_parent_added = True
                                continue
                        
                        if not new_entries and is_input_table(block):
                            if current_entry:
                                sub_cnt = 0
                                for row in rows:
                                    fc = row.cells[0].text.strip()
                                    if not fc: continue
                                    sub_cnt += 1
                                    new_entries.append({ "변수명": f"{current_entry['변수명']}_{sub_cnt}", "질문 내용": f"{current_entry['질문 내용']} ({fc})", "보기 값": "(숫자입력)", "유형": "Open" })

                        if not new_entries and current_entry:
                            table_vals_str, is_body_mapped = extract_table_scale(block)
                            is_matrix = False
                            if len(rows) > 1:
                                for row in rows[1:]:
                                    fc = row.cells[0].text.strip()
                                    # 동그라미 숫자 등의 척도값은 건너뛰고 질문라벨만 체크
                                    if fc and not fc.isdigit() and fc not in ["○", "●", "V"] and fc not in CIRCLE_MAP: 
                                        is_matrix = True; break
                            
                            # [수정] B1-B4 매트릭스 척도 처리 강화
                            if is_matrix:
                                sub_cnt = 0
                                for row in rows[1:]:
                                    fc = row.cells[0].text.strip()
                                    if not fc or fc in CIRCLE_MAP: continue
                                    sub_cnt += 1
                                    m_var = f"{current_entry['변수명']}_{sub_cnt}"
                                    new_entries.append({ "변수명": m_var, "질문 내용": f"[{current_entry['변수명']} 세부] {fc}", "보기 값": table_vals_str if table_vals_str else "(헤더참조)", "유형": "Matrix" })
                            elif not is_parent_added and not is_input_table(block):
                                split = check_and_split_time(current_entry)
                                if len(split) == 1: split = check_and_split_date(split[0])
                                if len(split) == 1: split = check_and_split_money(split[0])
                                if len(split) == 1: split = check_and_split_percent(split[0])
                                new_entries = split

            if new_entries:
                for item in new_entries:
                    variable_map[item['변수명']] = len(extracted_data)
                    extracted_data.append(item)
                is_parent_added = True

    if current_entry and not is_parent_added:
        flushed_data = flush_entry(current_entry)
        if flushed_data: 
            for item in flushed_data:
                variable_map[item['변수명']] = len(extracted_data)
                extracted_data.append(item)
            
    return pd.DataFrame(extracted_data)

def to_excel_with_usage_flag(df):
    rows = []
    code_start_pattern = re.compile(r"^(\d+|[①-⑩]|[a-zA-Z]|[가-하])[\.\)\s=]\s*(.*)")
    for idx, row in df.iterrows():
        var_name = row['변수명']; raw_q = str(row['질문 내용']); clean_q = re.sub(r"^\[.*?\]\s*", "", raw_q)
        if "_" in var_name:
            base_var, suffix = var_name.rsplit("_", 1)
            if raw_q.startswith("["): final_q_label = raw_q
            else: final_q_label = f"{base_var}. {suffix}) {clean_q}"
        else: final_q_label = f"{var_name}. {clean_q}"
        vals_str = str(row['보기 값']); formatted_values = ""
        if vals_str and vals_str.strip() != "" and vals_str != "nan":
            lines = vals_str.split('\n'); options = []; current_code = None; current_label_parts = []
            for line in lines:
                line = line.strip()
                if not line: continue
                is_new_code = False; temp_code = ""; temp_label = ""
                if "=" in line: is_new_code = True; temp_code, temp_label = line.split("=", 1)
                else:
                    match = code_start_pattern.match(line)
                    if match: is_new_code = True; temp_code, temp_label = match.groups()
                if is_new_code:
                    if current_code is not None: options.append(f"{current_code.strip()} = {' '.join(current_label_parts).strip()}")
                    current_code = temp_code; current_label_parts = [temp_label]
                else:
                    if current_code is not None: current_label_parts.append(line)
                    else: options.append(line)
            if current_code is not None: options.append(f"{current_code.strip()} = {' '.join(current_label_parts).strip()}")
            formatted_values = "\n".join(options) if options else vals_str
        rows.append({ "사용여부": "O", "V변수": "", "변수명": var_name, "질문 내용": final_q_label, "보기(Values)": formatted_values })
    result_df = pd.DataFrame(rows)
    var_list = df['변수명'].tolist(); var_counts = Counter(var_list); duplicates = [var for var, count in var_counts.items() if count > 1]
    highlight_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=False)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='Codebook')
        worksheet = writer.sheets['Codebook']
        for cell in worksheet[1]: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column <= 3: cell.alignment = align_center; 
                if cell.column == 3 and cell.value in duplicates: cell.fill = highlight_fill
                if cell.column > 3: cell.alignment = align_left
        worksheet.column_dimensions['A'].width = 8; worksheet.column_dimensions['B'].width = 15; worksheet.column_dimensions['C'].width = 20; worksheet.column_dimensions['D'].width = 50; worksheet.column_dimensions['E'].width = 40
    return output.getvalue()

def compress_var_list(var_list):
    if not var_list: return ""
    compressed = []; current_chunk = []; pattern = re.compile(r"^(.*?)(\d+)$")
    for var in var_list:
        if not current_chunk: current_chunk.append(var); continue
        prev_var = current_chunk[-1]; match_prev = pattern.match(prev_var); match_curr = pattern.match(var)
        is_continuous = False
        if match_prev and match_curr:
            prev_prefix, prev_num = match_prev.groups(); curr_prefix, curr_num = match_curr.groups()
            if prev_prefix == curr_prefix and int(curr_num) == int(prev_num) + 1: is_continuous = True
        if is_continuous: current_chunk.append(var)
        else:
            if len(current_chunk) >= 3: compressed.append(f"{current_chunk[0]} TO {current_chunk[-1]}")
            else: compressed.extend(current_chunk)
            current_chunk = [var]
    if len(current_chunk) >= 3: compressed.append(f"{current_chunk[0]} TO {current_chunk[-1]}")
    else: compressed.extend(current_chunk)
    return " ".join(compressed)

# [FIX] utils 에러 방지를 위한 내부 함수
def generate_spss_final(df_edited, encoding_type='utf-8'):
    enc_str = "UTF-8" if encoding_type == 'utf-8' else "CP949"
    syntax_lines = ["* SPSS Syntax Generated by Streamlit (Final).", f"* Encoding: {enc_str}.", "", "* 0. Set Working Directory and Load Data.", "CD '경로'.", "GET FILE='project_CE.sav'.", ""]
    if encoding_type == 'utf-8': syntax_lines.insert(2, "SET UNICODE=ON.")
    if '사용여부' in df_edited.columns: df_target = df_edited[df_edited['사용여부'].isin(['O', 'R'])].copy()
    else: df_target = df_edited.copy()
    syntax_lines.append("* 1. Rename Variables (B -> C)."); rename_count = 0
    unique_rows = df_target.drop_duplicates(subset=['변수명'], keep='first')
    for idx, row in unique_rows.iterrows():
        v_clean = str(row['변수명']).strip(); v_raw = str(row['V변수']).strip()
        if v_raw and v_raw.lower() != 'nan' and v_raw != v_clean: syntax_lines.append(f"Rename Var {v_raw}={v_clean}."); rename_count += 1
    if rename_count > 0: syntax_lines.append("EXECUTE."); syntax_lines.append("")
    syntax_lines.append("* 1.5 Recode Variables (Reverse Coding)."); recode_count = 0
    for idx, row in df_target.iterrows():
        if row['사용여부'] == 'R':
            v_name = row['변수명']; val_text = str(row['보기(Values)'])
            if not v_name or val_text == 'nan' or not val_text.strip(): continue
            codes = []
            for line in val_text.split('\n'):
                if '=' in line: c = line.split('=', 1)[0].strip(); 
                if c.isdigit(): codes.append(int(c))
            if codes:
                min_c, max_c = min(codes), max(codes); recode_pairs = []; 
                for c in codes: new_c = max_c + min_c - c; recode_pairs.append(f"({c}={new_c})")
                recode_str = " ".join(recode_pairs); syntax_lines.append(f"RECODE {v_name} {recode_str}."); recode_count += 1
    if recode_count > 0: syntax_lines.append("EXECUTE."); syntax_lines.append("")
    syntax_lines.append("VARIABLE LABELS"); unique_vars = df_target.drop_duplicates(subset=['변수명'], keep='first')
    for idx, row in unique_vars.iterrows():
        v = str(row['변수명']).strip(); l = str(row['질문 내용']).strip().replace('"', "'")
        if v: syntax_lines.append(f'  {v} "{l}"')
    syntax_lines.append(".\nEXECUTE.\n"); syntax_lines.append("VALUE LABELS"); value_map = {}
    for idx, row in df_target.iterrows():
        v = str(row['변수명']).strip(); val_text = str(row['보기(Values)']); is_reverse = (row['사용여부'] == 'R')
        if not v or val_text == 'nan' or not val_text.strip(): continue
        lines = val_text.split('\n'); codes_labels = []; codes_int = []
        if is_reverse:
            for line in lines:
                if '=' in line: c = line.split('=', 1)[0].strip(); 
                if c.isdigit(): codes_int.append(int(c))
            if codes_int: min_c, max_c = min(codes_int), max(codes_int)
        for line in lines:
            line = line.strip()
            if '=' in line:
                parts = line.split('=', 1); code = parts[0].strip(); label = parts[1].strip(); final_code = code
                if is_reverse and code.isdigit() and codes_int: c_int = int(code); new_c_int = max_c + min_c - c_int; final_code = str(new_c_int)
                if final_code and label: codes_labels.append((final_code, label))
        if codes_labels:
            try: codes_labels.sort(key=lambda x: int(x[0]))
            except: pass
            val_tuple = tuple(codes_labels)
            if val_tuple not in value_map: value_map[val_tuple] = []
            value_map[val_tuple].append(v)
    group_count = 0; total_groups = len(value_map)
    for val_tuple, var_list in value_map.items():
        group_count += 1
        var_block_str = compress_var_list(var_list); wrapped_vars = textwrap.wrap(var_block_str, width=80)
        for line in wrapped_vars: syntax_lines.append(f"  {line}")
        for code, label in val_tuple:
            label_clean = label.replace('"', "'"); syntax_lines.append(f'    {code} "{code}) {label_clean}"')
        if group_count < total_groups: syntax_lines.append("  /")
        else: syntax_lines.append("  .")
    syntax_lines.append("EXECUTE."); syntax_lines.append(""); syntax_lines.append("* 4. Save Data.")
    keep_vars = df_target['변수명'].drop_duplicates().tolist()
    if keep_vars:
        syntax_lines.append("SAVE OUTFILE='Project_DATA.sav'"); syntax_lines.append("  /KEEP="); 
        for var in keep_vars: syntax_lines.append(f"    {var}")
        syntax_lines.append("  .")
    else: syntax_lines.append("SAVE OUTFILE='Project_DATA.sav'.")
    syntax_lines.append("EXECUTE."); syntax_lines.append(""); syntax_lines.append("* 5. Export to Excel."); syntax_lines.append("GET FILE='Project_DATA.sav'."); syntax_lines.append("EXECUTE.")
    syntax_lines.append(""); syntax_lines.append("*_ SAVE - Values _."); syntax_lines.append("SAVE TRANSLATE OUTFILE='(RAW) Project_DATA.xlsx' /TYPE=XLS /VERSION=12 /MAP /REPLACE /FIELDNAMES /CELLS=VALUES.")
    syntax_lines.append(""); syntax_lines.append("*_ SAVE - Labels _."); syntax_lines.append("SAVE TRANSLATE OUTFILE='(LABEL) Project_DATA.xlsx' /TYPE=XLS /VERSION=12 /MAP /REPLACE /FIELDNAMES /CELLS=LABELS.")
    return "\n".join(syntax_lines)

# ==============================================================================
# Streamlit UI
# ==============================================================================
st.markdown("""
**[기능 설명]**
* **스마트 스캐닝:** 표 전체를 먼저 분석하여 **[자녀정보], [시간/분 입력], [단순 입력], [고정 합계], [가로형 입력], [가로형 척도]** 등의 유형을 자동으로 판단합니다.
* **복합 문항 지원:** A7 처럼 텍스트 안에 입력 칸이 여러 개 있는 경우(회/시간 등)도 자동으로 분리합니다.
* **질문 요약 (Beta):** 체크박스를 선택하면, 질문 내용의 불필요한 수식어를 제거하고 간결하게 요약합니다.
""")

tab1, tab2 = st.tabs(["1단계: 워드 ➡️ 엑셀 생성", "2단계: 엑셀 ➡️ SPSS 생성"])

with tab1:
    st.header("1. 워드 파일 파싱")
    uploaded_word = st.file_uploader("설문지(.docx) 업로드", type=["docx"], key="word_uploader")
    if uploaded_word:
        if st.button("분석 시작", key="btn_analyze"):
            with st.spinner("문서 구조 정밀 분석 중..."):
                try: 
                    df_raw = parse_word_to_df(uploaded_word)
                    st.session_state['df_raw'] = df_raw
                    st.success(f"분석 완료! {len(df_raw)}개 항목 추출됨")
                except Exception as e: 
                    st.error(f"오류 발생: {e}")
                    
    if 'df_raw' in st.session_state:
        st.subheader("📊 분석 결과 미리보기")
        st.dataframe(st.session_state['df_raw'], use_container_width=True, height=400)
        
        # 요약 옵션
        st.markdown("---")
        use_summary = st.checkbox("✂️ 긴 질문 내용을 간략하게 요약하기 (Beta)", 
                                  help="질문 끝의 '~입니까?', '귀하의' 같은 불필요한 문구를 자동으로 제거합니다.")
        
        st.info("아래 엑셀 파일을 다운로드하여 내용을 수정하세요.")
        
        if use_summary:
            df_to_download = st.session_state['df_raw'].copy()
            df_to_download['질문 내용'] = df_to_download['질문 내용'].apply(summarize_label_regex)
            excel_data = to_excel_with_usage_flag(df_to_download)
        else:
            excel_data = to_excel_with_usage_flag(st.session_state['df_raw'])
            
        st.download_button(
            label="📥 편집용 코드북 다운로드 (Codebook.xlsx)",
            data=excel_data,
            file_name="Codebook_Draft.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

with tab2:
    st.header("2. SPSS 신택스 생성")
    uploaded_excel = st.file_uploader("수정된 코드북(.xlsx) 업로드", type=["xlsx"], key="excel_uploader")
    if uploaded_excel:
        try:
            df_edited = pd.read_excel(uploaded_excel)
            if '사용여부' not in df_edited.columns: 
                st.error("⚠️ 1단계에서 생성된 엑셀 파일을 사용해주세요.")
            else:
                st.success("파일 로드 성공!")
                df_filtered = df_edited[df_edited['사용여부'].isin(['O', 'R'])].copy()
                st.write(f"총 {len(df_edited)}개 중 {len(df_filtered)}개 문항 선택됨")
                
                col1, col2 = st.columns(2)
                
                # Option 1: UTF-8
                with col1:
                    try:
                        spss_utf8 = utils.generate_spss_final(df_edited, encoding_type='utf-8')
                    except:
                        spss_utf8 = generate_spss_final(df_edited, encoding_type='utf-8')
                        
                    st.download_button(
                        label="💾 (추천) SPSS 신택스 다운로드 (UTF-8)",
                        data=spss_utf8.encode('utf-8-sig'), 
                        file_name="Syntax_UTF8.sps",
                        mime="text/plain",
                        type="primary",
                        use_container_width=True
                    )
                    st.caption("최신 버전 SPSS 사용 시 권장")

                # Option 2: CP949
                with col2:
                    try:
                        spss_cp949 = utils.generate_spss_final(df_edited, encoding_type='cp949')
                    except:
                        spss_cp949 = generate_spss_final(df_edited, encoding_type='cp949')

                    st.download_button(
                        label="💾 (구버전) SPSS 신택스 다운로드 (CP949)",
                        data=spss_cp949.encode('cp949', errors='ignore'), 
                        file_name="Syntax_CP949.sps",
                        mime="text/plain",
                        type="secondary",
                        use_container_width=True
                    )
                    st.caption("SPSS에서 한글이 깨질 때 사용")
                
                with st.expander("신택스 내용 미리보기 (UTF-8 기준)"):
                    st.code(spss_utf8, language="spss")
        except Exception as e: 
            st.error(f"파일 처리 중 오류: {e}")
