"""
╔══════════════════════════════════════════════════════════════════════════╗
║  파일명 : pages/3_🛠️_SPSS_정제.py                                          ║
║                                                                          ║
║  Raw 데이터와 Code북을 비교해 SPSS 변수명 변경 신텍스를 생성합니다.            ║
║  매칭 로직(analyze)은 Streamlit 과 분리해 두어 단독 테스트가 가능합니다.       ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import collections
import io
import os
import re
import sys
import traceback

import pandas as pd
import streamlit as st

# (주의) utils 모듈이 같은 폴더나 상위 폴더에 있어야 합니다.
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import utils
import excel_style

assert utils.MODULE_ROLE == "utils", "utils.py 가 아닌 파일이 import 되었습니다."

st.set_page_config(page_title="SPSS 변수명 정제", layout="wide")

# ==============================================================================
# 0. 변수명 규칙
# ==============================================================================

# SPSS 예약어. 변수명으로 쓸 수 없습니다.
RESERVED_WORDS = {"ALL", "AND", "BY", "EQ", "GE", "GT", "LE", "LT",
                  "NE", "NOT", "OR", "TO", "WITH"}

# 매핑에서 제외할 관리용 컬럼
SKIP_COLS = {"no", "id", "번호", "순번"}


def make_valid_name(candidate, fallback=""):
    """SPSS 변수명 규칙에 맞는 이름을 만든다.

    규칙: 영문자로 시작 / 영문·숫자·밑줄만 / 64바이트 이하 / 예약어 불가.

    utils.sanitize_var_name 은 한글을 전부 제거하므로, 라벨이 한글뿐이면
    빈 문자열이 됩니다(예: '성별' -> ''). 그대로 두면 중복 처리 과정에서
    `_1`, `_2` 같은 이름이 만들어져 SPSS 에서 실행이 실패합니다.
    그래서 비거나 숫자로 시작하면 Code 변수명으로 되돌립니다.
    """
    name = utils.sanitize_var_name(candidate or "").strip("_")
    if not name or not name[0].isalpha():
        name = utils.sanitize_var_name(fallback or "").strip("_")
    if not name:
        return ""
    if not name[0].isalpha():
        name = "V" + name
    if name.upper() in RESERVED_WORDS:
        name = name + "_"
    while len(name.encode("utf-8")) > 64:
        name = name[:-1]
    return name


def is_valid_name(name):
    """SPSS 가 받아들일 이름인지."""
    name = str(name or "").strip()
    return bool(
        name
        and name[0].isalpha()
        and re.fullmatch(r"[A-Za-z][A-Za-z0-9_]*", name)
        and name.upper() not in RESERVED_WORDS
        and len(name.encode("utf-8")) <= 64
    )


# ==============================================================================
# 0-1. 엑셀 서식 (원본 납품 파일과 동일하게)
# ==============================================================================

HEADER_FILL = "FFFFCC"     # 연노랑 — 원본 파일의 머리행 색
BODY_FONT = "맑은 고딕"
BODY_SIZE = 9
HEADER_ROW_HEIGHT = 17.55
FIRST_COL_WIDTH = 9.0

# Code 시트에서 '코드값/보기' 같은 안내 행은 칠하지 않는다
CODE_GUIDE_WORDS = {"코드값", "보기"}


def is_code_header_row(row_values):
    """Code 시트에서 '변수명 + 문항명' 행인지. (색칠 대상)

    변수명은 영문으로 시작하고(Q1, SQ1 …), 코드값 행은 숫자로 시작한다.
    """
    if not row_values:
        return False
    first = str(row_values[0] or "").strip()
    second = str(row_values[1] or "").strip() if len(row_values) > 1 else ""
    if not first or not second:
        return False
    if first in CODE_GUIDE_WORDS:
        return False
    return bool(re.match(r"^[A-Za-z]", first))


def style_sheet(ws, header_rows, bold_header, freeze_row=None):
    """시트에 원본과 동일한 서식을 입힌다.

    header_rows: 색을 칠할 1-based 행 번호 집합
    bold_header: 머리행을 굵게 할지 (Data/Label 은 True, Code 는 False)
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill = PatternFill("solid", fgColor=HEADER_FILL)
    body_font = Font(name=BODY_FONT, size=BODY_SIZE)
    head_font = Font(name=BODY_FONT, size=BODY_SIZE, bold=bold_header)
    center = Alignment(horizontal="center", vertical="center")
    vcenter = Alignment(vertical="center")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        is_head = row[0].row in header_rows
        for cell in row:
            if is_head:
                cell.fill = fill
                cell.font = head_font
                cell.alignment = center if bold_header else vcenter
            else:
                cell.font = body_font
                cell.alignment = vcenter

    for r in header_rows:
        ws.row_dimensions[r].height = HEADER_ROW_HEIGHT
    ws.column_dimensions[get_column_letter(1)].width = FIRST_COL_WIDTH
    if freeze_row:
        ws.freeze_panes = f"A{freeze_row}"


# ==============================================================================
# 1. 매칭 로직 (Streamlit 비의존 — 단독 테스트 가능)
# ==============================================================================

def get_priority(status):
    """낮을수록 우선. 정확 매칭 > 세트 문항 > 기타/파생."""
    if status in ("매칭 성공", "매칭 성공 (순위 문항)"):
        return 1
    if status == "매칭 성공 (세트 문항)":
        return 2
    return 3


def analyze(df_raw, df_code, label_col=1):
    """Raw 데이터와 Code북을 대조해 변수명 매핑을 만든다.

    label_col: Code북에서 질문 라벨이 들어 있는 열 번호 (0=A, 1=B, 2=C …)

    반환: (final_data, code_updates, warnings)
    """
    warnings = []

    # 소문자 키로 매핑하므로 대소문자만 다른 동명 컬럼은 하나가 묻힌다. 미리 알린다.
    lowered = collections.Counter(str(c).strip().lower() for c in df_raw.columns)
    collided = [k for k, n in lowered.items() if n > 1]
    if collided:
        warnings.append(
            "Raw 데이터에 대소문자만 다른 같은 이름의 열이 있습니다: "
            + ", ".join(collided[:5]) + " — 하나만 매칭됩니다.")

    raw_cols_map = {str(c).strip().lower(): str(c).strip() for c in df_raw.columns}
    temp_vars = []

    # --- [Step 1] Code북 순회하며 후보군 수집 ---
    for idx, row in df_code.iterrows():
        if len(row) <= max(1, label_col):
            continue
        if pd.isna(row.iloc[0]):
            continue

        col_a_val = utils.clean_text(row.iloc[0])          # 예: 문1
        # Code명이 '문1', '문 2_1' 이면 'Q1', 'Q2_1' 로 치환 (두 자리 이상도 처리)
        if col_a_val:
            col_a_val = re.sub(r"^문\s*(\d+)", r"Q\1", col_a_val)
        if not col_a_val:
            continue

        col_label = utils.clean_text(row.iloc[label_col])   # 질문 라벨

        # 라벨에서 순위 표기 제거
        clean_label = re.sub(r"[\(\[<]?\s*\d+\s*순위\s*[\)\]>]?\s*", "", col_label).strip()
        current_label_base = utils.extract_base_name(clean_label)
        if current_label_base:
            current_label_base = re.sub(r"^문\s*(\d+)", r"Q\1", current_label_base)
        else:
            current_label_base = col_a_val

        is_matched = False
        search_base_raw = col_a_val.lower()
        search_label_base = current_label_base

        # [로직 1] 정확히 일치
        if col_a_val.lower() in raw_cols_map:
            temp_vars.append({
                "Raw 변수명": raw_cols_map[col_a_val.lower()],
                "Code 변수명": col_a_val,
                "질문 내용": col_label,
                "변경할 변수명": make_valid_name(current_label_base, col_a_val),
                "상태": "매칭 성공",
                "code_idx": idx,
            })
            is_matched = True

        # [로직 2] 순위 문항 (Q5_RK1 -> q5_1)
        if not is_matched:
            rk_match = re.search(r"^(.*?)_?rk(\d+)$", col_a_val.lower())
            if rk_match:
                base_raw, rank_num = rk_match.group(1), rk_match.group(2)
                expected_raw_col = f"{base_raw}_{rank_num}"
                if expected_raw_col in raw_cols_map:
                    temp_vars.append({
                        "Raw 변수명": raw_cols_map[expected_raw_col],
                        "Code 변수명": col_a_val,
                        "질문 내용": col_label,
                        "변경할 변수명": make_valid_name(
                            f"{current_label_base}_{rank_num}", f"{base_raw}_{rank_num}"),
                        "상태": "매칭 성공 (순위 문항)",
                        "code_idx": idx,
                    })
                    is_matched = True
                    search_base_raw = base_raw
                    search_label_base = current_label_base

        # [로직 3] 기타/파생 변수 (복수응답 세트 포함)
        prefix = search_base_raw + "_"
        for rc_lower, rc_original in raw_cols_map.items():
            if not rc_lower.startswith(prefix):
                continue
            suffix = rc_original[len(search_base_raw):]
            if not suffix.startswith(("_", "-")):
                suffix = "_" + suffix

            if is_matched:
                state_msg = "매칭 성공 (기타/파생 변수)"
                display_label = f"{clean_label} [기타]"
            else:
                state_msg = "매칭 성공 (세트 문항)"
                display_label = col_label

            temp_vars.append({
                "Raw 변수명": rc_original,
                "Code 변수명": col_a_val,
                "질문 내용": display_label,
                "변경할 변수명": make_valid_name(search_label_base + suffix,
                                            col_a_val + suffix),
                "상태": state_msg,
                "code_idx": idx,
            })

    # --- [Step 2] 최적 매칭 선정 (중복/경합 방지) ---
    best_match_dict = {}
    for item in temp_vars:
        raw_col = item["Raw 변수명"]
        if raw_col.lower() in SKIP_COLS:
            continue          # 관리용 컬럼은 이름 중복 계산에서도 빼야 한다
        prev = best_match_dict.get(raw_col)
        if prev is None or get_priority(item["상태"]) < get_priority(prev["상태"]):
            best_match_dict[raw_col] = item

    name_freq = collections.Counter(
        item["변경할 변수명"] for item in best_match_dict.values() if item["변경할 변수명"])
    name_counter = collections.defaultdict(int)

    # --- [Step 3] Raw 데이터 원본 순서대로 결과 구성 ---
    final_data = []
    for raw_col in df_raw.columns:
        raw_col_str = str(raw_col).strip()
        if raw_col_str.lower() in SKIP_COLS:
            continue

        item = best_match_dict.get(raw_col_str)
        if item is None:
            final_data.append({
                "Raw 변수명": raw_col_str,
                "Code 변수명": "-",
                "질문 내용": "-",
                "변경할 변수명": "",
                "상태": "매칭 실패 (확인 필요)",
            })
            continue

        candidate = item["변경할 변수명"]
        if candidate and name_freq[candidate] > 1:
            name_counter[candidate] += 1
            item["변경할 변수명"] = f"{candidate}_{name_counter[candidate]}"
        final_data.append(item)

    # --- [Step 4] Code북 갱신 딕셔너리 ---
    code_updates = {}
    for item in final_data:
        if "code_idx" not in item or not item["변경할 변수명"]:
            continue
        r_idx = item["code_idx"]
        # 같은 행에 여러 변수가 걸리면 우선순위가 더 높은 것만 덮어쓴다.
        # 우선순위가 같으면 먼저 온 것(예: 복수응답의 Q1_1)을 유지한다.
        prev = code_updates.get(r_idx)
        if prev is not None and get_priority(item["상태"]) >= get_priority(prev["status"]):
            continue
        code_updates[r_idx] = {"name": item["변경할 변수명"], "status": item["상태"]}

    return final_data, {k: v["name"] for k, v in code_updates.items()}, warnings


def read_source_sav(file_bytes):
    """원본 .sav 에서 값라벨·측도·결측을 읽어온다.

    엑셀에는 응답 라벨(1=남성 …)이 들어 있지 않다. 원본 .sav 를 함께 올리면
    변수명이 바뀌어도 값라벨을 그대로 옮길 수 있다.
    키는 소문자 원본 변수명이다.
    """
    import tempfile
    from pathlib import Path as _Path

    import pyreadstat

    with tempfile.TemporaryDirectory() as tmp:
        path = _Path(tmp) / "src.sav"
        path.write_bytes(file_bytes)
        # 값라벨을 코드 그대로 두고 읽는다 (apply_value_formats=False)
        df, meta = pyreadstat.read_sav(str(path), user_missing=True)

    return {
        "df": df,
        "value_labels": {k.lower(): v for k, v in (meta.variable_value_labels or {}).items()},
        "var_labels": {k.lower(): v for k, v in (meta.column_names_to_labels or {}).items() if v},
        "measures": {k.lower(): v for k, v in (meta.variable_measure or {}).items()},
        "missing": {k.lower(): v for k, v in (meta.missing_ranges or {}).items()},
        "columns": list(meta.column_names or []),
        "columns_lower": {str(c).lower() for c in (meta.column_names or [])},
    }


def sav_safe_name(original):
    """.sav 저장이 가능한 열 이름으로 정리.

    SPSS 유니코드 모드는 한글 변수명을 허용하므로(`응답일시` 등) 굳이 영문으로
    바꾸지 않는다. 공백·특수문자·숫자 시작·예약어만 손본다.
    변경이 필요 없으면 원래 이름을 그대로 돌려준다.
    """
    name = str(original).strip()
    name = re.sub(r"[\s\-]+", "_", name)                 # 공백·하이픈 -> 밑줄
    name = re.sub(r"[^\w가-힣]", "", name, flags=re.UNICODE)  # 나머지 특수문자 제거
    name = re.sub(r"__+", "_", name).strip("_")
    if not name:
        return ""
    if name[0].isdigit():
        name = "V" + name
    if name.upper() in RESERVED_WORDS:
        name = name + "_"
    while len(name.encode("utf-8")) > 64:
        name = name[:-1]
    return name


def build_sav(df_raw, edited_df, source=None):
    """변경된 변수명을 적용한 .sav 를 만든다. (bytes, 적용 내역 dict)

    - 변수라벨은 Code북의 '질문 내용' 을 넣는다 (SPSS 한도 256바이트에서 절단).
    - source 를 주면 원본 .sav 의 값라벨·측도·결측을 새 변수명으로 옮겨 담는다.
      (엑셀에는 응답 라벨이 없으므로 원본 .sav 가 있어야 살릴 수 있다)
    - 이름을 바꾸지 않은 열은 그대로 둔다. 단 공백·특수문자가 있으면 저장이
      실패하므로 그 부분만 정리하고, 바뀐 열은 반환값에 담아 화면에 알린다.
      (한글 변수명은 SPSS 유니코드 모드에서 유효하므로 유지한다)
    """
    import tempfile
    from pathlib import Path as _Path

    import pyreadstat

    rename_map, label_map = {}, {}
    for _, row in edited_df.iterrows():
        old = str(row["Raw 변수명"]).strip()
        new = str(row["변경할 변수명"]).strip()
        if new and new.lower() != "nan":
            rename_map[old] = new
        label = str(row.get("질문 내용", "")).strip()
        if label and label != "-":
            label_map[old] = label

    df = df_raw.copy()
    names, labels, auto_fixed, used = [], [], [], set()
    for col in df.columns:
        original = str(col).strip()
        name = rename_map.get(original) or sav_safe_name(original)
        if not name:
            name = "V" + str(len(names) + 1)
        if original not in rename_map and name != original:
            auto_fixed.append(f"{original} → {name}")
        base = name
        n = 2
        while name in used:            # 이름이 겹치면 SPSS 가 파일을 거부한다
            name = f"{base}_{n}"
            n += 1
        used.add(name)
        names.append(name)
        labels.append(byte_trim(label_map.get(original, original), 256))

    df.columns = names

    # 원본 .sav 의 값라벨·측도·결측을 새 변수명 기준으로 옮긴다
    value_labels, measures, missing_ranges = {}, {}, {}
    carried, not_found = [], []
    if source:
        for original, new_name in zip([str(c).strip() for c in df_raw.columns], names):
            key = original.lower()
            if key in source["value_labels"]:
                value_labels[new_name] = source["value_labels"][key]
                carried.append(new_name)
            elif key not in source.get("columns_lower", set()):
                not_found.append(original)
            if key in source["measures"]:
                measures[new_name] = source["measures"][key]
            if key in source["missing"]:
                missing_ranges[new_name] = source["missing"][key]
    for name in names:
        if df[name].dtype == object:
            converted = pd.to_numeric(df[name], errors="coerce")
            # 원래 값이 있는데 숫자로 못 바꾼 칸이 하나도 없으면 숫자열로 본다
            if not (df[name].notna() & converted.isna()).any():
                df[name] = converted
            else:
                df[name] = df[name].astype(object)

    with tempfile.TemporaryDirectory() as tmp:
        path = _Path(tmp) / "out.sav"
        pyreadstat.write_sav(
            df, str(path),
            column_labels=labels,
            variable_value_labels=value_labels or None,
            variable_measure=measures or None,
            missing_ranges=missing_ranges or None,
        )
        return path.read_bytes(), {
            "vars": len(names), "auto_fixed": auto_fixed,
            "value_labels": len(value_labels), "not_in_source": not_found,
        }


def byte_trim(text, limit):
    """UTF-8 바이트 기준 절단 (한글 1자 = 3바이트)."""
    raw = str(text).encode("utf-8")
    if len(raw) <= limit:
        return str(text)
    return raw[:limit].decode("utf-8", errors="ignore").rstrip()


def build_syntax(edited_df, file_stem):
    """RENAME VARIABLES 구문 생성. (구문 문자열, 변환 건수)"""
    pairs = []
    for _, row in edited_df.iterrows():
        old_v = str(row["Raw 변수명"]).strip()
        new_v = str(row["변경할 변수명"]).strip()
        if old_v and new_v and new_v.lower() != "nan" and old_v.lower() != new_v.lower():
            pairs.append((old_v, new_v))

    lines = [f"* Auto Generated Syntax for {file_stem}.",
             f'GET FILE="{file_stem}.sav".']
    if pairs:
        lines.append("RENAME VARIABLES")
        lines += [f"  ({o} = {n})" for o, n in pairs]
        lines.append(".")
        lines.append("EXECUTE.")
    else:
        # 변환할 것이 없는데 RENAME VARIABLES 만 쓰면 SPSS 문법 오류가 납니다.
        lines.append("* 변경할 변수명이 없어 RENAME 구문을 생성하지 않았습니다.")
    lines.append(f'SAVE OUTFILE="{file_stem}_Renamed.sav".')
    lines.append("EXECUTE.")
    return "\n".join(lines), len(pairs)


# ==============================================================================
# 2. 화면
# ==============================================================================

if not utils.check_password():
    st.stop()


def ss(key, default=None):
    return st.session_state[key] if key in st.session_state else default


st.header("📊 SPSS 변수명 자동 정제 & 신텍스 생성")
st.markdown("""
**Raw 데이터**와 **Code북**을 비교하여 SPSS 변수명 변경 신텍스를 생성합니다.
* **기능 1:** 라벨의 앞부분(SQ1)을 추출하여 변수명으로 자동 변환
* **기능 2:** Code북에 `문1`, `문2_1`로 표기된 변수를 `Q1`, `Q2_1`로 자동 치환하여 인식
* **기능 3:** 척도 문항 중복 시 `_1`, `_2` 자동 부여 및 순위 문항(RK) 완벽 매칭
* **기능 4:** 파생 변수 탐색 시 라벨에 `[기타]` 추가 / 복수응답 Code북 업데이트 시 첫 번째 문항으로 고정
* **기능 5:** 엑셀 다운로드 시 Code북 시트 자동 갱신 및 순수 데이터(디자인 없음) 내보내기
* **기능 6:** SPSS 변수명 규칙 검사 — 한글·숫자 시작·예약어·중복을 걸러냅니다
* **기능 7:** 원본 엑셀의 서식(1행 색·굵게·열 너비·틀 고정)을 그대로 유지한 채 내보내기
* **기능 7:** 변수명과 변수라벨이 적용된 **SPSS 데이터(.sav)** 바로 내보내기
""")

st.markdown("### 1. 파일 업로드")
up1, up2 = st.columns(2)
with up1:
    uploaded_file = st.file_uploader("엑셀 파일(.xlsx) — Code북 필수", type=["xlsx"],
                                     key="spss_file_uploader")
with up2:
    up_sav = st.file_uploader("원본 .sav (선택)", type=["sav"], key="spss_src_sav",
                              help="응답 라벨(1=남성 …)은 엑셀에 없습니다. 원본 .sav 를 올리면 "
                                   "값라벨·측도·결측을 가져오고, 데이터도 .sav 에서 읽을 수 있습니다.")

source_meta = None
if up_sav is not None:
    try:
        source_meta = read_source_sav(up_sav.getvalue())
        st.success(f"원본 .sav — 변수 {len(source_meta['columns'])}개, 케이스 "
                   f"{len(source_meta['df']):,}개, 값라벨 보유 {len(source_meta['value_labels'])}개")
    except Exception as e:
        st.error(f"원본 .sav 를 읽지 못했습니다: {type(e).__name__}: {e}")
        source_meta = None

if uploaded_file:
    try:
        xl = pd.ExcelFile(uploaded_file)
        sheet_names = xl.sheet_names

        col1, col2, col3 = st.columns(3)
        with col1:
            raw_sheet = st.selectbox("Raw 데이터 시트", sheet_names, index=0, key="raw_sheet_select")
        with col2:
            code_idx = 2 if len(sheet_names) > 2 else (1 if len(sheet_names) > 1 else 0)
            code_sheet = st.selectbox("Code북 시트", sheet_names, index=code_idx, key="code_sheet_select")
        with col3:
            label_letter = st.selectbox(
                "Code북의 질문 라벨 열", ["B", "C", "D", "E"], index=0, key="label_col_select",
                help="변수명은 A열, 질문 라벨은 이 열에서 읽습니다.")

        if raw_sheet == code_sheet:
            st.warning("Raw 시트와 Code북 시트가 같습니다. 서로 다른 시트를 선택하세요.")

        use_sav_data = False
        if source_meta is not None:
            use_sav_data = st.radio(
                "분석·저장에 사용할 데이터",
                [".sav 원본 (권장)", "엑셀 Raw 시트"], horizontal=True, index=0,
                help=".sav 를 쓰면 자료형·결측이 원본 그대로 유지되고 엑셀과의 불일치도 없습니다.",
            ).startswith(".sav")

            # 두 파일의 열 구성이 다르면 미리 알린다
            excel_cols = {str(c).strip().lower()
                          for c in pd.read_excel(uploaded_file, sheet_name=raw_sheet, nrows=0).columns}
            sav_cols = source_meta["columns_lower"]
            only_excel = sorted(excel_cols - sav_cols)
            only_sav = sorted(sav_cols - excel_cols)
            if only_excel or only_sav:
                with st.expander(f"엑셀과 .sav 의 열 구성이 다릅니다 "
                                 f"(엑셀만 {len(only_excel)}개 / .sav만 {len(only_sav)}개)"):
                    if only_excel:
                        st.write("엑셀에만 있음: " + ", ".join(only_excel[:30]))
                    if only_sav:
                        st.write(".sav 에만 있음: " + ", ".join(only_sav[:30]))

        if st.button("분석 시작", key="analyze_btn"):
            with st.spinner("데이터 분석 및 매칭 중..."):
                all_sheets = pd.read_excel(uploaded_file, sheet_name=None)
                df_raw = source_meta["df"] if (source_meta is not None and use_sav_data) \
                    else all_sheets[raw_sheet]
                df_code = pd.read_excel(uploaded_file, sheet_name=code_sheet, header=None)

                label_col = "BCDE".index(label_letter) + 1
                final_data, code_updates, warns = analyze(df_raw, df_code, label_col)

                # 서식을 살려 내보내려면 원본 파일 자체가 필요하다 (pandas 로는 서식을 못 읽음)
                st.session_state["spss_original_bytes"] = uploaded_file.getvalue()
                st.session_state["spss_raw_columns"] = [str(c).strip() for c in df_raw.columns]
                st.session_state["spss_sheet_names"] = sheet_names
                st.session_state["spss_all_sheets"] = all_sheets
                st.session_state["spss_df_code"] = df_code
                st.session_state["spss_target_sheets"] = [raw_sheet]
                st.session_state["spss_code_sheet"] = code_sheet      # 아래 내보내기에서 사용
                st.session_state["spss_source_meta"] = source_meta
                st.session_state["spss_use_sav_data"] = bool(source_meta is not None and use_sav_data)
                st.session_state["spss_df_for_sav"] = df_raw
                st.session_state["spss_code_updates"] = code_updates
                st.session_state["spss_result_df"] = pd.DataFrame(final_data)
                # 파일명에 점이 여러 개여도 확장자만 떼어낸다
                st.session_state["spss_file_name"] = uploaded_file.name.rsplit(".", 1)[0]
                st.session_state.pop("spss_exports", None)

                for w in warns:
                    st.warning(w)
                st.success("분석이 완료되었습니다! 아래 표에서 결과를 확인하세요.")

    except Exception as e:
        st.error(f"오류가 발생했습니다: {e}")
        st.code(traceback.format_exc())

# ------------------------------------------------------------------ 결과 확인
if "spss_result_df" in st.session_state:
    st.markdown("---")
    st.markdown("### 2. 결과 확인 및 수정")
    st.info("💡 **'변경할 변수명'** 컬럼을 더블클릭하여 직접 수정할 수 있습니다.")

    edited_df = st.data_editor(
        st.session_state["spss_result_df"],
        column_config={
            "상태": st.column_config.TextColumn("상태", disabled=True),
            "Raw 변수명": st.column_config.TextColumn(disabled=True),
            "Code 변수명": st.column_config.TextColumn(disabled=True),
            "질문 내용": st.column_config.TextColumn(disabled=True),
        },
        use_container_width=True,
        height=600,
        hide_index=True,
        key="data_editor",
    )

    # --- 변수명 검사: 여기서 걸러야 SPSS 에서 실행 실패를 막는다 ---
    names = edited_df["변경할 변수명"].astype(str).str.strip()
    filled = names[names.ne("") & names.ne("nan")]

    dup = [n for n, c in collections.Counter(filled).items() if c > 1]
    bad = sorted({n for n in filled if not is_valid_name(n)})

    if dup:
        st.error("**변수명 중복** — SPSS 에서 실행되지 않습니다: " + ", ".join(dup[:8])
                 + (f" 외 {len(dup) - 8}개" if len(dup) > 8 else ""))
    if bad:
        st.error("**SPSS 변수명 규칙 위반** (영문 시작·영문/숫자/밑줄·64바이트·예약어): "
                 + ", ".join(bad[:8]) + (f" 외 {len(bad) - 8}개" if len(bad) > 8 else ""))
    if not dup and not bad:
        st.success(f"변수명 검사 통과 — 변환 대상 {len(filled)}개")

    unmatched = int((edited_df["상태"] == "매칭 실패 (확인 필요)").sum())
    if unmatched:
        st.warning(f"매칭 실패 {unmatched}개 — 이 변수들은 이름이 바뀌지 않습니다.")

    # ------------------------------------------------------------ 내보내기
    st.markdown("---")
    st.markdown("### 3. 파일 내보내기")

    _src = ss("spss_source_meta")
    if _src is None:
        st.caption("원본 .sav 없이 생성하면 값라벨(1=남성 …)이 비어 있습니다. "
                   "1단계에서 원본 .sav 를 함께 올리세요.")
    else:
        st.caption(f"값라벨 출처: 원본 .sav (값라벨 보유 변수 {len(_src['value_labels'])}개) · "
                   f"데이터 출처: {'.sav 원본' if ss('spss_use_sav_data') else '엑셀 Raw 시트'}")

    sheet_names_all = ss("spss_sheet_names", [])
    code_sheet_name = ss("spss_code_sheet")
    default_targets = [n for n in sheet_names_all
                       if n != code_sheet_name
                       and (n == (ss("spss_target_sheets") or [""])[0]
                            or "DATA" in n.upper() or "LABEL" in n.upper())]
    st.session_state["spss_export_targets"] = st.multiselect(
        "새 변수명 행을 넣을 시트", [n for n in sheet_names_all if n != code_sheet_name],
        default=default_targets, key="export_targets_select",
        help="선택한 시트의 맨 위에 새 변수명 행이 추가됩니다. Code북 시트는 A열 값만 바뀝니다.")
    st.session_state["spss_keep_old_header"] = st.checkbox(
        "원래 변수명 행 남기기 (2행으로 내림)", value=True, key="keep_old_header_cb")

    if "spss_original_bytes" in st.session_state and st.session_state["spss_export_targets"]:
        info = excel_style.preview_header_style(
            st.session_state["spss_original_bytes"], st.session_state["spss_export_targets"][0])
        if info:
            st.caption(
                f"원본에서 인식한 머리행 서식 — 배경 #{info.get('채우기') or '없음'} · "
                f"{'굵게' if info.get('굵게') else '보통'} · {info.get('글꼴')} {info.get('크기')}pt"
                f"{' · 가운데 정렬' if info.get('가운데') else ''}  (출처 {info.get('출처')})")
        else:
            st.caption("원본 머리행에서 서식을 찾지 못했습니다. 기본 서식(연노랑 + 굵게)을 적용합니다.")

    enc_label = st.radio("신텍스 인코딩", ["cp949 (한글 Windows SPSS)", "utf-8 (유니코드 모드)"],
                         horizontal=True, index=0,
                         help="SPSS 가 유니코드 모드면 utf-8 을 선택하세요. 한글이 깨지면 반대로 바꿔 보세요.")

    # 파일 생성은 버튼을 눌렀을 때만. (표를 수정할 때마다 엑셀을 다시 만들면 느려집니다)
    if st.button("파일 생성", type="primary", key="build_btn"):
        with st.spinner("파일을 만드는 중..."):
            stem = st.session_state["spss_file_name"]
            exports = {}

            syntax, count = build_syntax(edited_df, stem)
            if enc_label.startswith("cp949"):
                try:
                    syntax_bytes = syntax.encode("cp949")
                except UnicodeEncodeError:
                    syntax_bytes = syntax.encode("utf-8-sig")
                    st.warning("⚠️ cp949 로 표현할 수 없는 문자가 있어 UTF-8 로 저장했습니다.")
            else:
                syntax_bytes = syntax.encode("utf-8-sig")
            exports["sps"] = (syntax_bytes, f"{stem}_Rename.sps", count)

            out_map = io.BytesIO()
            with pd.ExcelWriter(out_map, engine="xlsxwriter") as writer:
                edited_df.to_excel(writer, index=False)
            exports["map"] = (out_map.getvalue(), f"{stem}_Mapping.xlsx", len(edited_df))

            if "spss_original_bytes" in st.session_state:
                # 원본 서식을 유지하기 위해 원본 통합문서를 열어 값만 바꾼다.
                # {원래 열 이름: 새 이름}. 빈 값과 'nan' 은 변경 대상이 아니다.
                rename_map = {
                    str(r["Raw 변수명"]).strip(): str(r["변경할 변수명"]).strip()
                    for _, r in edited_df.iterrows()
                    if str(r["변경할 변수명"]).strip() not in ("", "nan")
                }
                raw_cols = ss("spss_raw_columns", [])
                new_header = [rename_map.get(c, c) for c in raw_cols]
                try:
                    blob = excel_style.export_renamed_workbook(
                        st.session_state["spss_original_bytes"],
                        new_header=new_header,
                        rename_map=rename_map,
                        target_sheets=ss("spss_export_targets", []),
                        code_sheet=ss("spss_code_sheet"),
                        code_updates=ss("spss_code_updates", {}),
                        keep_old_header=ss("spss_keep_old_header", True),
                    )
                    exports["data"] = (blob, f"{stem}_Renamed.xlsx", 0)
                except Exception as e:
                    st.error(f"엑셀 내보내기 실패: {type(e).__name__}: {e}")

            # --- SPSS 데이터(.sav) ---
            # 변수명에 오류가 있으면 SPSS 가 파일을 거부하므로 만들지 않는다.
            df_for_sav = ss("spss_df_for_sav")
            if df_for_sav is not None and not dup and not bad:
                try:
                    sav_blob, sav_info = build_sav(df_for_sav, edited_df, ss("spss_source_meta"))
                    exports["sav"] = (sav_blob, f"{stem}_Renamed.sav", sav_info["vars"])
                    st.session_state["spss_sav_info"] = sav_info
                except Exception as e:
                    st.error(f".sav 생성 실패: {type(e).__name__}: {e}")
                    st.code(traceback.format_exc())
            elif dup or bad:
                st.warning("변수명 오류가 있어 .sav 는 만들지 않았습니다. 위 오류를 먼저 해결하세요.")

            st.session_state["spss_exports"] = exports

    exports = ss("spss_exports")
    if exports:
        c1, c2, c3, c4 = st.columns(4)
        XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        with c1:
            blob, fname, count = exports["sps"]
            st.download_button("📄 Syntax 다운로드 (.sps)", blob, file_name=fname,
                               mime="text/plain", type="primary")
            st.caption(f"✅ 총 {count}개의 변환 구문이 포함됩니다."
                       if count else "⚠️ 변환할 변수명이 없어 RENAME 구문이 비어 있습니다.")
        with c2:
            blob, fname, _ = exports["map"]
            st.download_button("📄 매핑 테이블(XLSX)", blob, file_name=fname, mime=XLSX_MIME)
        with c3:
            if "data" in exports:
                blob, fname, _ = exports["data"]
                st.download_button("📊 변환된 데이터(XLSX)", blob, file_name=fname, mime=XLSX_MIME)
        with c4:
            if "sav" in exports:
                blob, fname, nvar = exports["sav"]
                st.download_button("💾 SPSS 데이터(.sav)", blob, file_name=fname,
                                   mime="application/octet-stream")
                labeled = (ss("spss_sav_info") or {}).get("value_labels", 0)
                st.caption(f"변수 {nvar}개 · 변수라벨 포함"
                           + (f" · 값라벨 {labeled}개" if labeled else " · 값라벨 없음"))
            else:
                st.button("💾 SPSS 데이터(.sav)", disabled=True,
                          help="변수명 오류를 먼저 해결하세요.")

        info = ss("spss_sav_info")
        if info and not info.get("value_labels"):
            st.info("생성된 .sav 에 값라벨이 없습니다. 위에서 원본 .sav 를 올리면 "
                    "응답 라벨을 그대로 가져옵니다. (또는 .sps 를 원본 .sav 에 실행하면 "
                    "값라벨이 유지된 채 변수명만 바뀝니다)")
        if info and info.get("not_in_source"):
            with st.expander(f"원본 .sav 에 없는 열 {len(info['not_in_source'])}개"):
                st.write(", ".join(info["not_in_source"][:50]))
        if info and info["auto_fixed"]:
            with st.expander(f"sav 생성 시 자동으로 정리한 열 이름 {len(info['auto_fixed'])}개"):
                st.write(", ".join(info["auto_fixed"][:50]))
            st.caption("SPSS 변수명 규칙에 맞지 않는 열(한글 헤더 등)은 자동으로 정리했습니다. "
                       "매핑 테이블에 반영되지 않으니 필요하면 표에서 직접 지정하세요.")
