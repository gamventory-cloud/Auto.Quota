# banner_table_engine.py
#
# 뱅크표(배너표) 계산 엔진입니다. 화면 코드는 뱅크표_생성.py 에 있습니다.
# 페이지 파일이 아니므로 Home.py 의 SKIP 목록에 넣어 주세요.
#
# 구성
#   1) TableBlock   — "표 하나"를 나타내는 구조 (축, 통계, 필터, 제목)
#   2) parse_sps()  — Embrain 'Table' 매크로 신텍스 → TableBlock 목록
#   3) build_block() — 화면에서 고른 변수 선택 → TableBlock
#   4) compute_table() — TableBlock + .sav → TableResult (숫자 원본)
#   5) result_to_frame() / write_tables_xlsx() — 화면 표시 / 엑셀 출력
#
# ── 표 방향 (중요) ────────────────────────────────────────────────────
# SPSS 산출물과 같은 방향입니다.
#
#            │ 사례수 │ 서울 │ 경기,인천 │ … │  계
#   ─────────┼────────┼──────┼───────────┼───┼──────
#   ■ 전체 ■ │ (400)  │  5.5 │      28.5 │ … │ 100
#   [권역] 서울 │ (22)  │100.0 │       0.0 │ … │ 100
#          경기,인천│(114)│  0.0 │     100.0 │ … │ 100
#   [지자체] 서울│ (22) │100.0 │       0.0 │ … │ 100
#            ⋮
#
#   행(왼쪽 두 칸) = 배너   → 신텍스 /table= 의 'by' 앞 목록
#   열(위)        = 문항 보기 → 'by' 뒤 목록 (사례수 + 보기 + 계)
#
#   % 는 각 배너 행의 사례수를 분모로 가로로 퍼지며 합이 100 이 됩니다.
#   (다중응답은 100 을 넘는 것이 정상입니다)
#
# ── 계산 방식 ─────────────────────────────────────────────────────────
# 원본 매크로 정의(DEFINE ... !ENDDEFINE)는 신텍스에 없지만, 실제 SPSS
# 산출물(시설물_TABLE_조건.xls)과 숫자를 대조해 아래를 확인했습니다.
#   · 전체(■ 전체 ■) 행도 %를 표시한다  → 확인됨
#   · '계'는 보기 칸들의 가로 합이다     → 확인됨 (% 표는 100, N 표는 사례수)
#   · m 접두 변수는 원본과 값이 같다     → 확인됨
# 연속형(평균 등) 표는 전체 행 라벨이 '사례수' 입니다. (val lab @t3 값)

from __future__ import annotations

import io
import math
import re
from dataclasses import dataclass, field

import pandas as pd

TOTAL_LABEL = "■ 전체 ■"

# 표 방향
BANNER_ROW = "banner_row"     # 배너가 왼쪽 행 (SPSS 산출물 기본)
BANNER_COL = "banner_col"     # 배너가 위쪽 열


# =============================================================================
# 1) 표 하나를 나타내는 구조
# =============================================================================
@dataclass
class StatSpec:
    """통계 하나. count / cpct / ValidN / MEAN / MEDIAN / MIN / MAX"""
    name: str
    var: str
    fmt: str
    label: str
    restrict: list[str] | None = None


@dataclass
class MergeSpec:
    """여러 변수를 하나의 축으로 묶은 것. 신텍스의 /mrg= 에 해당."""
    name: str
    label: str
    varlist_raw: str


@dataclass
class SummarySpec:
    """척도 요약 칸 하나. '계' 뒤에 붙는다.

    kind='group' — 지정한 코드들을 묶은 비율/사례수 (Top2, Bottom2, Middle 등)
    kind='mean'  — 보기 코드값의 평균 (1~5 척도의 평균점)
    kind='std'   — 코드값의 표준편차
    """
    label: str
    kind: str
    codes: list[float] = field(default_factory=list)
    decimals: int = 1


@dataclass
class SigSpec:
    """유의성 검정 설정.

    같은 배너 그룹 안의 세그먼트끼리 짝지어 비교하고, 유의하게 높은 쪽 칸에
    상대 세그먼트의 글자(a/b/c)를 적는다. 비율은 두 비율 z검정(합동분산),
    평균은 Welch t검정이다. '전체' 세그먼트는 다른 세그먼트를 포함하므로
    비교에서 빼고 글자도 주지 않는다.
    """
    enabled: bool = False
    level: float = 0.95                 # 0.95 | 0.99
    min_base: int = 30                  # 사례수가 이보다 작은 세그먼트는 검정 제외


@dataclass
class TableBlock:
    target_raw: str
    extra_cond: str | None
    total_label: str
    merges: list[MergeSpec]
    obser_var: str | None
    banner_axis: list[str]              # 배너 축 (권역/지자체/… 세그먼트)
    value_axis: list[str]               # 보기 축 (사례수 + 보기/통계 + 계)
    stats: list[StatSpec]
    title: str
    ptotal_labels: dict
    ftotal_labels: dict
    row_ma_mode: str = "category"       # 'category' | 'dummy'
    orientation: str = BANNER_ROW       # BANNER_ROW | BANNER_COL
    summaries: list = field(default_factory=list)   # list[SummarySpec]
    sig: SigSpec | None = None          # 유의성 검정 (없으면 안 함)
    min_base_show: int = 0              # 이보다 작은 세그먼트는 값을 '-' 로 (0=안 함)
    sort_values: bool = False           # 보기를 전체 % 내림차순으로 정렬
    # 척도 종합표(서머리 표): 행이 문항, 열이 보기 또는 배너.
    battery_vars: list[str] = field(default_factory=list)
    battery_metric: str | None = None   # None=보기 분포형 / 'mean'·'Top2' 등=격자형

    @property
    def is_battery(self) -> bool:
        return bool(self.battery_vars)

    @property
    def is_obser(self) -> bool:
        return self.obser_var is not None

    @property
    def row_var_merge(self) -> MergeSpec | None:
        for m in self.merges:
            if m.name == "m_down":
                return m
        return None


# =============================================================================
# 계산 결과 구조
# =============================================================================
@dataclass
class ValueColumn:
    """값 컬럼 하나. kind 로 엑셀 숫자 서식을 정한다."""
    label: str
    kind: str                           # 'pct' | 'count' | 'stat'
    decimals: int = 1

    @property
    def excel_format(self) -> str:
        return f"###0.{'0' * self.decimals}" if self.decimals else "###0"


@dataclass
class BannerRow:
    group: str                          # '[권역]' 또는 '' (전체 행)
    category: str                       # '서울' 또는 '' (전체 행)
    n: int


@dataclass
class TableResult:
    """계산 결과. 방향과 무관하게 항상 '배너 = rows, 보기 = columns' 로 담고,
    화면·엑셀에서 orientation 에 따라 그대로 쓰거나 뒤집어 쓴다."""
    title: str
    n_label: str                        # '사례수'
    rows: list[BannerRow]
    columns: list[ValueColumn]
    matrix: list[list]                  # rows × columns 의 숫자 (없으면 None)
    orientation: str = BANNER_ROW
    # 유의성 검정을 켰을 때만 채워진다.
    letters: list[str] = field(default_factory=list)      # 행(세그먼트)별 글자
    col_letters: list[str] = field(default_factory=list)  # 열별 글자 (종합표 격자형)
    marks: list[list[str]] = field(default_factory=list)  # 칸별 상대 글자
    notes: list[str] = field(default_factory=list)        # 계산 중 알릴 것
    hidden: list[bool] = field(default_factory=list)      # 소표본이라 값을 감춘 행
    row_kind: str = "banner"            # 'banner' | 'question' (척도 종합표)

    @property
    def row_axis_names(self) -> tuple[str, str]:
        return ("", "문항") if self.row_kind == "question" else ("배너", "구분")

    @property
    def rows_have_groups(self) -> bool:
        return any(r.group for r in self.rows)

    @property
    def banner_on_rows(self) -> bool:
        return self.orientation != BANNER_COL

    @property
    def has_marks(self) -> bool:
        return any(any(row) for row in self.marks)

    def mark_at(self, i: int, j: int) -> str:
        if not self.marks:
            return ""
        try:
            return self.marks[i][j]
        except IndexError:
            return ""


# =============================================================================
# 2) 신텍스(.sps) 읽기
# =============================================================================
_STAT_RE = re.compile(
    r"(?P<name>\w+)\s*\(\s*(?P<var>\S+)\s*\(\s*(?P<fmt>[^)]*)\)\s*'(?P<label>[^']*)'"
    r"(?:\s*:\s*(?P<restrict>[^)]*))?\)"
)
_MRG_RE = re.compile(r"^/mrg=(\S+)\s+'([^']*)'\s+(.+?)\s*/?\s*$", re.M)
_SELECT_RE = re.compile(
    r"Select if nval\(([^)]+)\)\s*>\s*0\s*(?:&\s*\(([^)]+)\))?\s*\.", re.S
)
_TOTLAB_RE = re.compile(r"val lab @t3 1'([^']*)'")
_OBSER_RE = re.compile(r"^/obser=(\S+)", re.M)
_TABLE_RE = re.compile(
    r"/table=([\s\S]*?)\bby\b([\s\S]*?)(?=/statistics=|/ptotal=|/ftotal=|\Z)"
)
_PTOTAL_RE = re.compile(r"^/(ptotal|ftotal)=(\S+)\s+'([^']*)'", re.M)
_TITLE_RE = re.compile(r"/title='([^']*)'")


def read_sps_text(file_bytes: bytes) -> str:
    """신텍스 파일 인코딩 자동 판별. Embrain 신텍스는 보통 CP949 입니다."""
    for enc in ("utf-8", "cp949", "euc-kr"):
        try:
            return file_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("cp949", errors="replace")


def _axis_tokens(axis_text: str) -> list[str]:
    return [p.strip() for p in axis_text.strip().split("+") if p.strip()]


_VALUE_MARKERS = {"t1", "t2", "m_down"}


def _is_value_axis(tokens: list[str], obser_var: str | None) -> bool:
    """사례수·계·보기(또는 연속형 변수)가 들어 있으면 보기 축이다."""
    return any(t in _VALUE_MARKERS or (obser_var and t == obser_var) for t in tokens)


def parse_sps(text: str) -> list[TableBlock]:
    blocks: list[TableBlock] = []

    for raw in re.split(r"(?m)^Temp\.\s*$", text)[1:]:
        sel = _SELECT_RE.search(raw)
        if not sel:
            continue

        merges = [
            MergeSpec(name=n, label=lbl, varlist_raw=vl.strip())
            for n, lbl, vl in _MRG_RE.findall(raw)
        ]

        table_m = _TABLE_RE.search(raw)
        axis_a = _axis_tokens(table_m.group(1)) if table_m else []
        axis_b = _axis_tokens(table_m.group(2)) if table_m else []

        stats: list[StatSpec] = []
        for chunk in re.findall(r"/statistics=([\s\S]*?)(?=/title=|/statistics=|\Z)", raw):
            for m in _STAT_RE.finditer(chunk):
                restrict = m.group("restrict")
                stats.append(
                    StatSpec(
                        name=m.group("name"),
                        var=m.group("var"),
                        fmt=m.group("fmt").strip(),
                        label=m.group("label"),
                        restrict=restrict.split() if restrict else None,
                    )
                )

        ptotal_labels: dict[str, str] = {}
        ftotal_labels: dict[str, str] = {}
        for kind, name, lbl in _PTOTAL_RE.findall(raw):
            (ptotal_labels if kind == "ptotal" else ftotal_labels)[name] = lbl

        totlab = _TOTLAB_RE.search(raw)
        title = _TITLE_RE.search(raw)
        obser = _OBSER_RE.search(raw)
        obser_var = obser.group(1) if obser else None

        # '/table=A by B' 에서 어느 쪽이 보기 축인지 표시로 판별한다.
        # 사례수(t2) · 계(t1) · 보기(m_down) · 연속형 변수가 있는 쪽이 보기 축.
        # 이렇게 해 두면 신텍스가 축을 반대로 써도 알아서 맞춘다.
        if _is_value_axis(axis_a, obser_var) and not _is_value_axis(axis_b, obser_var):
            axis_a, axis_b = axis_b, axis_a
        banner_axis, value_axis = axis_a, axis_b

        blocks.append(
            TableBlock(
                target_raw=sel.group(1).strip(),
                extra_cond=sel.group(2).strip() if sel.group(2) else None,
                total_label=totlab.group(1) if totlab else TOTAL_LABEL,
                merges=merges,
                obser_var=obser_var,
                banner_axis=banner_axis,
                value_axis=value_axis,
                stats=stats,
                title=title.group(1).strip() if title else "(제목 없음)",
                ptotal_labels=ptotal_labels,
                ftotal_labels=ftotal_labels,
            )
        )
    return blocks


# =============================================================================
# 3) 화면에서 고른 선택 → TableBlock
# =============================================================================
@dataclass
class BannerSpec:
    """배너 축 하나.
    kind='single' → 변수 하나의 값 라벨이 그대로 배너 행이 된다.
    kind='merge'  → 여러 변수를 하나의 다중응답 배너로 묶는다.
    """
    kind: str
    var: str | None = None
    label: str | None = None
    varlist: list[str] = field(default_factory=list)


_OBSER_LABELS = {"MEAN": "평균", "MEDIAN": "중위값", "MIN": "최소값", "MAX": "최대값"}


def build_battery_block(
    *,
    battery_vars: list[str],
    title: str,
    banners: list[BannerSpec] | None = None,
    metric: str | None = None,          # None=보기 분포형 / 'mean'·'Top2' 등
    summaries: list | None = None,
    show_pct: bool = True,
    decimals: int = 1,
    show_total_row: bool = True,
    extra_cond: str | None = None,
    orientation: str = BANNER_ROW,
    sig: SigSpec | None = None,
    min_base_show: int = 0,
    sort_rows: bool = False,
) -> TableBlock:
    """척도 종합표 하나를 만든다.

    metric 을 주면 격자형(행=문항, 열=배너)이고, 안 주면 보기 분포형
    (행=문항, 열=보기+계+요약)이다. 격자형에서만 배너를 쓴다.
    """
    merges: list[MergeSpec] = []
    banner_axis: list[str] = []
    if metric:
        banner_axis.append("@t3")
        for i, b in enumerate(banners or []):
            if b.kind == "single":
                banner_axis.append(b.var)
            else:
                token = f"bnr{i}"
                merges.append(MergeSpec(name=token, label=b.label or "",
                                        varlist_raw=" ".join(b.varlist)))
                banner_axis.append(token)

    stats: list[StatSpec] = []
    if show_pct:
        stats.append(StatSpec(name="cpct", var="battery", fmt=f"F.{decimals}",
                              label=""))
    value_axis = ["t1"] if (show_total_row and not metric) else []

    return TableBlock(
        target_raw=" ".join(battery_vars),
        extra_cond=extra_cond,
        total_label=TOTAL_LABEL,
        merges=merges,
        obser_var=None,
        banner_axis=banner_axis,
        value_axis=value_axis,
        stats=stats,
        title=title,
        ptotal_labels={"t2": "사례수"},
        ftotal_labels={"t1": "      계"},
        orientation=orientation,
        summaries=list(summaries or []),
        sig=sig,
        min_base_show=min_base_show,
        sort_values=sort_rows,
        battery_vars=list(battery_vars),
        battery_metric=metric,
    )


def build_block(
    *,
    row_type: str,                          # 'single' | 'multi' | 'obser'
    row_vars: list[str],
    banners: list[BannerSpec],
    title: str,
    row_ma_mode: str = "category",
    obser_stats: list[str] | None = None,
    extra_cond: str | None = None,
    show_pct: bool = True,
    decimals: int = 1,
    obser_decimals: int = 2,
    show_total_row: bool = True,
    orientation: str = BANNER_ROW,
    obser_show_values: bool = False,
    summaries: list | None = None,
    sig: SigSpec | None = None,
    min_base_show: int = 0,
    sort_values: bool = False,
) -> TableBlock:
    merges: list[MergeSpec] = []
    banner_axis: list[str] = ["@t3"]

    for i, b in enumerate(banners):
        if b.kind == "single":
            banner_axis.append(b.var)
        else:
            token = f"bnr{i}"
            merges.append(
                MergeSpec(name=token, label=b.label or "", varlist_raw=" ".join(b.varlist))
            )
            banner_axis.append(token)

    stats: list[StatSpec] = []
    value_axis: list[str] = ["t2"]

    if row_type == "obser":
        obser_var = row_vars[0]
        stats.append(StatSpec(name="ValidN", var="t2", fmt="paren5.0", label=""))

        # 응답값 분포를 함께 보여주면 단수 표와 같은 모양이 되고,
        # 그 뒤에 평균·중위값 같은 통계 칸이 붙는다.
        if obser_show_values:
            merges.append(MergeSpec(name="m_down", label="", varlist_raw=obser_var))
            if show_pct:
                stats.append(
                    StatSpec(name="cpct", var="m_down", fmt=f"F.{decimals}", label="")
                )
            value_axis.append("m_down")
            if show_total_row:
                value_axis.append("t1")
            total_label = TOTAL_LABEL
        else:
            total_label = "사례수"              # 통계만 있는 표는 SPSS 도 사례수

        for s in obser_stats or list(_OBSER_LABELS):
            stats.append(
                StatSpec(name=s, var=obser_var, fmt=f"F.{obser_decimals}",
                         label=_OBSER_LABELS.get(s, s))
            )
        value_axis.append(obser_var)
        # 척도 요약은 맨 뒤 — 단수 표에서 '계' 뒤에 오는 것과 같은 자리다.
        # (수치형은 평균·중위값 칸이 그 앞에 있어서 통계 뒤가 된다)
        if summaries:
            value_axis.append("summary")
        target_raw, obser_field = obser_var, obser_var
    else:
        stats.append(StatSpec(name="count", var="t2", fmt="paren5.0", label=""))
        merges.append(MergeSpec(name="m_down", label="", varlist_raw=" ".join(row_vars)))
        if show_pct:
            stats.append(StatSpec(name="cpct", var="m_down", fmt=f"F.{decimals}", label=""))
        value_axis.append("m_down")
        if show_total_row:
            value_axis.append("t1")
        # 척도 요약(Top2·Middle·Bottom2·평균)은 '계' 뒤에 붙는다
        if summaries:
            value_axis.append("summary")
        target_raw, obser_field = " ".join(row_vars), None
        total_label = TOTAL_LABEL

    return TableBlock(
        target_raw=target_raw,
        extra_cond=extra_cond,
        total_label=total_label,
        merges=merges,
        obser_var=obser_field,
        banner_axis=banner_axis,
        value_axis=value_axis,
        stats=stats,
        title=title,
        ptotal_labels={"t2": "사례수", "t3": total_label},
        ftotal_labels={"t1": "      계"},
        row_ma_mode=row_ma_mode,
        orientation=orientation,
        summaries=list(summaries or []),
        sig=sig,
        min_base_show=min_base_show,
        sort_values=sort_values,
    )


# =============================================================================
# 4) 계산
# =============================================================================
def expand_var_range(raw: str, columns: list[str]) -> list[str]:
    """'A to B' 는 파일 순서 기준 구간으로, 공백으로 나열된 것은 그 목록으로."""
    raw = raw.strip()
    if " to " in raw:
        start, end = [p.strip() for p in raw.split(" to ", 1)]
        i0, i1 = columns.index(start), columns.index(end)
        if i0 > i1:
            i0, i1 = i1, i0
        return list(columns[i0:i1 + 1])
    parts = raw.split()
    return parts if len(parts) > 1 else [raw]


def multi_category_map(df: pd.DataFrame, value_labels: dict, varlist: list[str]):
    """다중응답(카테고리 코딩): 변수마다 자기 코드값만 갖고 나머지는 결측."""
    out = []
    for v in varlist:
        vl = value_labels.get(v, {})
        codes = sorted(df[v].dropna().unique().tolist())
        if not codes:
            continue
        code = codes[0]
        out.append((v, code, vl.get(code, vl.get(float(code), str(code)))))
    return out


def title_with_marker(base: str, kind: str | None) -> str:
    """표 제목 뒤에 ' - %' 또는 ' - N' 표시를 붙인다.

    같은 문항으로 % 표와 N 표를 각각 만들면 이름이 같아져 목록·엑셀에서
    구분이 안 되므로, 표시 방식을 이름에 남긴다.
    이미 그 표시로 끝나는 제목이면 덧붙이지 않는다.
    kind: 'pct' | 'n' | None(붙이지 않음)
    """
    base = (base or "").strip()
    if not kind:
        return base
    mark = "%" if kind == "pct" else "N"
    tail = base.replace(" ", "").upper()
    if tail.endswith(f"-{mark.upper()}"):
        return base
    return f"{base} - {mark}" if base else mark


# =============================================================================
# 척도 요약 정의 읽기 (Top2 / Middle / Bottom2 / 평균)
# =============================================================================
_TOP_WORDS = ("상위", "상", "top", "t")
_BOTTOM_WORDS = ("하위", "하", "bottom", "b")
_MID_WORDS = ("중간", "중", "mid", "middle", "m")


def parse_summary_spec(raw: str, codes: list[float], *, decimals: int = 1,
                       mean_decimals: int = 2) -> tuple[list, list[str]]:
    """'상2,중,하2' / 'Top2,Bottom2' / '긍정=4,5' / '평균' 을 읽는다.

    codes 는 그 문항의 보기 코드들(작은 값부터). '상N' 은 큰 코드 N 개,
    '하N' 은 작은 코드 N 개, '중' 은 상·하에 안 들어간 나머지다.
    코드를 직접 적으려면 '이름=코드,코드' 로 쓴다.
    반환: (SummarySpec 목록, 문제 목록)
    """
    specs: list[SummarySpec] = []
    problems: list[str] = []
    raw = (raw or "").strip()
    if not raw:
        return specs, problems

    ordered = sorted(codes)
    used: set[float] = set()
    mid_slots: list[int] = []              # '중' 은 나머지를 알아야 하므로 뒤에 채운다

    # '이름=1,2' 처럼 등호가 있는 조각은 콤마로 잘리면 안 되므로 먼저 분리한다
    pieces: list[str] = []
    for chunk in raw.replace(";", "|").split("|"):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "=" in chunk:
            pieces.append(chunk)
        else:
            pieces.extend(p.strip() for p in chunk.split(",") if p.strip())

    for piece in pieces:
        low = piece.lower().replace(" ", "")

        if "=" in piece:                   # 이름=코드 목록
            name, codelist = piece.split("=", 1)
            picked: list[float] = []
            for tok in codelist.replace(";", ",").split(","):
                tok = tok.strip()
                if not tok:
                    continue
                try:
                    picked.append(float(tok))
                except ValueError:
                    problems.append(f"요약 '{piece}': 코드 '{tok}' 가 숫자가 아닙니다.")
            if picked:
                specs.append(SummarySpec(name.strip() or "묶음", "group",
                                         picked, decimals))
                used.update(picked)
            continue

        if low in ("평균", "mean", "avg"):
            specs.append(SummarySpec("평균", "mean", [], mean_decimals))
            continue
        if low in ("표준편차", "std", "sd"):
            specs.append(SummarySpec("표준편차", "std", [], mean_decimals))
            continue

        m = re.match(r"^([가-힣a-z]+)(\d*)$", low)
        if not m:
            problems.append(f"요약 '{piece}' 를 알 수 없습니다. 예: 상2, 하2, 중, 평균")
            continue
        word, num = m.group(1), m.group(2)
        count = int(num) if num else 1

        if word in _TOP_WORDS:
            picked = ordered[-count:] if count <= len(ordered) else list(ordered)
            specs.append(SummarySpec(f"Top{len(picked)}", "group", picked, decimals))
            used.update(picked)
        elif word in _BOTTOM_WORDS:
            picked = ordered[:count] if count <= len(ordered) else list(ordered)
            specs.append(SummarySpec(f"Bottom{len(picked)}", "group", picked, decimals))
            used.update(picked)
        elif word in _MID_WORDS:
            specs.append(SummarySpec("Middle", "group", [], decimals))
            mid_slots.append(len(specs) - 1)
        else:
            problems.append(f"요약 '{piece}' 를 알 수 없습니다. 예: 상2, 하2, 중, 평균")

    # 상·하가 겹치면 같은 보기가 두 묶음에 들어간다 — 대개 실수다
    tops = [x for x in specs if x.label.startswith("Top")]
    bots = [x for x in specs if x.label.startswith("Bottom")]
    for t in tops:
        for b in bots:
            both = sorted(set(t.codes) & set(b.codes))
            if both:
                problems.append(
                    f"요약: {t.label} 과 {b.label} 이 보기 {both} 를 함께 씁니다. "
                    f"보기가 {len(ordered)}개인데 묶음이 너무 큽니다."
                )

    # '중' 은 상·하에 안 들어간 코드들
    for idx in mid_slots:
        rest = [c for c in ordered if c not in used]
        specs[idx].codes = rest
        if not rest:
            problems.append("요약 '중': 상·하 묶음이 보기를 다 써서 남은 보기가 없습니다.")

    return specs, problems


def _value_label(code) -> str:
    """값 라벨이 없는 변수의 보기 이름 — 응답된 숫자를 그대로 쓴다."""
    try:
        f = float(code)
    except (TypeError, ValueError):
        return str(code)
    return f"{int(f):,}" if f.is_integer() else f"{f:,}"


def _decimals_of(fmt: str, default: int = 1) -> int:
    try:
        return int((fmt or "").split(".")[-1])
    except ValueError:
        return default


# =============================================================================
# 유의성 검정
# =============================================================================
# scipy 를 쓰지 않습니다. requirements.txt 를 늘리지 않으려고 필요한 분포
# 함수만 직접 넣었습니다. 정확도는 scipy 와 대조해 확인했습니다
# (표준정규 1e-15, t분포 1e-12 이내).

def _norm_sf(z: float) -> float:
    """표준정규분포의 오른쪽 꼬리 확률."""
    return 0.5 * math.erfc(z / math.sqrt(2.0))


def _betacf(a: float, b: float, x: float) -> float:
    """정규화 불완전베타함수의 연분수 (Lentz 방법)."""
    tiny, eps, max_iter = 1e-30, 3e-16, 300
    qab, qap, qam = a + b, a + 1.0, a - 1.0
    c = 1.0
    d = 1.0 - qab * x / qap
    if abs(d) < tiny:
        d = tiny
    d = 1.0 / d
    h = d
    for m in range(1, max_iter + 1):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < tiny:
            d = tiny
        c = 1.0 + aa / c
        if abs(c) < tiny:
            c = tiny
        d = 1.0 / d
        h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < tiny:
            d = tiny
        c = 1.0 + aa / c
        if abs(c) < tiny:
            c = tiny
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1.0) < eps:
            break
    return h


def _betai(a: float, b: float, x: float) -> float:
    """정규화 불완전베타함수 I_x(a, b)."""
    if x <= 0.0:
        return 0.0
    if x >= 1.0:
        return 1.0
    lbeta = (math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b)
             + a * math.log(x) + b * math.log1p(-x))
    front = math.exp(lbeta)
    if x < (a + 1.0) / (a + b + 2.0):
        return front * _betacf(a, b, x) / a
    return 1.0 - math.exp(
        math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b)
        + b * math.log1p(-x) + a * math.log(x)
    ) * _betacf(b, a, 1.0 - x) / b


def _t_sf(t: float, dof: float) -> float:
    """t분포의 오른쪽 꼬리 확률."""
    if dof <= 0:
        return float("nan")
    x = dof / (dof + t * t)
    half = 0.5 * _betai(0.5 * dof, 0.5, x)
    return half if t > 0 else 1.0 - half


def prop_test(x1: int, n1: int, x2: int, n2: int) -> float:
    """두 비율 차이의 양쪽 p값 (합동분산 z검정). 못 하면 nan."""
    if n1 <= 0 or n2 <= 0:
        return float("nan")
    p1, p2 = x1 / n1, x2 / n2
    p = (x1 + x2) / (n1 + n2)
    if p <= 0.0 or p >= 1.0:
        return float("nan")           # 두 집단 모두 0% 거나 모두 100%
    se = math.sqrt(p * (1.0 - p) * (1.0 / n1 + 1.0 / n2))
    if se == 0.0:
        return float("nan")
    return 2.0 * _norm_sf(abs((p1 - p2) / se))


def mean_test(m1: float, sd1: float, n1: int,
              m2: float, sd2: float, n2: int) -> float:
    """두 평균 차이의 양쪽 p값 (Welch t검정). 못 하면 nan.

    두 집단 모두 값이 하나뿐(표준편차 0)이면 nan 을 준다. 수식대로면 t 가
    무한대라 'p=0, 유의함' 이 되지만, 그건 자료가 상수라서 생기는 허수이지
    검정 결과가 아니다. (배너 변수와 문항 변수가 같은 표에서만 생긴다 —
    예: '권역' 표를 '권역' 배너로 돌리면 각 칸 안의 권역 값은 상수다.
    scipy 도 이 경우 결과를 믿지 말라는 경고를 낸다.)
    """
    if n1 < 2 or n2 < 2:
        return float("nan")
    v1, v2 = sd1 * sd1 / n1, sd2 * sd2 / n2
    denom = v1 + v2
    if denom <= 0.0:
        return float("nan")
    t = (m1 - m2) / math.sqrt(denom)
    dof = denom * denom / (v1 * v1 / (n1 - 1) + v2 * v2 / (n2 - 1))
    return 2.0 * _t_sf(abs(t), dof)


_SIG_LETTERS = "abcdefghijklmnopqrstuvwxyz"


def _segment_letters(rows: list[BannerRow], total_label: str) -> list[str]:
    """배너 그룹마다 a, b, c … 를 새로 매긴다.

    비교는 같은 그룹 안에서만 하므로 그룹마다 다시 시작하는 편이 읽기 쉽다.
    '전체' 행과, 세그먼트가 하나뿐인 그룹은 비교 상대가 없어 글자를 주지 않는다.
    """
    counts: dict[str, int] = {}
    for r in rows:
        if r.group == total_label and not r.category:
            continue
        counts[r.group] = counts.get(r.group, 0) + 1

    letters: list[str] = []
    used: dict[str, int] = {}
    for r in rows:
        if r.group == total_label and not r.category:
            letters.append("")
            continue
        if counts.get(r.group, 0) < 2:
            letters.append("")
            continue
        i = used.get(r.group, 0)
        letters.append(_SIG_LETTERS[i] if i < len(_SIG_LETTERS) else "")
        used[r.group] = i + 1
    return letters


def _compute_marks(rows, letters, columns, col_tests, sig: SigSpec):
    """칸마다 '이 칸이 유의하게 높은 상대들의 글자' 를 만든다."""
    alpha = 1.0 - sig.level
    marks = [["" for _ in columns] for _ in rows]

    # 같은 그룹 + 글자가 있는 행끼리만 비교한다
    by_group: dict[str, list[int]] = {}
    for i, (r, ltr) in enumerate(zip(rows, letters)):
        if ltr and r.n >= sig.min_base:
            by_group.setdefault(r.group, []).append(i)

    for j, test in enumerate(col_tests):
        if not test:
            continue
        for idxs in by_group.values():
            for i in idxs:
                hits = []
                for k in idxs:
                    if k == i:
                        continue
                    if test["kind"] == "prop":
                        p = prop_test(test["x"][i], test["n"][i],
                                      test["x"][k], test["n"][k])
                        higher = (test["x"][i] / max(test["n"][i], 1)
                                  > test["x"][k] / max(test["n"][k], 1))
                    else:
                        m_i, m_k = test["m"][i], test["m"][k]
                        if m_i is None or m_k is None:
                            continue
                        p = mean_test(m_i, test["sd"][i], test["n"][i],
                                      m_k, test["sd"][k], test["n"][k])
                        higher = m_i > m_k
                    if not math.isnan(p) and p < alpha and higher:
                        hits.append(letters[k])
                marks[i][j] = "".join(sorted(hits))
    return marks


# =============================================================================
# 보기 정렬
# =============================================================================
# '기타', '모름', '무응답' 계열은 응답이 많아도 맨 아래에 두는 것이 관례입니다.
_TAIL_WORDS = ("기타", "그 외", "그외", "모름", "무응답", "없음", "해당 없음",
               "해당없음", "잘 모르", "잘모르", "비해당", "dk", "na", "etc")


def _is_tail_label(label: str) -> bool:
    t = str(label).strip().lower()
    return any(w in t for w in _TAIL_WORDS)


def _sort_categories(cat_cols: list[tuple]) -> list[tuple]:
    """(라벨, 세그먼트별 사례수) 목록을 전체 응답이 많은 순으로 정렬한다.

    전체 사례수는 첫 세그먼트('전체' 행)가 아니라 모든 세그먼트를 합친 값으로
    본다. 배너에 '전체' 를 안 넣은 표에서도 같은 기준이 되도록.
    기타·모름 계열은 원래 순서를 유지한 채 맨 뒤로 보낸다.
    """
    body = [(lbl, cnt) for lbl, cnt in cat_cols if not _is_tail_label(lbl)]
    tail = [(lbl, cnt) for lbl, cnt in cat_cols if _is_tail_label(lbl)]
    body.sort(key=lambda p: -sum(p[1]))
    return body + tail


class _Segment:
    """배너 행 하나 + 그 행에 해당하는 응답자 표시."""
    __slots__ = ("group", "label", "mask")

    def __init__(self, group: str, label: str, mask: pd.Series):
        self.group = group
        self.label = label
        self.mask = mask


def _build_segments(dff: pd.DataFrame, meta, block: TableBlock) -> list[_Segment]:
    """신텍스 /table= 의 'by' 앞 목록 → 배너 행들."""
    columns_order = list(dff.columns)
    value_labels = meta.variable_value_labels
    col_labels = meta.column_names_to_labels
    merges_by_name = {m.name: m for m in block.merges}

    segments: list[_Segment] = []
    for token in block.banner_axis:
        if token == "@t3":
            segments.append(
                _Segment(block.total_label, "", pd.Series(True, index=dff.index))
            )
        elif token in dff.columns:
            vl = value_labels.get(token, {})
            group = col_labels.get(token) or token
            for code in sorted(vl.keys()):
                segments.append(_Segment(group, vl[code], dff[token] == code))
        elif token in merges_by_name:
            m = merges_by_name[token]
            varlist = expand_var_range(m.varlist_raw, columns_order)
            for var, code, label in multi_category_map(dff, value_labels, varlist):
                segments.append(_Segment(m.label, label, dff[var] == code))
    return segments


def select_mask(df: pd.DataFrame, block: TableBlock) -> pd.Series:
    """이 표가 쓰는 응답자만 골라내는 표시. (신텍스의 Temp. + Select if)"""
    columns_order = list(df.columns)
    mask = pd.Series(False, index=df.index)
    for v in expand_var_range(block.target_raw, columns_order):
        mask = mask | df[v].notna()
    if block.extra_cond:
        var, val = [p.strip() for p in block.extra_cond.split("=")]
        mask = mask & (df[var] == float(val))
    return mask


def value_breakdown_size(df: pd.DataFrame, block: TableBlock) -> int:
    """응답값 분포를 보여줄 때 보기가 몇 개 나올지 미리 센다.
    값이 수백 개인 변수를 골랐을 때 미리 알려주기 위한 것."""
    merge = block.row_var_merge
    if merge is None:
        return 0
    varlist = expand_var_range(merge.varlist_raw, list(df.columns))
    if len(varlist) != 1 or varlist[0] not in df.columns:
        return 0
    dff = df[select_mask(df, block)]
    return int(dff[varlist[0]].dropna().nunique())


def compute_table(df: pd.DataFrame, meta, block: TableBlock) -> TableResult:
    """표 하나를 계산한다. 결과는 서식 없는 숫자 + 컬럼 정의."""
    if block.is_battery:
        return compute_battery(df, meta, block)
    columns_order = list(df.columns)
    value_labels = meta.variable_value_labels

    dff = df[select_mask(df, block)].copy()

    segments = _build_segments(dff, meta, block)

    # 사례수가 0 인 배너 행은 뺀다. (SPSS 산출물과 같은 동작 —
    # 예: 공설만 걸러낸 표에는 '[주체] 사설' 행이 아예 없다)
    segments = [s for s in segments if int(s.mask.sum()) > 0]

    rows = [BannerRow(s.group, s.label, int(s.mask.sum())) for s in segments]
    bases = [max(r.n, 1) for r in rows]

    stat_by_var: dict[str, list] = {}
    for s in block.stats:
        stat_by_var.setdefault(s.var, []).append(s)

    columns: list[ValueColumn] = []
    cols_data: list[list] = []           # 컬럼별 값 목록
    # 컬럼별 검정 재료. 서식과 무관한 원자료(분자·분모 또는 n·평균·표준편차)를
    # 따로 모아 둔다. 검정을 안 켜면 쓰이지 않고, %인지 N인지에 영향받지 않는다.
    col_tests: list[dict | None] = []
    total_col_start = None               # '계' 계산에 쓸 시작 위치
    ns = [r.n for r in rows]

    def _mean_material(var: str, codes: list[float] | None = None):
        """세그먼트별 (n, 평균, 표준편차) — 평균 검정 재료."""
        m_, sd_, n_ = [], [], []
        for seg in segments:
            s = dff.loc[seg.mask, var].dropna()
            if codes is not None:
                s = s[s.isin(codes)]
            n_.append(int(len(s)))
            m_.append(float(s.mean()) if len(s) else None)
            sd_.append(float(s.std()) if len(s) > 1 else 0.0)
        return {"kind": "mean", "n": n_, "m": m_, "sd": sd_}

    for token in block.value_axis:
        # ── 사례수 열은 rows 에 이미 있으므로 여기서는 건너뛴다 ──
        if token == "t2":
            total_col_start = len(columns)
            continue

        # ── 연속형: 평균 / 중위값 / 최소값 / 최대값 ──
        if block.is_obser and token == block.obser_var:
            for s in stat_by_var.get(token, []):
                vals = []
                for seg in segments:
                    series = dff.loc[seg.mask, token]
                    if s.name == "MEAN":
                        v = series.mean()
                    elif s.name == "MEDIAN":
                        v = series.median()
                    elif s.name == "MIN":
                        v = series.min() if len(series) else None
                    elif s.name == "MAX":
                        v = series.max() if len(series) else None
                    else:
                        v = None
                    vals.append(None if v is None or pd.isna(v) else float(v))
                columns.append(
                    ValueColumn(s.label, "stat", _decimals_of(s.fmt, 2))
                )
                cols_data.append(vals)
                # 평균만 검정한다. 중위값·최소·최대는 표준 검정이 없다.
                col_tests.append(_mean_material(token) if s.name == "MEAN" else None)

        # ── 단일 / 다중응답 보기별 열 ──
        elif block.row_var_merge is not None and token == block.row_var_merge.name:
            merge = block.row_var_merge
            varlist = expand_var_range(merge.varlist_raw, columns_order)
            pct_stat = next(
                (s for s in stat_by_var.get(token, []) if s.name == "cpct"), None
            )

            if len(varlist) == 1 and varlist[0] in df.columns:
                var = varlist[0]
                vl = value_labels.get(var, {})
                if vl:
                    cats = [(var, code, vl[code]) for code in sorted(vl.keys())]
                else:
                    # 값 라벨이 없는 변수(수치형)는 실제 응답된 값을 보기로 쓴다.
                    # 걸러낸 응답자 안에 있는 값만 — 아무도 답하지 않은 값은 빼려고.
                    cats = [
                        (var, code, _value_label(code))
                        for code in sorted(dff[var].dropna().unique().tolist())
                    ]
            elif block.row_ma_mode == "dummy":
                labels = meta.column_names_to_labels
                cats = [(v, 1, labels.get(v) or v) for v in varlist]
            else:
                cats = multi_category_map(dff, value_labels, varlist)

            cat_cols: list[tuple] = []          # (label, counts) — 정렬용으로 모은다
            for var, code, label in cats:
                counts = [int((dff.loc[seg.mask, var] == code).sum()) for seg in segments]
                cat_cols.append((label, counts))

            # ── 보기 정렬: 전체 기준 응답이 많은 순 ──
            if block.sort_values:
                cat_cols = _sort_categories(cat_cols)

            for label, counts in cat_cols:
                if pct_stat is not None:
                    cols_data.append([c / b * 100 for c, b in zip(counts, bases)])
                    columns.append(
                        ValueColumn(label, "pct", _decimals_of(pct_stat.fmt, 1))
                    )
                else:
                    cols_data.append([float(c) for c in counts])
                    columns.append(ValueColumn(label, "count", 0))
                col_tests.append({"kind": "prop", "x": list(counts), "n": list(ns)})

        # ── '계' 열: 보기 열들의 가로 합 ──
        elif token == "t1":
            start = total_col_start or 0
            body = cols_data[start:]
            if body:
                totals = [
                    sum((col[i] or 0.0) for col in body) for i in range(len(rows))
                ]
                kind = columns[start].kind if columns[start:] else "pct"
                dec = columns[start].decimals if columns[start:] else 1
                columns.append(
                    ValueColumn(block.ftotal_labels.get(token, "      계"), kind, dec)
                )
                cols_data.append(totals)
                col_tests.append(None)      # '계' 는 검정 대상이 아니다

        # ── 척도 요약: Top2 · Middle · Bottom2 · 평균 ('계' 뒤에 붙는다) ──
        elif token == "summary" and block.summaries:
            merge = block.row_var_merge
            if merge is not None:
                varlist = expand_var_range(merge.varlist_raw, columns_order)
                as_pct = any(s.name == "cpct" for s in stat_by_var.get(merge.name, []))
            else:
                # 분포를 안 보여주는 수치형 표 — 요약은 변수 값으로 바로 계산한다
                varlist = [block.obser_var] if block.obser_var else []
                as_pct = False
            if len(varlist) != 1 or varlist[0] not in df.columns:
                continue                   # 요약은 단수·수치형 문항에서만 뜻이 있다
            var = varlist[0]

            for spec in block.summaries:
                vals, hits = [], []
                for seg, base in zip(segments, bases):
                    series = dff.loc[seg.mask, var].dropna()
                    if spec.kind == "group":
                        hit = int(series.isin(spec.codes).sum())
                        hits.append(hit)
                        vals.append(hit / base * 100 if as_pct else float(hit))
                    elif spec.kind == "mean":
                        vals.append(float(series.mean()) if len(series) else None)
                    elif spec.kind == "std":
                        vals.append(float(series.std()) if len(series) > 1 else None)
                    else:
                        vals.append(None)
                columns.append(
                    ValueColumn(
                        spec.label,
                        "pct" if (spec.kind == "group" and as_pct) else
                        ("count" if spec.kind == "group" else "stat"),
                        spec.decimals,
                    )
                )
                cols_data.append(vals)
                if spec.kind == "group":
                    col_tests.append({"kind": "prop", "x": hits, "n": list(ns)})
                elif spec.kind == "mean":
                    col_tests.append(_mean_material(var))
                else:
                    col_tests.append(None)     # 표준편차는 검정하지 않는다

    matrix = [[cols_data[j][i] for j in range(len(columns))] for i in range(len(rows))]
    notes: list[str] = []

    # ── 유의성 검정 ──
    letters: list[str] = []
    marks: list[list[str]] = []
    if block.sig and block.sig.enabled:
        # 소표본이라 값을 감춘 배너는 검정에서도 뺀다. 감춰서 안 보이는 행의
        # 글자가 다른 칸에 상대로 적히면 읽는 사람이 확인할 방법이 없다.
        eff = SigSpec(
            enabled=True, level=block.sig.level,
            min_base=max(block.sig.min_base, block.min_base_show),
        )
        letters = _segment_letters(rows, block.total_label)
        marks = _compute_marks(rows, letters, columns, col_tests, eff)
        # 값을 감추는 기준과 검정 기준이 같으면 아래 '감춘 배너' 안내로 갈음한다
        if eff.min_base > block.min_base_show:
            small = [r.category or r.group for r, l in zip(rows, letters)
                     if l and r.n < eff.min_base]
            if small:
                notes.append(
                    f"사례수가 {eff.min_base}명 미만이라 검정에서 뺀 배너: "
                    + ", ".join(small)
                )
        if any(m.name.startswith("bnr") for m in block.merges):
            notes.append(
                "다중응답 배너는 한 응답자가 여러 세그먼트에 들어갈 수 있어 "
                "세그먼트끼리 독립이 아닙니다. 그 배너의 검정 결과는 참고로만 보세요."
            )

    # ── 소표본 감추기: 사례수가 기준 미만인 배너 행은 값을 비운다 ──
    # 검정보다 나중에 한다. 지우기 전 숫자로 검정을 마쳐야 하기 때문이 아니라,
    # 어차피 검정에서도 같은 행을 빼기 때문에 순서와 무관하게 결과가 같다.
    hidden = [False] * len(rows)
    if block.min_base_show > 0:
        names = []
        for i, r in enumerate(rows):
            if r.n < block.min_base_show:
                matrix[i] = [None] * len(columns)
                hidden[i] = True
                if marks:
                    marks[i] = [""] * len(columns)
                names.append(r.category or r.group)
        if names:
            tail = " (검정에서도 뺐습니다)" if (block.sig and block.sig.enabled) else ""
            notes.append(
                f"사례수 {block.min_base_show}명 미만이라 값을 감춘 배너: "
                + ", ".join(names) + tail
            )

    return TableResult(
        title=block.title,
        n_label=block.ptotal_labels.get("t2", "사례수"),
        rows=rows,
        columns=columns,
        matrix=matrix,
        orientation=block.orientation,
        letters=letters,
        marks=marks,
        notes=notes,
        hidden=hidden,
    )


# =============================================================================
# 4-2) 척도 종합표 (서머리 표)
# =============================================================================
# 문항 여러 개를 한 표에 세로로 쌓는 표입니다. 두 가지 모양이 있습니다.
#
#  (가) 보기 분포형 — 열이 보기 + 계 + 요약
#       문항                    사례수  전혀  아니  보통  그렇  매우    계  Top2  평균
#       Q5-1. 시설이 깨끗하다    (200)  10.5  20.5  27.0  26.5  15.5  100  42.0  3.16
#       Q5-2. 직원이 친절하다    (200)   4.5  15.5  34.5  22.0  23.5  100  45.5  3.44
#
#  (나) 격자형 — 열이 배너, 값은 지표 하나(평균이나 Top2)
#       문항                    전체   남성   여성   20대   30대
#       Q5-1. 시설이 깨끗하다    3.16   3.18   3.14   3.30   3.11
#       Q5-2. 직원이 친절하다    3.44   3.37   3.51   3.52   3.52
#
# 유의성 검정은 (나) 에서만 합니다. (가) 는 행끼리 비교하는 표인데 같은
# 응답자가 모든 문항에 답했으므로 독립표본 검정을 쓰면 안 됩니다.

def _battery_labels(vars_: list[str], value_labels: dict) -> tuple[list, list[str]]:
    """문항들이 공유하는 보기 목록. 서로 다르면 합집합을 쓰고 알린다."""
    sets, codes = [], {}
    for v in vars_:
        vl = value_labels.get(v, {})
        sets.append(tuple(sorted(vl.keys())))
        codes.update(vl)
    notes = []
    if len(set(sets)) > 1:
        notes.append(
            "문항마다 보기가 달라서 합집합을 썼습니다. 어떤 문항에 없는 보기는 "
            "0 으로 나옵니다 — 척도가 같은 문항끼리 묶는 것이 좋습니다."
        )
    return sorted(codes.items()), notes


def _compute_marks_by_col(rows, col_letters, columns, cell_tests, sig: SigSpec):
    """격자형 종합표용 — 한 행 안에서 열끼리 비교한다.

    cell_tests[i] 는 그 행의 열별 검정 재료다. 열(배너 세그먼트)에 글자를
    주고, 유의하게 높은 칸에 상대 열의 글자를 적는다.
    """
    alpha = 1.0 - sig.level
    marks = [["" for _ in columns] for _ in rows]
    for i in range(len(rows)):
        test = cell_tests[i]
        if not test:
            continue
        idxs = [j for j, ltr in enumerate(col_letters)
                if ltr and test["n"][j] >= sig.min_base]
        for j in idxs:
            hits = []
            for k in idxs:
                if k == j:
                    continue
                if test["kind"] == "prop":
                    p = prop_test(test["x"][j], test["n"][j],
                                  test["x"][k], test["n"][k])
                    higher = (test["x"][j] / max(test["n"][j], 1)
                              > test["x"][k] / max(test["n"][k], 1))
                else:
                    m_j, m_k = test["m"][j], test["m"][k]
                    if m_j is None or m_k is None:
                        continue
                    p = mean_test(m_j, test["sd"][j], test["n"][j],
                                  m_k, test["sd"][k], test["n"][k])
                    higher = m_j > m_k
                if not math.isnan(p) and p < alpha and higher:
                    hits.append(col_letters[k])
            marks[i][j] = "".join(sorted(hits))
    return marks


def compute_battery(df: pd.DataFrame, meta, block: TableBlock) -> TableResult:
    """척도 종합표를 계산한다."""
    value_labels = meta.variable_value_labels
    col_labels = meta.column_names_to_labels
    notes: list[str] = []

    dff = df[select_mask(df, block)].copy()
    vars_ = [v for v in block.battery_vars if v in df.columns]
    if not vars_:
        return TableResult(title=block.title, n_label="사례수", rows=[],
                           columns=[], matrix=[], orientation=block.orientation,
                           notes=["종합표에 쓸 문항 변수를 찾지 못했습니다."])

    cats, lab_notes = _battery_labels(vars_, value_labels)
    notes.extend(lab_notes)

    rows = [BannerRow("", col_labels.get(v) or v, int(dff[v].notna().sum()))
            for v in vars_]
    bases = [max(r.n, 1) for r in rows]
    as_pct = any(s.name == "cpct" for s in block.stats)

    columns: list[ValueColumn] = []
    matrix: list[list] = [[] for _ in vars_]
    letters: list[str] = []
    col_letters: list[str] = []
    marks: list[list[str]] = []

    # ── (나) 격자형: 열이 배너, 값은 지표 하나 ──
    if block.battery_metric:
        segments = [s for s in _build_segments(dff, meta, block)
                    if int(s.mask.sum()) > 0]
        spec = next((x for x in block.summaries
                     if x.label == block.battery_metric), None)
        kind = "stat" if block.battery_metric in ("mean", "std") else (
            "pct" if as_pct else "count")
        dec = spec.decimals if spec else (2 if block.battery_metric == "mean" else 1)

        for seg in segments:
            label = f"{seg.group} {seg.label}".strip() if seg.label else seg.group
            columns.append(ValueColumn(label, kind, dec))

        cell_tests: list[dict | None] = []
        for i, v in enumerate(vars_):
            xs, ns, ms, sds = [], [], [], []
            for seg in segments:
                s = dff.loc[seg.mask, v].dropna()
                ns.append(int(len(s)))
                if block.battery_metric == "mean":
                    val = float(s.mean()) if len(s) else None
                    ms.append(val)
                    sds.append(float(s.std()) if len(s) > 1 else 0.0)
                    xs.append(0)
                elif block.battery_metric == "std":
                    val = float(s.std()) if len(s) > 1 else None
                    ms.append(None)
                    sds.append(0.0)
                    xs.append(0)
                else:
                    hit = int(s.isin(spec.codes).sum()) if spec else 0
                    xs.append(hit)
                    ms.append(None)
                    sds.append(0.0)
                    val = (hit / max(len(s), 1) * 100) if as_pct else float(hit)
                matrix[i].append(val)
            if block.battery_metric == "mean":
                cell_tests.append({"kind": "mean", "n": ns, "m": ms, "sd": sds})
            elif block.battery_metric == "std":
                cell_tests.append(None)
            else:
                cell_tests.append({"kind": "prop", "x": xs, "n": ns})

        if block.sig and block.sig.enabled:
            col_letters = _letters_for_columns(segments, block.total_label)
            marks = _compute_marks_by_col(rows, col_letters, columns,
                                          cell_tests, block.sig)

    # ── (가) 보기 분포형: 열이 보기 + 계 + 요약 ──
    else:
        for code, label in cats:
            columns.append(ValueColumn(label, "pct" if as_pct else "count",
                                       1 if as_pct else 0))
        for i, v in enumerate(vars_):
            s = dff[v].dropna()
            for code, _label in cats:
                hit = int((s == code).sum())
                matrix[i].append(hit / bases[i] * 100 if as_pct else float(hit))

        if block.sort_values:
            order = _battery_row_order(vars_, dff, block, cats)
            vars_ = [vars_[i] for i in order]
            rows = [rows[i] for i in order]
            bases = [bases[i] for i in order]
            matrix = [matrix[i] for i in order]

        if "t1" in block.value_axis:
            columns.append(ValueColumn("      계", "pct" if as_pct else "count",
                                       1 if as_pct else 0))
            for i in range(len(vars_)):
                matrix[i].append(sum(x or 0.0 for x in matrix[i]))

        for spec in block.summaries:
            columns.append(ValueColumn(
                spec.label,
                "pct" if (spec.kind == "group" and as_pct) else
                ("count" if spec.kind == "group" else "stat"),
                spec.decimals,
            ))
            for i, v in enumerate(vars_):
                s = dff[v].dropna()
                if spec.kind == "group":
                    hit = int(s.isin(spec.codes).sum())
                    matrix[i].append(hit / bases[i] * 100 if as_pct else float(hit))
                elif spec.kind == "mean":
                    matrix[i].append(float(s.mean()) if len(s) else None)
                elif spec.kind == "std":
                    matrix[i].append(float(s.std()) if len(s) > 1 else None)
                else:
                    matrix[i].append(None)

        if block.sig and block.sig.enabled:
            notes.append(
                "보기 분포형 종합표는 유의성 검정을 하지 않습니다. 행끼리 비교하는 "
                "표인데 같은 응답자가 모든 문항에 답했으므로 독립표본 검정을 쓰면 "
                "안 됩니다. 문항별 표에서 배너끼리 비교하세요."
            )

    hidden = [False] * len(rows)
    if block.min_base_show > 0:
        names = []
        for i, r in enumerate(rows):
            if r.n < block.min_base_show:
                matrix[i] = [None] * len(columns)
                hidden[i] = True
                if marks:
                    marks[i] = [""] * len(columns)
                names.append(r.category)
        if names:
            notes.append(
                f"사례수 {block.min_base_show}명 미만이라 값을 감춘 문항: "
                + ", ".join(names)
            )

    return TableResult(
        title=block.title,
        n_label="사례수",
        rows=rows,
        columns=columns,
        matrix=matrix,
        orientation=block.orientation,
        letters=letters,
        col_letters=col_letters,
        marks=marks,
        notes=notes,
        hidden=hidden,
        row_kind="question",
    )


# =============================================================================
# 4-3) 차수 비교
# =============================================================================
def diff_result(now: TableResult, before: TableResult, *,
                title: str) -> TableResult:
    """두 차수의 같은 표에서 증감표를 만든다 — 이번 값 − 지난 값.

    행(배너)과 열(보기)은 이름으로 맞춥니다. 한쪽에만 있는 행·열은 비교할 수
    없으니 빼고, 무엇을 뺐는지 알립니다. 위치로 맞추면 보기가 하나 늘거나
    줄었을 때 엉뚱한 값끼리 빼는 일이 생깁니다.
    """
    notes: list[str] = []
    before_rows = {(r.group, r.category): i for i, r in enumerate(before.rows)}
    before_cols = {c.label: j for j, c in enumerate(before.columns)}

    keep_rows = [(i, before_rows[(r.group, r.category)])
                 for i, r in enumerate(now.rows)
                 if (r.group, r.category) in before_rows]
    keep_cols = [(j, before_cols[c.label])
                 for j, c in enumerate(now.columns) if c.label in before_cols]

    dropped_r = [r.category or r.group for i, r in enumerate(now.rows)
                 if (r.group, r.category) not in before_rows]
    dropped_c = [c.label.strip() for j, c in enumerate(now.columns)
                 if c.label not in before_cols]
    if dropped_r:
        notes.append("지난 차수에 없어서 증감을 못 낸 배너: " + ", ".join(dropped_r))
    if dropped_c:
        notes.append("지난 차수에 없어서 증감을 못 낸 보기: " + ", ".join(dropped_c))

    rows = [BannerRow(now.rows[i].group, now.rows[i].category,
                      now.rows[i].n) for i, _ in keep_rows]
    columns = [ValueColumn(now.columns[j].label, "stat",
                           now.columns[j].decimals) for j, _ in keep_cols]
    matrix = []
    for i, bi in keep_rows:
        line = []
        for j, bj in keep_cols:
            a, b = now.matrix[i][j], before.matrix[bi][bj]
            if a is None or b is None:
                line.append(None)
                continue
            d = float(a) - float(b)
            # 부동소수 오차로 '-0.0' 이 찍히는 것을 막는다
            line.append(0.0 if abs(d) < 1e-9 else d)
        matrix.append(line)

    return TableResult(
        title=title, n_label=f"{now.n_label} (이번)", rows=rows, columns=columns,
        matrix=matrix, orientation=now.orientation,
        notes=notes, row_kind=now.row_kind,
    )


def compare_waves(df_now, meta_now, df_before, meta_before,
                  blocks: list[TableBlock], *,
                  label_now: str = "이번 차수",
                  label_before: str = "지난 차수") -> tuple[list[TableResult], list[str]]:
    """표 정의 하나마다 이번 차수 · 지난 차수 · 증감 세 표를 만든다.

    증감을 한 표에 겹쳐 넣지 않는 이유: 같은 칸에 값이 두 개 들어가면 숫자
    서식을 하나만 줄 수 없어서 엑셀에서 계산에 못 씁니다. 표 세 개로 두면
    각각 그대로 쓸 수 있습니다.
    """
    results: list[TableResult] = []
    problems: list[str] = []

    for block in blocks:
        missing = missing_vars(block, list(df_before.columns))
        if missing:
            problems.append(
                f"'{block.title}' — 지난 차수 데이터에 없는 변수: "
                + ", ".join(missing) + ". 이 표는 이번 차수만 넣었습니다."
            )
            results.append(compute_table(df_now, meta_now, block))
            continue

        r_now = compute_table(df_now, meta_now, block)
        r_bef = compute_table(df_before, meta_before, block)
        r_now.title = f"{block.title} [{label_now}]"
        r_bef.title = f"{block.title} [{label_before}]"
        r_diff = diff_result(r_now, r_bef,
                             title=f"{block.title} [증감 %p]")
        results.extend([r_now, r_bef, r_diff])
        problems.extend(f"'{block.title}' — {m}" for m in r_diff.notes)

    return results, problems


def _letters_for_columns(segments, total_label: str) -> list[str]:
    """격자형 종합표의 열(배너 세그먼트)에 글자를 매긴다.
    비교는 같은 배너 그룹 안에서만 하므로 그룹마다 a 부터 다시 시작한다."""
    counts: dict[str, int] = {}
    for s in segments:
        if s.group == total_label and not s.label:
            continue
        counts[s.group] = counts.get(s.group, 0) + 1
    out, used = [], {}
    for s in segments:
        if (s.group == total_label and not s.label) or counts.get(s.group, 0) < 2:
            out.append("")
            continue
        i = used.get(s.group, 0)
        out.append(_SIG_LETTERS[i] if i < len(_SIG_LETTERS) else "")
        used[s.group] = i + 1
    return out


def _battery_row_order(vars_, dff, block: TableBlock, cats) -> list[int]:
    """종합표의 문항 순서. 평균이 있으면 평균 높은 순, 없으면 첫 요약 묶음 높은 순."""
    mean_spec = next((x for x in block.summaries if x.kind == "mean"), None)
    group_spec = next((x for x in block.summaries if x.kind == "group"), None)

    def key(i):
        s = dff[vars_[i]].dropna()
        if not len(s):
            return 0.0
        if mean_spec is not None:
            return -float(s.mean())
        if group_spec is not None:
            return -float(s.isin(group_spec.codes).sum()) / max(len(s), 1)
        return 0.0

    return sorted(range(len(vars_)), key=key)


# =============================================================================
# 5) 설정 저장 / 불러오기
# =============================================================================
# 표 정의(TableBlock)만 JSON 으로 주고받습니다. 계산 결과는 저장하지 않으므로,
# 같은 구조의 새 .sav 를 올려 그대로 다시 계산하면 바뀐 데이터가 반영됩니다.
#
# 파일로 주고받는 이유: 클라우드(Streamlit Community Cloud)는 앱이 다시 뜨면
# 서버에 저장한 파일이 사라집니다. JSON 을 내려받아 두면 로컬에서도, 클라우드
# 에서도, 다음 프로젝트에서도 똑같이 씁니다.

SETTINGS_VERSION = 1

# 파일명에 쓸 수 없는 글자
_BAD_FILENAME = re.compile(r'[\\/:*?"<>|\r\n\t]+')


def safe_stem(filename: str, fallback: str = "뱅크표") -> str:
    """업로드한 파일 이름에서 확장자와 경로를 떼고 파일명으로 쓸 수 있게 다듬는다.

    설정·엑셀 파일 이름을 원본 데이터 이름으로 짓기 위한 것이다.
    데이터마다 변수 구성이 다르므로, 이름이 같으면 서로 섞이기 쉽다.
    """
    stem = (filename or "").replace("\\", "/").split("/")[-1]
    if "." in stem:
        stem = stem.rsplit(".", 1)[0]
    stem = _BAD_FILENAME.sub("_", stem).strip(" ._")
    return stem[:80] or fallback


def block_to_dict(block: TableBlock) -> dict:
    return {
        "title": block.title,
        "target_raw": block.target_raw,
        "extra_cond": block.extra_cond,
        "total_label": block.total_label,
        "obser_var": block.obser_var,
        "banner_axis": list(block.banner_axis),
        "value_axis": list(block.value_axis),
        "row_ma_mode": block.row_ma_mode,
        "orientation": block.orientation,
        "ptotal_labels": dict(block.ptotal_labels),
        "summaries": [
            {"label": x.label, "kind": x.kind,
             "codes": list(x.codes), "decimals": x.decimals}
            for x in block.summaries
        ],
        "ftotal_labels": dict(block.ftotal_labels),
        "sig": (
            {"enabled": block.sig.enabled, "level": block.sig.level,
             "min_base": block.sig.min_base} if block.sig else None
        ),
        "min_base_show": block.min_base_show,
        "sort_values": block.sort_values,
        "battery_vars": list(block.battery_vars),
        "battery_metric": block.battery_metric,
        "merges": [
            {"name": m.name, "label": m.label, "varlist_raw": m.varlist_raw}
            for m in block.merges
        ],
        "stats": [
            {"name": s.name, "var": s.var, "fmt": s.fmt,
             "label": s.label, "restrict": s.restrict}
            for s in block.stats
        ],
    }


def block_from_dict(d: dict) -> TableBlock:
    return TableBlock(
        target_raw=d["target_raw"],
        extra_cond=d.get("extra_cond"),
        total_label=d.get("total_label", TOTAL_LABEL),
        merges=[MergeSpec(**m) for m in d.get("merges", [])],
        obser_var=d.get("obser_var"),
        banner_axis=list(d.get("banner_axis", [])),
        value_axis=list(d.get("value_axis", [])),
        stats=[StatSpec(**st) for st in d.get("stats", [])],
        title=d.get("title", "(제목 없음)"),
        ptotal_labels=d.get("ptotal_labels", {}),
        ftotal_labels=d.get("ftotal_labels", {}),
        row_ma_mode=d.get("row_ma_mode", "category"),
        orientation=d.get("orientation", BANNER_ROW),
        summaries=[SummarySpec(**x) for x in d.get("summaries", [])],
        sig=(SigSpec(**d["sig"]) if d.get("sig") else None),
        min_base_show=int(d.get("min_base_show", 0) or 0),
        sort_values=bool(d.get("sort_values", False)),
        battery_vars=list(d.get("battery_vars", [])),
        battery_metric=d.get("battery_metric"),
    )


def blocks_to_json(blocks: list[TableBlock], *, source_file: str = "",
                   note: str = "") -> bytes:
    """표 정의를 설정 파일(JSON)로. 어느 파일에서 만든 것인지도 같이 적는다."""
    import json
    from datetime import datetime, timezone

    used: list[str] = []
    for b in blocks:
        for v in referenced_vars(b, []):
            if v not in used:
                used.append(v)

    payload = {
        "kind": "banner_table_settings",
        "version": SETTINGS_VERSION,
        "source_file": source_file,          # 만들 때 쓴 .sav 또는 .sps 이름
        "saved_at": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "note": note,
        "table_count": len(blocks),
        "variables_used": used,              # 이 설정이 필요로 하는 변수 목록
        "tables": [block_to_dict(b) for b in blocks],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def load_settings(data: bytes) -> tuple[list[TableBlock], dict]:
    """설정 파일을 읽어 (표 정의들, 정보) 를 돌려준다. 형식이 아니면 ValueError."""
    import json

    try:
        payload = json.loads(data.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as e:
        raise ValueError(f"JSON 을 읽을 수 없습니다 — {e}") from e

    if not isinstance(payload, dict) or payload.get("kind") != "banner_table_settings":
        raise ValueError("이 앱에서 저장한 설정 파일이 아닙니다.")
    if payload.get("version", 1) > SETTINGS_VERSION:
        raise ValueError(
            f"더 새 버전({payload.get('version')})의 설정 파일입니다. 앱을 업데이트해 주세요."
        )
    tables = payload.get("tables") or []
    if not tables:
        raise ValueError("설정 파일에 표가 없습니다.")

    info = {
        "source_file": payload.get("source_file", ""),
        "saved_at": payload.get("saved_at", ""),
        "note": payload.get("note", ""),
        "variables_used": payload.get("variables_used", []),
    }
    return [block_from_dict(t) for t in tables], info


def blocks_from_json(data: bytes) -> list[TableBlock]:
    """load_settings 의 표 정의만 필요할 때."""
    return load_settings(data)[0]


def referenced_vars(block: TableBlock, columns: list[str]) -> list[str]:
    """이 표가 쓰는 변수 이름들. 새 .sav 에 다 있는지 확인할 때 쓴다."""
    names: list[str] = []

    def add(raw: str):
        raw = (raw or "").strip()
        if not raw:
            return
        if " to " in raw:
            # 구간 표기는 양 끝 이름만 확인한다 (사이 변수는 파일 순서에 달림)
            names.extend(p.strip() for p in raw.split(" to ", 1))
        else:
            names.extend(raw.split())

    add(block.target_raw)
    if block.obser_var:
        names.append(block.obser_var)
    if block.extra_cond and "=" in block.extra_cond:
        names.append(block.extra_cond.split("=")[0].strip())
    for m in block.merges:
        add(m.varlist_raw)
    merge_names = {m.name for m in block.merges}
    for token in block.banner_axis:
        if token != "@t3" and token not in merge_names:
            names.append(token)

    seen, out = set(), []
    for n in names:
        if n and n not in seen:
            seen.add(n)
            out.append(n)
    return out


def missing_vars(block: TableBlock, columns: list[str]) -> list[str]:
    have = set(columns)
    return [v for v in referenced_vars(block, list(columns)) if v not in have]


# =============================================================================
# 6) 화면 표시
# =============================================================================
def _fmt_num(value, decimals: int, *, blank: str = ".") -> str:
    if value is None or pd.isna(value):
        return blank
    return f"{value:,.{decimals}f}"


def _row_label(result: TableResult, i: int) -> str:
    """배너 행 이름. 유의성 검정을 켜면 글자를 뒤에 붙인다 — '남성 (a)'."""
    base = result.rows[i].category
    if result.letters and i < len(result.letters) and result.letters[i]:
        return f"{base} ({result.letters[i]})" if base else f"({result.letters[i]})"
    return base


def _col_label(result: TableResult, j: int, *, strip: bool = False) -> str:
    """값 열 이름. 격자형 종합표에서 검정을 켜면 글자를 뒤에 붙인다.

    strip 은 화면 표시용이다. 엑셀은 '      계' 처럼 앞 공백으로 오른쪽에
    붙여 놓은 원본 서식을 그대로 써야 한다.
    """
    raw = result.columns[j].label
    base = (raw.strip() or raw) if strip else raw
    if result.col_letters and j < len(result.col_letters) and result.col_letters[j]:
        return f"{base} ({result.col_letters[j]})"
    return base


def _is_hidden(result: TableResult, i: int) -> bool:
    return bool(result.hidden) and i < len(result.hidden) and result.hidden[i]


def _cell_text(result: TableResult, i: int, j: int) -> str:
    """화면용 칸 글자. 유의 표시가 있으면 숫자 뒤에 붙인다."""
    col = result.columns[j]
    blank = "-" if _is_hidden(result, i) else "."
    txt = _fmt_num(result.matrix[i][j], col.decimals, blank=blank)
    mark = result.mark_at(i, j)
    return f"{txt} {mark}" if mark else txt


def result_to_frame(result: TableResult) -> pd.DataFrame:
    """화면 표시용 DataFrame. 방향에 따라 축을 바꿔 만든다."""
    # 딕셔너리로 모으면 이름이 같은 칸(요약의 '평균' 과 통계의 '평균')이
    # 서로 덮어써서 한 칸이 조용히 사라진다. 그래서 목록으로 모은다.
    labels = [result.n_label] + [
        _col_label(result, j, strip=True) for j in range(len(result.columns))
    ]
    series = [[f"({r.n:,})" for r in result.rows]] + [
        [_cell_text(result, i, j) for i in range(len(result.rows))]
        for j in range(len(result.columns))
    ]
    banner_index = pd.MultiIndex.from_tuples(
        [(r.group, _row_label(result, i)) for i, r in enumerate(result.rows)],
        names=list(result.row_axis_names),
    )

    if result.banner_on_rows:
        # 배너 = 행 (SPSS 산출물과 같은 방향)
        frame = pd.DataFrame(list(zip(*series)), index=banner_index, columns=labels)
        return frame

    # 배너 = 열
    return pd.DataFrame(series, index=labels, columns=banner_index)


# =============================================================================
# 7) 엑셀 출력
# =============================================================================
def _styles():
    from openpyxl.styles import Alignment, Border, Font, Side

    return {
        "thick": Side(style="medium"),
        "thin": Side(style="hair"),
        "font": Font(name="맑은 고딕", size=9),
        "bold": Font(name="맑은 고딕", size=9, bold=True),
        "link": Font(name="맑은 고딕", size=9, color="0563C1", underline="single"),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "right": Alignment(horizontal="right", vertical="center"),
        "Border": Border,
    }


def _num(value):
    return None if value is None or pd.isna(value) else float(value)


def _excel_cell(result: TableResult, i: int, j: int):
    """엑셀에 넣을 값과 서식.

    보통은 숫자 + 숫자서식으로 넣어 엑셀에서 그대로 계산할 수 있게 한다.
    유의 표시가 붙는 칸은 '42.0 a' 처럼 글자가 섞이므로 문자로 넣는다
    (SPSS 산출물도 같은 방식이다). 소표본으로 감춘 칸은 '-' 문자.
    """
    col = result.columns[j]
    if _is_hidden(result, i):
        return "-", None
    mark = result.mark_at(i, j)
    if mark:
        return f"{_fmt_num(result.matrix[i][j], col.decimals)} {mark}", None
    return _num(result.matrix[i][j]), col.excel_format


def _write_banner_rows(tab, row: int, result: TableResult, S) -> int:
    """배너 = 행. A=그룹, B=보기, C=사례수, D부터 값."""
    Border = S["Border"]
    ncols = 3 + len(result.columns)

    cell = tab.cell(row=row, column=1, value=result.title)
    cell.font, cell.alignment = S["bold"], S["left"]
    tab.merge_cells(start_row=row, start_column=1, end_row=row,
                    end_column=max(ncols - 3, 2))
    row += 1

    head = row
    tab.merge_cells(start_row=head, start_column=1, end_row=head, end_column=2)
    for c in range(1, ncols + 1):
        cell = tab.cell(row=head, column=c)
        cell.font, cell.alignment = S["font"], S["center"]
        cell.border = Border(top=S["thick"], bottom=S["thick"],
                             left=S["thick"] if c == 1 else S["thin"],
                             right=S["thick"] if c == ncols else S["thin"])
    tab.cell(row=head, column=3, value=result.n_label)
    for j in range(len(result.columns)):
        tab.cell(row=head, column=4 + j, value=_col_label(result, j))
    tab.row_dimensions[head].height = 30
    row += 1

    first = row
    for i, brow in enumerate(result.rows):
        last = i == len(result.rows) - 1
        if (i == 0 and not brow.category) or not result.rows_have_groups:
            # 전체 행, 그리고 그룹이 없는 표(척도 종합표)는 A:B 를 합쳐 쓴다
            tab.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            tab.cell(row=row, column=1, value=brow.group or _row_label(result, i))
        else:
            tab.cell(row=row, column=1, value=brow.group)
            tab.cell(row=row, column=2, value=_row_label(result, i))
        tab.cell(row=row, column=3, value=f"({brow.n:,})")
        for j in range(len(result.columns)):
            value, fmt = _excel_cell(result, i, j)
            cell = tab.cell(row=row, column=4 + j, value=value)
            if fmt:
                cell.number_format = fmt
        for c in range(1, ncols + 1):
            cell = tab.cell(row=row, column=c)
            cell.font = S["font"]
            cell.alignment = S["left"] if c <= 2 else S["right"]
            cell.border = Border(
                top=S["thick"] if row == first else None,
                bottom=S["thick"] if last else None,
                left=S["thick"] if c in (1, 3) else S["thin"],
                right=S["thick"] if c in (2, ncols) else S["thin"],
            )
        tab.row_dimensions[row].height = 16
        row += 1

    # 배너 그룹명 세로 병합. 그룹이 없는 표(척도 종합표)는 이미 A:B 를 가로로
    # 합쳐 문항 이름을 넣었으므로 건드리면 안 된다 — 세로로 합치면 첫 줄만
    # 남고 나머지 문항 이름이 지워진다.
    if not result.rows_have_groups:
        return row

    start, prev = None, None
    for i, brow in enumerate(result.rows):
        if i == 0 and not brow.category:
            prev = None
            continue
        if brow.group != prev:
            if start is not None and i - 1 > start:
                tab.merge_cells(start_row=first + start, start_column=1,
                                end_row=first + i - 1, end_column=1)
            start, prev = i, brow.group
    if start is not None and len(result.rows) - 1 > start:
        tab.merge_cells(start_row=first + start, start_column=1,
                        end_row=first + len(result.rows) - 1, end_column=1)
    return row


def _write_banner_cols(tab, row: int, result: TableResult, S) -> int:
    """배너 = 열. A=보기(사례수/보기/계), B부터 배너가 2단 머리글로."""
    Border = S["Border"]
    ncols = 1 + len(result.rows)

    cell = tab.cell(row=row, column=1, value=result.title)
    cell.font, cell.alignment = S["bold"], S["left"]
    tab.merge_cells(start_row=row, start_column=1, end_row=row,
                    end_column=max(ncols, 2))
    row += 1

    h1, h2 = row, row + 1
    tab.merge_cells(start_row=h1, start_column=1, end_row=h2, end_column=1)
    for j, brow in enumerate(result.rows):
        c = 2 + j
        tab.cell(row=h1, column=c, value=brow.group)
        if brow.category:
            tab.cell(row=h2, column=c, value=_row_label(result, j))
        else:
            # 전체 열: 두 줄을 위아래로 병합
            tab.merge_cells(start_row=h1, start_column=c, end_row=h2, end_column=c)
    for r in (h1, h2):
        for c in range(1, ncols + 1):
            cell = tab.cell(row=r, column=c)
            cell.font, cell.alignment = S["font"], S["center"]
            cell.border = Border(
                top=S["thick"] if r == h1 else None,
                bottom=S["thick"] if r == h2 else None,
                left=S["thick"] if c <= 2 else S["thin"],
                right=S["thick"] if c == ncols else S["thin"],
            )
    tab.row_dimensions[h1].height = 16
    tab.row_dimensions[h2].height = 30

    # 같은 배너 그룹이 이어지는 구간은 위쪽 머리글을 좌우 병합
    start, prev = None, None
    for j, brow in enumerate(result.rows):
        if not brow.category:
            prev = None
            continue
        if brow.group != prev:
            if start is not None and j - 1 > start:
                tab.merge_cells(start_row=h1, start_column=2 + start,
                                end_row=h1, end_column=2 + j - 1)
            start, prev = j, brow.group
    if start is not None and len(result.rows) - 1 > start:
        tab.merge_cells(start_row=h1, start_column=2 + start,
                        end_row=h1, end_column=2 + len(result.rows) - 1)
    row = h2 + 1

    # 사례수 행 + 보기 행들. 값과 서식을 칸마다 따로 정한다
    # (유의 표시가 붙은 칸만 문자가 되므로 행 단위로 정할 수 없다)
    body: list[tuple[str, list]] = [
        (result.n_label, [(f"({r.n:,})", None) for r in result.rows])
    ]
    for j in range(len(result.columns)):
        body.append((
            _col_label(result, j),
            [_excel_cell(result, i, j) for i in range(len(result.rows))],
        ))

    first = row
    for k, (label, cells) in enumerate(body):
        last = k == len(body) - 1
        tab.cell(row=row, column=1, value=label)
        for j, (v, fmt) in enumerate(cells):
            cell = tab.cell(row=row, column=2 + j, value=v)
            if fmt:
                cell.number_format = fmt
        for c in range(1, ncols + 1):
            cell = tab.cell(row=row, column=c)
            cell.font = S["font"]
            cell.alignment = S["left"] if c == 1 else S["right"]
            cell.border = Border(
                top=S["thick"] if row == first else None,
                bottom=S["thick"] if last else None,
                left=S["thick"] if c <= 2 else S["thin"],
                right=S["thick"] if c == ncols else S["thin"],
            )
        tab.row_dimensions[row].height = 16
        row += 1
    return row


def _set_table_widths(tab, results: list[TableResult]) -> None:
    """열 너비 — 두 방향을 섞어도 읽히게 넉넉히 잡는다."""
    from openpyxl.utils import get_column_letter

    if any(r.banner_on_rows for r in results):
        tab.column_dimensions["A"].width = 16
        tab.column_dimensions["B"].width = 14
        tab.column_dimensions["C"].width = 9
        start_col, width = 4, 13
    else:
        tab.column_dimensions["A"].width = 16
        start_col, width = 2, 13
    widest = max(
        (len(r.columns) if r.banner_on_rows else len(r.rows) for r in results),
        default=0,
    )
    for j in range(widest):
        tab.column_dimensions[get_column_letter(start_col + j)].width = width


# 엑셀 시트 이름에 쓸 수 없는 글자와 길이 제한(31자)
_BAD_SHEETNAME = re.compile(r"[\\/*?:\[\]]+")


def _sheet_name(title: str, used: set[str], index: int) -> str:
    """표 제목으로 시트 이름을 만든다. 못 쓰는 글자를 바꾸고 겹치면 번호를 붙인다."""
    base = _BAD_SHEETNAME.sub("_", (title or "").strip()) or f"표{index}"
    base = base[:28] or f"표{index}"
    name, k = base, 2
    while name in used:
        suffix = f"_{k}"
        name = base[:31 - len(suffix)] + suffix
        k += 1
    used.add(name)
    return name


def write_tables_xlsx(results: list[TableResult], *,
                      split_sheets: bool = False) -> bytes:
    """SPSS 산출물과 같은 형태로 엑셀을 만든다.

    기본은 시트 두 개
      · '목 차' — 번호 + 제목, 각 표로 가는 하이퍼링크
      · 'Table' — 표들을 위아래로 이어 붙이고 사이에 빈 행 하나

    split_sheets=True 면 표마다 시트를 하나씩 만듭니다. 표를 따로 떼어
    쓰거나 시트별로 인쇄할 때 편합니다. 목차의 링크도 각 시트로 갑니다.

    표마다 orientation 을 따르므로, 배너를 행으로 둔 표와 열로 둔 표를
    한 파일에 섞어도 됩니다.
    """
    from openpyxl import Workbook

    S = _styles()
    wb = Workbook()
    toc = wb.active
    toc.title = "목 차"

    toc.column_dimensions["A"].width = 8
    toc.column_dimensions["B"].width = 80
    cell = toc.cell(row=1, column=2, value="목  차")
    cell.font, cell.alignment = S["bold"], S["center"]

    links: list[str] = []                # 목차에서 쓸 하이퍼링크 목적지

    if split_sheets:
        used: set[str] = {"목 차"}
        for i, result in enumerate(results, start=1):
            name = _sheet_name(result.title, used, i)
            sheet = wb.create_sheet(name)
            if result.banner_on_rows:
                _write_banner_rows(sheet, 1, result, S)
            else:
                _write_banner_cols(sheet, 1, result, S)
            _set_table_widths(sheet, [result])
            links.append(f"#'{name}'!A1")
    else:
        tab = wb.create_sheet("Table")
        row = 1
        for result in results:
            links.append(f"#'Table'!A{row}")
            if result.banner_on_rows:
                row = _write_banner_rows(tab, row, result, S)
            else:
                row = _write_banner_cols(tab, row, result, S)
            row += 1        # 표 사이 빈 행
        _set_table_widths(tab, results)

    for i, (result, target) in enumerate(zip(results, links), start=1):
        num = toc.cell(row=1 + i, column=1, value=i)
        num.font, num.alignment = S["font"], S["center"]
        link = toc.cell(row=1 + i, column=2, value=result.title)
        link.font, link.alignment = S["link"], S["left"]
        link.hyperlink = target

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# 8) 빈도표 — 변수 여러 개를 한 번에
# =============================================================================
# 뱅크표와 달리 배너가 없습니다. 변수 하나에 표 하나이고, 열은 SPSS 빈도표와
# 같은 모양입니다 — 빈도 · 퍼센트 · 유효퍼센트 · 누적퍼센트.
#
# 값 라벨 처리에 두 가지 규칙을 뒀습니다. 둘 다 데이터를 점검하는 데 씁니다.
#   · 값 라벨에 정의된 보기는 응답이 0이어도 표시한다 — 아무도 안 고른 보기를
#     알아야 쿼터나 로직을 확인할 수 있습니다.
#   · 값 라벨에 없는 코드가 데이터에 있으면 표시하고 알린다 — 코딩 오류이거나
#     라벨을 안 붙인 것이므로, 조용히 넘어가면 안 됩니다.

FREQ_TEXT_LIMIT = 30        # 문자/무라벨 변수의 고유값이 이보다 많으면 줄여서 본다


@dataclass
class FreqRow:
    """빈도표 한 줄."""
    label: str
    count: int
    pct: float                      # 전체(결측 포함) 대비
    valid_pct: float | None         # 유효 응답 대비 (결측 줄은 None)
    cum_pct: float | None           # 유효퍼센트의 누적
    kind: str = "value"             # 'value' | 'undefined' | 'missing' | 'total'


@dataclass
class FreqTable:
    """변수 하나의 빈도표."""
    var: str
    label: str                      # 변수 라벨 (문항 문구). 없으면 변수명
    rows: list[FreqRow]
    total_n: int
    valid_n: int
    missing_n: int
    stats: dict | None = None       # 값 라벨 없는 숫자 변수의 요약
    notes: list[str] = field(default_factory=list)

    @property
    def title(self) -> str:
        return f"{self.var} — {self.label}" if self.label != self.var else self.var


def _freq_stats(series: pd.Series, decimals: int = 2) -> dict:
    """값 라벨 없는 숫자 변수의 요약."""
    s = pd.to_numeric(series, errors="coerce").dropna()
    if not len(s):
        return {}
    return {
        "평균": round(float(s.mean()), decimals),
        "표준편차": round(float(s.std()), decimals) if len(s) > 1 else None,
        "중위값": round(float(s.median()), decimals),
        "최소값": round(float(s.min()), decimals),
        "최대값": round(float(s.max()), decimals),
        "고유값": int(s.nunique()),
    }


def compute_frequencies(
    df: pd.DataFrame,
    meta,
    variables: list[str],
    *,
    show_missing: bool = True,
    sort_by_count: bool = False,
    decimals: int = 1,
    stat_decimals: int = 2,
    text_limit: int = FREQ_TEXT_LIMIT,
) -> list[FreqTable]:
    """고른 변수들의 빈도표를 한 번에 만든다.

    sort_by_count 를 켜면 응답 많은 보기부터 나오되, '기타'·'모름' 계열은
    뱅크표와 같은 규칙으로 맨 뒤에 둡니다.
    """
    value_labels = meta.variable_value_labels
    col_labels = meta.column_names_to_labels
    out: list[FreqTable] = []

    for var in variables:
        if var not in df.columns:
            continue

        series = df[var]
        total_n = int(len(series))
        missing_n = int(series.isna().sum())
        valid_n = total_n - missing_n
        notes: list[str] = []
        stats = None

        vl = value_labels.get(var, {})
        counts = series.value_counts(dropna=True)

        if vl:
            # 값 라벨이 있는 변수 — 정의된 보기를 전부 쓰고, 정의에 없는 값은 뒤에
            pairs = [(vl[code], int(counts.get(code, 0)), "value")
                     for code in sorted(vl.keys())]
            unknown = [c for c in counts.index if c not in vl]
            for code in sorted(unknown):
                shown = int(code) if float(code).is_integer() else code
                pairs.append((f"[라벨 없음] {shown}", int(counts[code]), "undefined"))
            if unknown:
                notes.append(
                    f"값 라벨에 없는 코드가 {len(unknown)}개 있습니다 — "
                    "코딩 오류이거나 라벨을 안 붙인 것입니다."
                )
        else:
            numeric = pd.api.types.is_numeric_dtype(series)
            if numeric:
                stats = _freq_stats(series, stat_decimals)
            if int(series.nunique()) > text_limit:
                # 값이 너무 많으면 표로 만들 수 없다. 숫자면 통계로 갈음하고,
                # 문자면 많이 나온 것만 보여준다 (주관식은 따로 봐야 한다).
                if numeric:
                    notes.append(
                        f"고유값이 {int(series.nunique())}개라 빈도표 대신 "
                        "통계 요약만 넣었습니다."
                    )
                    out.append(FreqTable(var, col_labels.get(var) or var, [],
                                         total_n, valid_n, missing_n, stats, notes))
                    continue
                top = counts.head(text_limit)
                pairs = [(str(k), int(v), "value") for k, v in top.items()]
                # 자른 나머지를 한 줄로 남긴다. 안 그러면 보이는 줄의 합이
                # '합계' 와 안 맞아서 읽는 사람이 계산을 의심하게 된다.
                rest_kinds = int(series.nunique()) - len(top)
                rest_count = int(counts.sum() - top.sum())
                if rest_count:
                    pairs.append((f"(나머지 {rest_kinds}종)", rest_count, "value"))
                notes.append(
                    f"고유값이 {int(series.nunique())}개라 많이 나온 "
                    f"{text_limit}개만 따로 넣고 나머지는 한 줄로 묶었습니다."
                )
            else:
                keys = sorted(counts.index, key=lambda x: (isinstance(x, str), x))
                pairs = []
                for k in keys:
                    shown = (int(k) if isinstance(k, float) and float(k).is_integer()
                             else k)
                    pairs.append((str(shown), int(counts[k]), "value"))

        if sort_by_count:
            body = [p for p in pairs if not _is_tail_label(p[0])]
            tail = [p for p in pairs if _is_tail_label(p[0])]
            body.sort(key=lambda p: -p[1])
            pairs = body + tail

        rows: list[FreqRow] = []
        cum = 0.0
        for label, cnt, kind in pairs:
            pct = cnt / total_n * 100 if total_n else 0.0
            vpct = cnt / valid_n * 100 if valid_n else None
            if vpct is not None:
                cum += vpct
            rows.append(FreqRow(label, cnt, round(pct, decimals),
                                None if vpct is None else round(vpct, decimals),
                                None if vpct is None else round(min(cum, 100.0),
                                                                decimals),
                                kind))

        rows.append(FreqRow("합계", valid_n,
                            round(valid_n / total_n * 100, decimals) if total_n else 0.0,
                            100.0 if valid_n else None,
                            None, "total"))
        if show_missing and missing_n:
            rows.append(FreqRow("무응답(결측)", missing_n,
                                round(missing_n / total_n * 100, decimals),
                                None, None, "missing"))

        out.append(FreqTable(var, col_labels.get(var) or var, rows,
                             total_n, valid_n, missing_n, stats, notes))

    return out


def freq_to_frame(table: FreqTable) -> pd.DataFrame:
    """화면 표시용."""
    if not table.rows:
        if table.stats:
            return pd.DataFrame(
                [[v for v in table.stats.values()]],
                columns=list(table.stats.keys()), index=["값"],
            )
        return pd.DataFrame()

    def txt(v, dec=1):
        return "" if v is None else f"{v:,.{dec}f}"

    data = [[f"{r.count:,}", txt(r.pct), txt(r.valid_pct), txt(r.cum_pct)]
            for r in table.rows]
    return pd.DataFrame(
        data,
        index=pd.Index([r.label for r in table.rows], name="보기"),
        columns=["빈도", "퍼센트", "유효퍼센트", "누적퍼센트"],
    )


def _write_freq_table(sheet, row: int, table: FreqTable, S) -> int:
    """빈도표 하나를 시트에 쓴다. 다음에 쓸 행 번호를 돌려준다."""
    Border = S["Border"]
    ncols = 5                       # 보기 + 빈도 + 퍼센트 + 유효퍼센트 + 누적퍼센트

    cell = sheet.cell(row=row, column=1, value=table.title)
    cell.font, cell.alignment = S["bold"], S["left"]
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    row += 1

    if not table.rows:
        if table.stats:
            for j, (k, v) in enumerate(table.stats.items(), start=1):
                head = sheet.cell(row=row, column=j, value=k)
                head.font, head.alignment = S["font"], S["center"]
                body = sheet.cell(row=row + 1, column=j, value=v)
                body.font, body.alignment = S["font"], S["right"]
            row += 2
        for note in table.notes:
            note_cell = sheet.cell(row=row, column=1, value=f"· {note}")
            note_cell.font, note_cell.alignment = S["font"], S["left"]
            row += 1
        return row

    head = row
    for j, name in enumerate(["보기", "빈도", "퍼센트", "유효퍼센트", "누적퍼센트"],
                             start=1):
        cell = sheet.cell(row=head, column=j, value=name)
        cell.font, cell.alignment = S["font"], S["center"]
        cell.border = Border(top=S["thick"], bottom=S["thick"],
                             left=S["thick"] if j == 1 else S["thin"],
                             right=S["thick"] if j == ncols else S["thin"])
    sheet.row_dimensions[head].height = 20
    row += 1

    first = row
    for i, r in enumerate(table.rows):
        last = i == len(table.rows) - 1
        sheet.cell(row=row, column=1, value=r.label)
        c = sheet.cell(row=row, column=2, value=r.count)
        c.number_format = "#,##0"
        for j, v in enumerate([r.pct, r.valid_pct, r.cum_pct], start=3):
            cell = sheet.cell(row=row, column=j, value=v)
            cell.number_format = "###0.0"
        for j in range(1, ncols + 1):
            cell = sheet.cell(row=row, column=j)
            cell.font = S["bold"] if r.kind == "total" else S["font"]
            cell.alignment = S["left"] if j == 1 else S["right"]
            cell.border = Border(
                top=S["thick"] if row == first else None,
                bottom=S["thick"] if last else None,
                left=S["thick"] if j == 1 else S["thin"],
                right=S["thick"] if j == ncols else S["thin"],
            )
        sheet.row_dimensions[row].height = 16
        row += 1

    if table.stats:
        txt = " · ".join(f"{k} {v:,}" for k, v in table.stats.items()
                         if v is not None and k != "고유값")
        cell = sheet.cell(row=row, column=1, value=txt)
        cell.font, cell.alignment = S["font"], S["left"]
        row += 1
    for note in table.notes:
        cell = sheet.cell(row=row, column=1, value=f"· {note}")
        cell.font, cell.alignment = S["font"], S["left"]
        row += 1
    return row


def write_freq_xlsx(tables: list[FreqTable], *, split_sheets: bool = False) -> bytes:
    """빈도표들을 엑셀로. 뱅크표 출력과 같은 서식·구조를 씁니다."""
    from openpyxl import Workbook

    S = _styles()
    wb = Workbook()
    toc = wb.active
    toc.title = "목 차"
    toc.column_dimensions["A"].width = 8
    toc.column_dimensions["B"].width = 60
    toc.column_dimensions["C"].width = 10
    cell = toc.cell(row=1, column=2, value="목  차")
    cell.font, cell.alignment = S["bold"], S["center"]
    n_head = toc.cell(row=1, column=3, value="유효 N")
    n_head.font, n_head.alignment = S["bold"], S["center"]

    def widths(sheet):
        for col, w in zip("ABCDE", (34, 10, 10, 12, 12)):
            sheet.column_dimensions[col].width = w

    links: list[str] = []
    if split_sheets:
        used: set[str] = {"목 차"}
        for i, table in enumerate(tables, start=1):
            name = _sheet_name(table.var, used, i)
            sheet = wb.create_sheet(name)
            _write_freq_table(sheet, 1, table, S)
            widths(sheet)
            links.append(f"#'{name}'!A1")
    else:
        sheet = wb.create_sheet("빈도표")
        row = 1
        for table in tables:
            links.append(f"#'빈도표'!A{row}")
            row = _write_freq_table(sheet, row, table, S) + 1
        widths(sheet)

    for i, (table, target) in enumerate(zip(tables, links), start=1):
        num = toc.cell(row=1 + i, column=1, value=i)
        num.font, num.alignment = S["font"], S["center"]
        link = toc.cell(row=1 + i, column=2, value=table.title)
        link.font, link.alignment = S["link"], S["left"]
        link.hyperlink = target
        n = toc.cell(row=1 + i, column=3, value=table.valid_n)
        n.font, n.alignment = S["font"], S["right"]
        n.number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
