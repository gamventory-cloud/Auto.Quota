# banner_table_form.py
#
# 엑셀 폼(양식)으로 표를 정의하는 부분입니다. 신텍스(.sps) 없이 엑셀에
# 표 목록과 배너를 적어 올리면 그대로 표가 나옵니다.
# 페이지 파일이 아니므로 Home.py 의 SKIP 목록에 넣어 주세요.
#
# 양식은 시트 네 개입니다.
#   · 표목록   — 표 하나가 한 줄. 제목·문항유형·변수·표시(%/N)·방향 등
#   · 배너     — 배너로 쓸 변수들. 한 번 적으면 모든 표가 같이 씁니다
#   · 변수목록 — 올린 .sav 의 변수와 값 라벨 (보고 베끼는 용도, 읽기 전용)
#   · 사용법   — 각 칸에 무엇을 적는지
#
# 만드는 쪽(write_form_template)과 읽는 쪽(read_form)이 같은 열 정의
# (_TABLE_COLS / _BANNER_COLS)를 쓰므로, 열을 바꿀 때는 그 한 곳만 고치면 됩니다.

from __future__ import annotations

import io

import pandas as pd

from banner_table_engine import (
    BANNER_COL,
    BANNER_ROW,
    BannerSpec,
    SigSpec,
    TableBlock,
    build_battery_block,
    build_block,
    expand_var_range,
    parse_summary_spec,
    title_with_marker,
    value_breakdown_size,
)

# 응답값 분포의 보기가 이보다 많으면 표가 너무 넓어지므로 미리 알려준다
BREAKDOWN_WARN = 30

SHEET_TABLES = "표목록"
SHEET_BANNER = "배너"
SHEET_VARS = "변수목록"
SHEET_HELP = "사용법"

DEFAULT_SET = "기본"

# 표목록 시트의 열 (이름, 너비, 안내)
_TABLE_COLS = [
    ("표제목", 26, "비우면 변수명으로 자동"),
    ("문항유형", 14, "단수 / 복수 / 수치형 / 척도종합 / 평균서머리"),
    ("문항변수", 30, "변수명. 복수는 콤마로 여러 개 또는 'A to B'"),
    ("표시", 10, "% / N / %+N. 수치형은 비우면 통계만"),
    ("요약", 22, "척도 요약. 예: 상2,중,하2,평균"),
    ("통계", 24, "수치형은 평균,중위값,최소값,최대값 / 평균서머리는 지표 하나"),
    ("소수점", 8, "통계 소수점. 기본 2 (% 는 1자리 고정)"),
    ("계표시", 8, "Y / N"),
    ("정렬", 8, "Y 면 응답 많은 보기부터 (기타·모름은 맨 아래)"),
    ("유의성", 10, "95 또는 99 면 배너끼리 검정해 a/b/c 표시. 비우면 안 함"),
    ("소표본", 10, "이 사례수 미만인 배너는 값을 '-' 로. 비우면 안 감춤"),
    ("배너방향", 12, "행(세로) / 열(가로)"),
    ("필터", 18, "예: 주체=공설 또는 주체=1"),
    ("배너세트", 12, f"비우면 '{DEFAULT_SET}'"),
]

# 배너 시트의 열
_BANNER_COLS = [
    ("배너세트", 12, f"비우면 '{DEFAULT_SET}'"),
    ("그룹명", 20, "비우면 변수 라벨. 다중응답으로 묶을 때는 필수"),
    ("변수", 34, "변수명 하나. 여러 개를 하나로 묶을 때는 콤마 또는 'A to B'"),
    ("다중응답", 10, "여러 변수를 하나의 배너로 묶을 때 Y"),
]

# 입력값 정규화 표
_TYPE_MAP = {
    "단수": "single", "단일": "single", "단일응답": "single", "single": "single",
    "복수": "multi", "다중": "multi", "다중응답": "multi", "ma": "multi", "multi": "multi",
    # 수치형이 기본 이름. 연속형·평균 등은 예전 양식 호환용
    "수치형": "obser", "수치": "obser", "연속형": "obser", "연속": "obser",
    "평균": "obser", "obser": "obser", "mean": "obser", "numeric": "obser",
    # 척도 종합표 — 행이 문항, 열이 보기(척도종합) 또는 배너(평균서머리)
    "척도종합": "battery", "종합": "battery", "종합표": "battery",
    "서머리": "battery", "battery": "battery",
    "평균서머리": "battery_grid", "격자": "battery_grid",
    "척도종합격자": "battery_grid", "grid": "battery_grid",
}
# '평균서머리' 의 '통계' 칸 → 격자에 넣을 지표
_METRIC_MAP = {
    "평균": "mean", "mean": "mean",
    "표준편차": "std", "sd": "std", "std": "std",
    "top2": "Top2", "상2": "Top2", "top3": "Top3", "상3": "Top3",
    "bottom2": "Bottom2", "하2": "Bottom2", "bottom3": "Bottom3", "하3": "Bottom3",
    "middle": "Middle", "중": "Middle", "중간": "Middle",
}
_SHOW_MAP = {
    "%": ("pct",), "퍼센트": ("pct",), "pct": ("pct",),
    "n": ("n",), "빈도": ("n",), "사례수": ("n",),
    "%+n": ("pct", "n"), "n+%": ("pct", "n"), "둘다": ("pct", "n"),
    "%,n": ("pct", "n"), "both": ("pct", "n"),
}
_ORIENT_MAP = {
    "행": BANNER_ROW, "세로": BANNER_ROW, "row": BANNER_ROW, "왼쪽": BANNER_ROW,
    "열": BANNER_COL, "가로": BANNER_COL, "col": BANNER_COL, "column": BANNER_COL,
    "위": BANNER_COL,
}
_STAT_MAP = {
    "평균": "MEAN", "mean": "MEAN",
    "중위값": "MEDIAN", "중앙값": "MEDIAN", "median": "MEDIAN",
    "최소값": "MIN", "최소": "MIN", "min": "MIN",
    "최대값": "MAX", "최대": "MAX", "max": "MAX",
}
_YES = {"y", "yes", "예", "o", "ㅇ", "true", "1"}
_NO = {"n", "no", "아니오", "아님", "x", "false", "0"}


def _txt(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _split_vars(raw: str, columns: list[str]) -> list[str]:
    """'A, B' · 'A to B' · 'A B' 를 모두 변수 목록으로."""
    raw = _txt(raw)
    if not raw:
        return []
    if " to " in raw.lower():
        low = raw.lower()
        i = low.index(" to ")
        return expand_var_range(f"{raw[:i].strip()} to {raw[i + 4:].strip()}", columns)
    parts = [p.strip() for p in raw.replace(";", ",").split(",")]
    if len(parts) == 1:
        parts = [p for p in parts[0].split() if p]
    return [p for p in parts if p]


# =============================================================================
# 양식 만들기
# =============================================================================
def write_form_template(df: pd.DataFrame, meta, *, example: bool = True) -> bytes:
    """올린 .sav 에 맞춘 빈 양식을 만든다. 변수 목록 시트가 함께 들어간다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    HEAD_FILL = PatternFill("solid", fgColor="E8EEF4")
    HINT_FONT = Font(name="맑은 고딕", size=8, italic=True, color="777777")
    HEAD_FONT = Font(name="맑은 고딕", size=9, bold=True)
    BODY_FONT = Font(name="맑은 고딕", size=9)
    WRAP = Alignment(vertical="center", wrap_text=True)

    wb = Workbook()

    def make_sheet(title: str, cols) -> "object":
        ws = wb.create_sheet(title)
        for j, (name, width, hint) in enumerate(cols, start=1):
            c = ws.cell(row=1, column=j, value=name)
            c.font, c.fill, c.alignment = HEAD_FONT, HEAD_FILL, WRAP
            h = ws.cell(row=2, column=j, value=hint)
            h.font, h.alignment = HINT_FONT, WRAP
            ws.column_dimensions[get_column_letter(j)].width = width
        ws.freeze_panes = "A3"
        return ws

    del wb["Sheet"]

    # ── 표목록 ──
    ws = make_sheet(SHEET_TABLES, _TABLE_COLS)
    validations = {
        "문항유형": '"단수,복수,수치형,척도종합,평균서머리"',
        "표시": '"%,N,%+N"',
        "배너방향": '"행,열"',
        "계표시": '"Y,N"',
        "정렬": '"Y,N"',
        "유의성": '"95,99"',
    }
    names = [c[0] for c in _TABLE_COLS]
    for col_name, formula in validations.items():
        j = names.index(col_name) + 1
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{get_column_letter(j)}3:{get_column_letter(j)}300")

    if example:
        # 올린 데이터에서 그럴듯한 예시 한 줄을 만들어 둔다
        cat_vars = [c for c in df.columns
                    if meta.variable_value_labels.get(c)
                    and 2 <= len(meta.variable_value_labels[c]) <= 30]
        num_vars = [c for c in df.columns
                    if not meta.variable_value_labels.get(c)
                    and pd.api.types.is_numeric_dtype(df[c])]
        rows = []
        if cat_vars:
            rows.append([cat_vars[0], "단수", cat_vars[0], "%+N", "", "", 1,
                         "Y", "N", "", "", "행", "", ""])
        if num_vars:
            rows.append([f"{num_vars[0]} 평균", "수치형", num_vars[0], "", "",
                         "평균,중위값,최소값,최대값", 2,
                         "N", "N", "", "", "행", "", ""])
        for i, row in enumerate(rows):
            for j, v in enumerate(row, start=1):
                cell = ws.cell(row=3 + i, column=j, value=v)
                cell.font, cell.alignment = BODY_FONT, WRAP

    # ── 배너 ──
    wsb = make_sheet(SHEET_BANNER, _BANNER_COLS)
    j = [c[0] for c in _BANNER_COLS].index("다중응답") + 1
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    wsb.add_data_validation(dv)
    dv.add(f"{get_column_letter(j)}3:{get_column_letter(j)}300")

    if example:
        cat_vars = [c for c in df.columns
                    if meta.variable_value_labels.get(c)
                    and 2 <= len(meta.variable_value_labels[c]) <= 30]
        for i, v in enumerate(cat_vars[:3]):
            for jj, val in enumerate([DEFAULT_SET, "", v, ""], start=1):
                cell = wsb.cell(row=3 + i, column=jj, value=val)
                cell.font, cell.alignment = BODY_FONT, WRAP

    # ── 변수목록 (읽기 전용 참고) ──
    wsv = wb.create_sheet(SHEET_VARS)
    for j, (name, width) in enumerate(
        [("변수명", 24), ("변수 라벨", 40), ("값 라벨", 80)], start=1
    ):
        c = wsv.cell(row=1, column=j, value=name)
        c.font, c.fill, c.alignment = HEAD_FONT, HEAD_FILL, WRAP
        wsv.column_dimensions[get_column_letter(j)].width = width
    wsv.freeze_panes = "A2"
    for i, col in enumerate(df.columns, start=2):
        vl = meta.variable_value_labels.get(col, {})
        vl_txt = ", ".join(f"{int(k) if float(k).is_integer() else k}={v}"
                           for k, v in sorted(vl.items())) if vl else ""
        for j, val in enumerate(
            [col, meta.column_names_to_labels.get(col) or "", vl_txt], start=1
        ):
            cell = wsv.cell(row=i, column=j, value=val)
            cell.font, cell.alignment = BODY_FONT, WRAP

    # ── 사용법 ──
    wsh = wb.create_sheet(SHEET_HELP)
    wsh.column_dimensions["A"].width = 18
    wsh.column_dimensions["B"].width = 96
    guide = [
        ("", "엑셀 폼으로 뱅크표 만들기"),
        ("", ""),
        ("순서", "① '배너' 시트에 배너로 쓸 변수를 적는다 (한 번만)"),
        ("", "② '표목록' 시트에 표를 한 줄씩 적는다"),
        ("", "③ 저장하고 앱의 '엑셀 폼으로 만들기' 탭에 올린다"),
        ("", ""),
        ("변수명", f"'{SHEET_VARS}' 시트에 올린 데이터의 변수명과 값 라벨이 있으니 보고 적으세요."),
        ("", ""),
        ("[표목록]", ""),
    ]
    for name, _w, hint in _TABLE_COLS:
        guide.append((name, hint))
    guide += [
        ("", ""),
        ("[배너]", ""),
    ]
    for name, _w, hint in _BANNER_COLS:
        guide.append((name, hint))
    guide += [
        ("", ""),
        ("배너방향", "행 = 배너가 왼쪽에 세로로 (SPSS 산출물과 같은 모양)"),
        ("", "열 = 배너가 위에 가로로"),
        ("표시", "%+N 으로 적으면 % 표와 N 표가 각각 한 개씩, 두 개 나옵니다."),
        ("필터", "값은 코드(1)로도 라벨(공설)로도 적을 수 있습니다."),
        ("문항유형", "단수 = 보기 하나만 고르는 문항"),
        ("", "복수 = 여러 개 고르는 문항(중복응답). 변수를 콤마로 나열하세요."),
        ("", "수치형 = 값 자체가 숫자인 문항(이용료·나이 등)"),
        ("", "척도종합 = 문항 여러 개를 한 표에. 행이 문항, 열이 보기+계+요약"),
        ("", "평균서머리 = 행이 문항, 열이 배너. '통계' 칸에 지표 하나를 적으세요"),
        ("", ""),
        ("척도종합", "척도가 같은 문항들을 '문항변수' 에 콤마로 나열하세요 (예: Q5_1, Q5_2, Q5_3)."),
        ("", "정렬=Y 면 평균 높은 문항부터 나옵니다."),
        ("", "행끼리(문항끼리) 비교하는 표라서 유의성 검정은 하지 않습니다 —"),
        ("", "같은 응답자가 모든 문항에 답했으므로 독립표본 검정을 쓸 수 없습니다."),
        ("평균서머리", "'통계' 칸에 평균 / 표준편차 / Top2 / Bottom2 / 중간 중 하나."),
        ("", "Top2 같은 묶음을 쓰려면 '요약' 칸에 상2 처럼 정의도 적어 주세요."),
        ("", "이 표는 열(배너)끼리 비교하므로 유의성 검정을 할 수 있습니다."),
        ("", ""),
        ("유의성", "95 나 99 를 적으면 같은 배너 그룹 안에서 세그먼트끼리 비교합니다."),
        ("", "비율은 두 비율 z검정, 평균은 Welch t검정입니다."),
        ("", "배너 이름에 (a)(b)(c) 가 붙고, 유의하게 높은 칸에 상대 글자가 적힙니다."),
        ("", "예: '남성 (a)' 행의 42.6 b → 여성(b)보다 유의하게 높다는 뜻."),
        ("", "글자는 배너 그룹마다 a 부터 다시 시작합니다 (비교가 그룹 안에서만 되므로)."),
        ("", "'전체' 는 다른 배너를 포함하므로 비교에서 빼고 글자도 안 줍니다."),
        ("", "사례수 30 미만 배너는 검정에서 뺍니다."),
        ("", "※ 켜면 그 표의 값이 '42.6 b' 같은 문자로 나가서 엑셀 계산에는 못 씁니다."),
        ("", "※ 다중응답 배너는 한 응답자가 여러 세그먼트에 들어가 독립이 아니니 참고로만."),
        ("소표본", "예: 30 을 적으면 사례수 30 미만인 배너는 값이 '-' 로 나옵니다."),
        ("", "N=3 에서 33.3% 같은 숫자가 그대로 나가는 것을 막기 위한 것입니다."),
        ("정렬", "Y 면 응답이 많은 보기부터 왼쪽에 옵니다 (복수응답에서 특히 유용)."),
        ("", "'기타', '모름', '무응답' 계열은 응답이 많아도 맨 뒤로 보냅니다."),
        ("", ""),
        ("요약", "리커트 척도의 Top2·Middle·Bottom2·평균을 '계' 뒤에 붙입니다."),
        ("", "예: 상2,중,하2,평균  /  Top2,Bottom2  /  긍정=4,5; 부정=1,2; 평균"),
        ("", "상2 = 코드가 큰 보기 2개, 하2 = 코드가 작은 보기 2개,"),
        ("", "중 = 상·하에 안 들어간 나머지. 코드를 직접 적으려면 '이름=1,2' 로."),
        ("", "평균은 보기 코드값의 평균입니다 (1~5 척도의 평균점)."),
        ("", ""),
        ("수치형의 표시", "비우면 평균·중위값 같은 통계만 나옵니다."),
        ("", "% 나 N 을 적으면 응답된 값의 분포가 단수 표처럼 먼저 나오고,"),
        ("", "'계' 뒤에 통계 칸이 붙습니다. (사례수 → 값별 %/N → 계 → 평균…)"),
        ("", "값 종류가 많은 변수(이용료 등)는 보기가 수십~수백 개가 되니 주의하세요."),
        ("배너세트", "배너를 여러 벌 쓰고 싶을 때만 이름을 다르게 적으세요."),
        ("", "'표목록' 의 배너세트가 비어 있으면 첫 번째 세트를 씁니다."),
    ]
    for i, (a, b) in enumerate(guide, start=1):
        ca = wsh.cell(row=i, column=1, value=a)
        cb = wsh.cell(row=i, column=2, value=b)
        ca.font = HEAD_FONT if a.startswith("[") or i == 1 else BODY_FONT
        cb.font = HEAD_FONT if i == 1 else BODY_FONT
        ca.alignment = cb.alignment = WRAP

    wb.move_sheet(SHEET_HELP, offset=-3)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# 양식 읽기
# =============================================================================
def _looks_like_hint_row(values: list[str], var_cells: list[str],
                         columns: set[str]) -> bool:
    """안내 줄(머리글 아래 설명 줄)인지 판단한다.

    안내 문구와 글자를 맞춰보는 방식은 쓰지 않는다. 문구를 한 번 고치면
    이미 내려받아 둔 양식이 안 읽히기 때문이다. 대신 안내 줄의 성질로 본다.
      · 실제 변수 이름을 하나도 가리키지 않는다
      · 단수/복수/Y/N 같은 정해진 값도 하나도 없다
      · 칸 내용이 설명문처럼 길다
    변수명을 잘못 적은 줄은 정해진 값(문항유형 등)이 남아 있어서 데이터 줄로
    남고, 그래서 '없는 변수' 라고 제대로 알려줄 수 있다.
    """
    filled = [v for v in values if v]
    if not filled:
        return False

    # 실제 변수를 가리키면 데이터 줄
    for cell in var_cells:
        for token in cell.replace(";", ",").replace(",", " ").split():
            if token in columns:
                return False

    # 정해진 값이 하나라도 있으면 데이터 줄
    known = set(_TYPE_MAP) | set(_SHOW_MAP) | set(_ORIENT_MAP) | _YES | _NO
    for v in filled:
        if v.lower().replace(" ", "") in known:
            return False

    # 설명문처럼 긴 칸이 절반 이상이면 안내 줄
    long_cells = sum(1 for v in filled if len(v) > 12)
    return long_cells * 2 >= len(filled)


def _read_sheet(data: bytes, sheet: str, cols, columns: set[str],
                var_col_names: list[str]) -> pd.DataFrame:
    """머리글 아래의 안내 줄과 빈 줄을 걸러내고 읽는다."""
    frame = pd.read_excel(io.BytesIO(data), sheet_name=sheet, header=0, dtype=object)
    names = [c[0] for c in cols]
    frame = frame.rename(columns={c: str(c).strip() for c in frame.columns})
    for n in names:
        if n not in frame.columns:
            frame[n] = None
    frame = frame[names]

    keep = []
    for _, r in frame.iterrows():
        values = [_txt(v) for v in r.tolist()]
        if not any(values):
            continue
        var_cells = [_txt(r[n]) for n in var_col_names]
        if _looks_like_hint_row(values, var_cells, columns):
            continue
        keep.append(r)
    return pd.DataFrame(keep, columns=names) if keep else pd.DataFrame(columns=names)


def _banner_sets(data: bytes, df: pd.DataFrame, meta,
                 problems: list[str]) -> dict[str, list[BannerSpec]]:
    columns = list(df.columns)
    frame = _read_sheet(data, SHEET_BANNER, _BANNER_COLS, set(columns), ["변수"])
    sets: dict[str, list[BannerSpec]] = {}

    for i, r in frame.iterrows():
        line = int(i) + 2                     # 엑셀에서 보이는 줄 번호
        set_name = _txt(r["배너세트"]) or DEFAULT_SET
        group = _txt(r["그룹명"])
        merge = _txt(r["다중응답"]).lower() in _YES
        varlist = _split_vars(r["변수"], columns)

        if not varlist:
            problems.append(f"[{SHEET_BANNER}] {line}행: '변수' 가 비어 있습니다.")
            continue
        gone = [v for v in varlist if v not in df.columns]
        if gone:
            problems.append(
                f"[{SHEET_BANNER}] {line}행: 이 .sav 에 없는 변수 — {', '.join(gone)}"
            )
            continue

        bucket = sets.setdefault(set_name, [])
        if merge or len(varlist) > 1:
            if not group:
                problems.append(
                    f"[{SHEET_BANNER}] {line}행: 여러 변수를 묶을 때는 '그룹명' 이 필요합니다."
                )
                continue
            bucket.append(BannerSpec(kind="merge", label=group, varlist=varlist))
        else:
            bucket.append(BannerSpec(kind="single", var=varlist[0]))
    return sets


def _parse_filter(raw: str, df: pd.DataFrame, meta, where: str,
                  problems: list[str]) -> tuple[str | None, bool]:
    """(조건, 정상여부). 필터를 적었는데 해석하지 못하면 정상여부가 False 다.

    이때 그 줄은 표를 만들지 않는다. 필터 없이 계산해 버리면 틀린 숫자가
    맞는 것처럼 나오기 때문이다.
    """
    raw = _txt(raw)
    if not raw:
        return None, True
    if "=" not in raw:
        problems.append(
            f"{where}: 필터는 '변수=값' 으로 적어 주세요 (받은 값: {raw}). 이 줄은 건너뜁니다."
        )
        return None, False
    var, val = [p.strip() for p in raw.split("=", 1)]
    if var not in df.columns:
        problems.append(
            f"{where}: 필터 변수 '{var}' 가 이 .sav 에 없습니다. 이 줄은 건너뜁니다."
        )
        return None, False

    labels = meta.variable_value_labels.get(var, {})
    for code, label in labels.items():
        if str(label).strip() == val:
            return f"{var}={code}", True
    try:
        float(val)
    except ValueError:
        opts = ", ".join(str(v) for v in labels.values()) or "(값 라벨 없음)"
        problems.append(
            f"{where}: '{var}' 의 값 '{val}' 을 찾지 못했습니다. "
            f"쓸 수 있는 값 — {opts}. 이 줄은 건너뜁니다."
        )
        return None, False
    return f"{var}={val}", True


def _parse_sig(raw, where: str, problems: list[str]):
    """'유의성' 칸 → SigSpec. 비우거나 N 이면 안 함."""
    txt = _txt(raw).lower().replace("%", "").replace(" ", "")
    if not txt or txt in _NO:
        return None
    if txt in _YES:
        return SigSpec(enabled=True, level=0.95)
    try:
        num = float(txt)
    except ValueError:
        problems.append(
            f"{where}: '유의성' 값 '{txt}' 을 몰라 검정하지 않았습니다. "
            "쓸 수 있는 값 — 95, 99, 또는 비우기"
        )
        return None
    if num > 1:
        num /= 100.0
    if num not in (0.90, 0.95, 0.99):
        problems.append(
            f"{where}: 유의수준 {num:.0%} 는 관례에 없어 95% 로 봤습니다."
        )
        num = 0.95
    return SigSpec(enabled=True, level=num)


def _parse_min_base(raw, where: str, problems: list[str]) -> int:
    """'소표본' 칸 → 값을 감출 기준 사례수. 비우면 0(안 감춤)."""
    txt = _txt(raw)
    if not txt:
        return 0
    try:
        n = int(float(txt))
    except ValueError:
        problems.append(f"{where}: '소표본' 값 '{txt}' 이 숫자가 아니라 무시했습니다.")
        return 0
    return max(n, 0)


def _parse_summary_cell(raw, row_vars: list[str], meta, *, decimals: int = 1,
                        mean_decimals: int = 2):
    """'요약' 칸 → (SummarySpec 목록, 문제 목록).

    보기 코드는 첫 문항의 값 라벨에서 가져온다 (척도가 같은 문항끼리 묶는
    것이 전제다). 값 라벨이 없으면 만들 수 없다.
    """
    txt = _txt(raw)
    if not txt:
        return [], []
    codes = sorted(meta.variable_value_labels.get(row_vars[0], {}).keys())
    if not codes:
        return [], [f"'{row_vars[0]}' 에 값 라벨이 없어 요약을 만들 수 없습니다."]
    return parse_summary_spec(txt, codes, decimals=decimals,
                              mean_decimals=mean_decimals)


def read_form(data: bytes, df: pd.DataFrame, meta) -> tuple[list[TableBlock], list[str]]:
    """채운 양식을 읽어 (표 정의들, 문제 목록) 을 돌려준다.

    한 줄에 문제가 있으면 그 줄만 건너뛰고 나머지는 만든다.
    """
    problems: list[str] = []
    columns = list(df.columns)

    try:
        sets = _banner_sets(data, df, meta, problems)
    except ValueError as e:
        raise ValueError(f"'{SHEET_BANNER}' 시트를 읽지 못했습니다 — {e}") from e

    try:
        frame = _read_sheet(data, SHEET_TABLES, _TABLE_COLS,
                            set(columns), ["문항변수"])
    except ValueError as e:
        raise ValueError(f"'{SHEET_TABLES}' 시트를 읽지 못했습니다 — {e}") from e

    if not sets:
        problems.append(f"'{SHEET_BANNER}' 시트에 쓸 수 있는 배너가 없습니다.")
        return [], problems

    default_set = DEFAULT_SET if DEFAULT_SET in sets else next(iter(sets))
    blocks: list[TableBlock] = []

    for i, r in frame.iterrows():
        line = int(i) + 2
        where = f"[{SHEET_TABLES}] {line}행"

        type_raw = _txt(r["문항유형"]).lower()
        row_type = _TYPE_MAP.get(type_raw)
        if not row_type:
            problems.append(
                f"{where}: '문항유형' 을 알 수 없습니다 (받은 값: "
                f"{_txt(r['문항유형']) or '빈칸'} / 쓸 수 있는 값: 단수, 복수, "
                "수치형, 척도종합, 평균서머리)"
            )
            continue

        row_vars = _split_vars(r["문항변수"], columns)
        if not row_vars:
            problems.append(f"{where}: '문항변수' 가 비어 있습니다.")
            continue
        gone = [v for v in row_vars if v not in df.columns]
        if gone:
            problems.append(f"{where}: 이 .sav 에 없는 변수 — {', '.join(gone)}")
            continue
        multi_ok = row_type in ("multi", "battery", "battery_grid")
        if not multi_ok and len(row_vars) > 1:
            problems.append(
                f"{where}: 문항유형이 '{_txt(r['문항유형'])}' 인데 변수가 "
                f"{len(row_vars)}개입니다. '복수' 로 적거나 변수를 하나만 남겨 주세요."
            )
            continue

        # 보기 분포형 종합표는 행이 문항, 열이 보기라서 배너를 쓰지 않는다
        needs_banner = row_type != "battery"
        set_name = _txt(r["배너세트"]) or default_set
        banners = sets.get(set_name)
        if not banners:
            if needs_banner:
                problems.append(
                    f"{where}: 배너세트 '{set_name}' 을 '{SHEET_BANNER}' 시트에서 "
                    f"찾지 못했습니다. (있는 세트: {', '.join(sets)})"
                )
                continue
            banners = []

        orient_raw = _txt(r["배너방향"]).lower()
        orientation = _ORIENT_MAP.get(orient_raw, BANNER_ROW)
        if orient_raw and orient_raw not in _ORIENT_MAP:
            problems.append(f"{where}: '배너방향' 값 '{orient_raw}' 을 몰라 '행' 으로 봤습니다.")

        extra_cond, filter_ok = _parse_filter(r["필터"], df, meta, where, problems)
        if not filter_ok:
            continue

        dec_raw = _txt(r["소수점"])
        try:
            decimals = int(float(dec_raw)) if dec_raw else (2 if row_type == "obser" else 1)
        except ValueError:
            problems.append(f"{where}: '소수점' 값 '{dec_raw}' 이 숫자가 아니라 기본값을 씁니다.")
            decimals = 2 if row_type == "obser" else 1

        total_raw = _txt(r["계표시"]).lower()
        show_total = total_raw not in _NO if total_raw else (row_type != "obser")

        sort_raw = _txt(r["정렬"]).lower()
        sort_values = sort_raw in _YES if sort_raw else False
        if sort_raw and sort_raw not in _YES and sort_raw not in _NO:
            problems.append(f"{where}: '정렬' 값 '{sort_raw}' 을 몰라 정렬하지 않았습니다.")

        sig = _parse_sig(r["유의성"], where, problems)
        min_base_show = _parse_min_base(r["소표본"], where, problems)

        title = _txt(r["표제목"]) or row_vars[0]

        # ── 척도 종합표 ──
        if row_type in ("battery", "battery_grid"):
            grid = row_type == "battery_grid"
            summaries, sum_problems = _parse_summary_cell(
                r["요약"], row_vars, meta, decimals=1
            )
            for msg in sum_problems:
                problems.append(f"{where}: {msg}")

            metric = None
            if grid:
                metric_raw = _txt(r["통계"]).lower().replace(" ", "")
                metric = _METRIC_MAP.get(metric_raw)
                if not metric:
                    if metric_raw:
                        problems.append(
                            f"{where}: '평균서머리' 의 '통계' 값 '{metric_raw}' 을 몰라 "
                            "평균으로 봤습니다. 쓸 수 있는 값 — 평균, 표준편차, "
                            "Top2, Bottom2, 중간"
                        )
                    metric = "mean"
                # Top2 같은 묶음 지표는 그 묶음 정의가 있어야 계산된다
                if metric not in ("mean", "std") and not any(
                    s.label == metric for s in summaries
                ):
                    auto = {"Top2": "상2", "Top3": "상3", "Bottom2": "하2",
                            "Bottom3": "하3", "Middle": "중"}[metric]
                    extra, _p = _parse_summary_cell(
                        "상2,중,하2" if metric == "Middle" else auto,
                        row_vars, meta, decimals=1,
                    )
                    summaries = summaries + [s for s in extra
                                             if s.label == metric]
                    if not any(s.label == metric for s in summaries):
                        problems.append(
                            f"{where}: '{metric}' 를 만들 보기 정보가 없어 이 줄을 "
                            "건너뜁니다. '요약' 칸에 상2 처럼 적어 주세요."
                        )
                        continue

            show_raw = _txt(r["표시"]).lower().replace(" ", "") or "%"
            as_pct = show_raw not in ("n", "빈도", "사례수")
            # 평균·표준편차 격자는 값이 비율이 아니므로 제목에 %/N 을 붙이지 않는다
            marker = None if metric in ("mean", "std") else (
                "pct" if as_pct else "n")
            blocks.append(build_battery_block(
                battery_vars=row_vars,
                title=title_with_marker(title, marker),
                banners=banners if grid else None,
                metric=metric,
                summaries=summaries,
                show_pct=as_pct,
                decimals=decimals if decimals else 1,
                show_total_row=show_total,
                extra_cond=extra_cond,
                orientation=orientation,
                sig=sig,
                min_base_show=min_base_show,
                sort_rows=sort_values,
            ))
            continue

        if row_type == "obser":
            stats_raw = _txt(r["통계"])
            stats: list[str] = []
            for piece in [p.strip().lower() for p in stats_raw.replace(";", ",").split(",")]:
                if not piece:
                    continue
                mapped = _STAT_MAP.get(piece)
                if mapped:
                    if mapped not in stats:
                        stats.append(mapped)
                else:
                    problems.append(f"{where}: '통계' 의 '{piece}' 를 몰라 건너뜁니다.")

            # 수치형에서 '표시' 를 적으면 응답된 값의 분포를 단수 표처럼 보여주고,
            # 그 뒤에 평균·중위값 같은 통계 칸을 붙인다. 비우면 통계만 나온다.
            show_raw = _txt(r["표시"]).lower().replace(" ", "")
            if show_raw:
                kinds = _SHOW_MAP.get(show_raw)
                if not kinds:
                    problems.append(
                        f"{where}: '표시' 값 '{show_raw}' 을 몰라 통계만 넣었습니다."
                    )
                    kinds = (None,)
            else:
                kinds = (None,)
                if total_raw in _YES:
                    problems.append(
                        f"{where}: '계' 는 응답값 분포가 있어야 나옵니다. "
                        "'표시' 에 % 나 N 을 적어 주세요. 이 표는 통계만 넣었습니다."
                    )

            # 수치형의 요약 — 값 라벨이 없으니 실제 응답된 값을 보기로 본다.
            # (평균·표준편차는 보기와 무관하게 값 자체로 계산된다)
            summaries = []
            summary_raw = _txt(r["요약"])
            if summary_raw:
                codes = sorted(meta.variable_value_labels.get(row_vars[0], {}).keys())
                if not codes:
                    codes = sorted(df[row_vars[0]].dropna().unique().tolist())
                summaries, sum_problems = parse_summary_spec(
                    summary_raw, codes, decimals=1, mean_decimals=decimals
                )
                for msg in sum_problems:
                    problems.append(f"{where}: {msg}")
                # '통계' 에 이미 평균이 있으면 요약의 평균은 같은 숫자라 뺀다.
                # ('통계' 를 비우면 네 가지가 모두 들어가므로 평균도 이미 있다)
                if "MEAN" in (stats or ["MEAN", "MEDIAN", "MIN", "MAX"]):
                    summaries = [s for s in summaries if s.kind != "mean"]

            # 표가 넓어진다는 안내는 줄마다 한 번만 (%+N 이면 표는 두 개라서)
            warned = False

            for kind in kinds:
                show_values = kind is not None
                block = build_block(
                    row_type="obser", row_vars=row_vars, banners=banners,
                    title=title_with_marker(title, kind),
                    obser_stats=stats or None, extra_cond=extra_cond,
                    obser_decimals=decimals, orientation=orientation,
                    obser_show_values=show_values,
                    show_pct=(kind == "pct"),
                    decimals=1,               # 분포 % 는 소수점 1자리 (통계는 '소수점' 칸)
                    show_total_row=show_total,
                    summaries=summaries,
                    sig=sig,
                    min_base_show=min_base_show,
                    sort_values=sort_values,
                )
                if show_values and not warned:
                    size = value_breakdown_size(df, block)
                    if size > BREAKDOWN_WARN:
                        warned = True
                        problems.append(
                            f"{where}: '{row_vars[0]}' 는 응답된 값이 {size}종이라 "
                            f"보기가 {size}개 나옵니다. 표가 너무 넓으면 '표시' 를 비워 "
                            "통계만 내거나, 값을 묶은 변수를 쓰세요."
                        )
                blocks.append(block)
            continue

        # ── 척도 요약 (Top2 · Middle · Bottom2 · 평균) ──
        summaries = []
        if _txt(r["요약"]):
            if row_type == "multi":
                problems.append(
                    f"{where}: 요약(Top2 등)은 단수 문항에서만 뜻이 있어 무시했습니다."
                )
            else:
                summaries, sum_problems = _parse_summary_cell(
                    r["요약"], row_vars, meta, decimals=decimals
                )
                for msg in sum_problems:
                    problems.append(f"{where}: {msg}")

        show_raw = _txt(r["표시"]).lower().replace(" ", "") or "%"
        shows = _SHOW_MAP.get(show_raw)
        if not shows:
            problems.append(f"{where}: '표시' 값 '{show_raw}' 을 몰라 '%' 로 봤습니다.")
            shows = ("pct",)

        for kind in shows:
            blocks.append(build_block(
                row_type=row_type, row_vars=row_vars, banners=banners,
                title=title_with_marker(title, kind),
                extra_cond=extra_cond,
                show_pct=(kind == "pct"), decimals=decimals,
                show_total_row=show_total, orientation=orientation,
                summaries=summaries,
                sig=sig,
                min_base_show=min_base_show,
                sort_values=sort_values,
            ))

    if not blocks and not problems:
        problems.append(f"'{SHEET_TABLES}' 시트에 표가 없습니다.")
    return blocks, problems


# =============================================================================
# 지금 만든 표들을 양식으로 내보내기 (되돌려 쓰기)
# =============================================================================
def blocks_to_form(blocks: list[TableBlock], df: pd.DataFrame, meta) -> bytes:
    """화면에서 만든 표들을 채워진 양식으로 내보낸다.
    다음에 조금 고쳐서 다시 올리기 쉽게 하려는 것."""
    from openpyxl import load_workbook

    data = write_form_template(df, meta, example=False)
    wb = load_workbook(io.BytesIO(data))
    ws, wsb = wb[SHEET_TABLES], wb[SHEET_BANNER]

    # 배너 세트: 표들이 쓰는 배너를 모아 이름을 붙인다
    seen: dict[str, str] = {}
    banner_rows: list[list] = []

    for b in blocks:
        key_parts = []
        for token in b.banner_axis:
            if token == "@t3":
                continue
            m = next((x for x in b.merges if x.name == token), None)
            key_parts.append(f"{m.label}:{m.varlist_raw}" if m else token)
        key = "|".join(key_parts)
        # 배너를 쓰지 않는 표(보기 분포형 종합표)는 세트를 만들지 않는다.
        # 빈 세트에 '기본' 이름을 붙여 두면, 정작 배너를 쓰는 표들이 '세트2' 로
        # 밀려서 되읽을 때 '기본' 을 못 찾는다.
        if not key:
            continue
        if key in seen:
            continue
        name = DEFAULT_SET if not seen else f"세트{len(seen) + 1}"
        seen[key] = name
        for token in b.banner_axis:
            if token == "@t3":
                continue
            m = next((x for x in b.merges if x.name == token), None)
            if m:
                banner_rows.append([name, m.label,
                                    ", ".join(expand_var_range(m.varlist_raw, list(df.columns))),
                                    "Y"])
            else:
                banner_rows.append([name, "", token, ""])

    for i, row in enumerate(banner_rows):
        for j, v in enumerate(row, start=1):
            wsb.cell(row=3 + i, column=j, value=v)

    def set_key(b: TableBlock) -> str:
        parts = []
        for token in b.banner_axis:
            if token == "@t3":
                continue
            m = next((x for x in b.merges if x.name == token), None)
            parts.append(f"{m.label}:{m.varlist_raw}" if m else token)
        return "|".join(parts)

    stat_ko = {"MEAN": "평균", "MEDIAN": "중위값", "MIN": "최소값", "MAX": "최대값"}
    metric_ko = {"mean": "평균", "std": "표준편차"}
    line = 3
    for b in blocks:
        if b.is_battery:
            row_vars = list(b.battery_vars)
            row_type = "평균서머리" if b.battery_metric else "척도종합"
            show = "%" if any(s.name == "cpct" for s in b.stats) else "N"
            stats = metric_ko.get(b.battery_metric or "", b.battery_metric or "")
            dec = next((int(s.fmt.split(".")[-1]) for s in b.stats
                        if s.name == "cpct"), 1)
        elif b.is_obser:
            row_type, row_vars = "수치형", [b.obser_var]
            stats = ",".join(stat_ko.get(s.name, s.name)
                             for s in b.stats if s.var == b.obser_var)
            dec = next(
                (int(s.fmt.split(".")[-1]) for s in b.stats if s.var == b.obser_var), 2
            )
            # 응답값 분포를 함께 넣은 표라면 표시 칸에 %/N 이 들어간다
            if b.row_var_merge is not None:
                show = "%" if any(s.name == "cpct" for s in b.stats) else "N"
            else:
                show = ""
        else:
            merge = b.row_var_merge
            row_vars = expand_var_range(merge.varlist_raw, list(df.columns)) if merge else []
            row_type = "복수" if len(row_vars) > 1 else "단수"
            has_pct = any(s.name == "cpct" for s in b.stats)
            show = "%" if has_pct else "N"
            stats = ""
            dec = next((int(s.fmt.split(".")[-1]) for s in b.stats if s.name == "cpct"), 1)

        filt = ""
        if b.extra_cond and "=" in b.extra_cond:
            var, val = b.extra_cond.split("=", 1)
            labels = meta.variable_value_labels.get(var, {})
            try:
                label = labels.get(float(val))
            except ValueError:
                label = None
            filt = f"{var}={label or val}"

        summary_txt = ""
        if getattr(b, "summaries", None):
            parts = []
            for x in b.summaries:
                if x.kind == "mean":
                    parts.append("평균")
                elif x.kind == "std":
                    parts.append("표준편차")
                elif x.label.startswith("Top"):
                    parts.append(f"상{len(x.codes)}")
                elif x.label.startswith("Bottom"):
                    parts.append(f"하{len(x.codes)}")
                elif x.label == "Middle":
                    parts.append("중")
                else:
                    codes = ",".join(str(int(c)) if float(c).is_integer() else str(c)
                                     for c in x.codes)
                    parts.append(f"{x.label}={codes}")
            summary_txt = ";".join(parts) if any("=" in p for p in parts) else ",".join(parts)

        sig_txt = ""
        if b.sig and b.sig.enabled:
            sig_txt = f"{b.sig.level:.0%}".replace("%", "")

        values = [
            b.title, row_type, ", ".join(row_vars),
            show, summary_txt, stats, dec,
            "Y" if "t1" in b.value_axis else "N",
            "Y" if b.sort_values else "N",
            sig_txt,
            b.min_base_show or "",
            "행" if b.orientation == BANNER_ROW else "열",
            filt,
            seen.get(set_key(b), "") if set_key(b) else "",
        ]
        for j, v in enumerate(values, start=1):
            ws.cell(row=line, column=j, value=v)
        line += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# =============================================================================
# 8) .sav 만 보고 양식을 자동으로 채우기
# =============================================================================
# 변수 라벨(문항 문구)과 값 라벨(보기)이 .sav 안에 있으므로, 표 목록의 상당
# 부분은 데이터만 보고 채울 수 있습니다. 사람이 판단해야 하는 것(어떤 변수를
# 배너로 쓸지, 어떤 문항을 뺄지)은 제안만 하고 지우기 쉽게 둡니다.

ID_LABEL_LIMIT = 30        # 값 라벨이 이보다 많으면 문항이 아니라 명칭/ID 로 본다
BANNER_MAX_CATEGORIES = 20  # 배너 후보로 제안할 보기 개수 상한
BANNER_SUGGEST_LIMIT = 6    # 배너 후보 제안 개수


def _numeric_suffix_family(columns: list[str]) -> dict[str, list[str]]:
    """'앞부분 + 숫자' 이름이 이어지는 변수들을 묶음 후보로 모은다.
    예: 봉안시설_1 ~ 봉안시설_4 → {'봉안시설': [...]}"""
    import re

    fam: dict[str, list[str]] = {}
    for c in columns:
        m = re.match(r"^(.*?)[_\-]?(\d+)$", c)
        if m and m.group(1):
            fam.setdefault(m.group(1), []).append(c)
    return {k: v for k, v in fam.items() if len(v) > 1}


def _is_category_coded_set(df: pd.DataFrame, members: list[str]) -> bool:
    """다중응답(카테고리 코딩)인지 — 변수마다 '자기 코드값' 하나만 갖는지.

    같은 값 라벨을 쓰는 묶음이라도, 각 변수가 보기 전체 범위를 값으로 가지면
    다중응답이 아니라 **평가 배터리**(문항마다 5점 척도 등)다. 그때는 변수
    하나하나가 별개의 단수 문항이므로 묶지 않는다.
    """
    for i, v in enumerate(members, start=1):
        vals = df[v].dropna().unique().tolist()
        if len(vals) != 1 or float(vals[0]) != float(i):
            return False
    return True


# 리커트 척도로 보이는 보기 라벨에 흔히 나오는 말
_SCALE_WORDS = (
    "매우", "전혀", "보통", "그렇", "아니", "만족", "불만", "동의",
    "좋", "싫", "많", "적", "높", "낮", "약간", "다소", "대체로",
)


def _scale_summary(labels: dict) -> str:
    """보기 라벨이 리커트 척도처럼 보이면 요약 정의를 만들어 준다.

    4~7점이고 라벨에 '매우 · 보통 · 그렇다' 같은 말이 두 개 이상 나오면
    척도로 본다. 아니면 빈칸(요약 없음)이다. 어디까지나 제안이므로 엑셀에서
    지우거나 고치면 된다.
    """
    if not (4 <= len(labels) <= 7):
        return ""
    texts = [str(v) for v in labels.values()]
    hits = sum(1 for t in texts if any(w in t for w in _SCALE_WORDS))
    if hits < 2:
        return ""
    return "상2,중,하2,평균" if len(labels) % 2 else "상2,하2,평균"


def suggest_form_rows(df: pd.DataFrame, meta) -> tuple[list[list], list[list], list[str]]:
    """.sav 를 보고 (표목록 줄들, 배너 줄들, 안내) 를 만든다.

    제외하는 것
      · 문자 변수 — 집계할 수 없다
      · 값 라벨이 아주 많거나 응답자마다 값이 다른 변수 — 명칭/ID 로 본다
      · 앞선 변수와 데이터가 완전히 같은 변수 — bv1/cv1/m~ 같은 파생 복제본
    """
    vl = meta.variable_value_labels
    cl = meta.column_names_to_labels
    columns = list(df.columns)
    notes: list[str] = []

    # ── 파생 복제본 찾기: 앞 변수와 데이터가 완전히 같으면 뒤쪽을 버린다 ──
    # 모든 변수쌍을 직접 비교하면 변수가 많을 때 느려지므로, 먼저 열마다
    # 지문(해시)을 구해 같은 지문끼리만 실제로 비교한다.
    fingerprint: dict[str, list[str]] = {}
    for c in columns:
        try:
            key = pd.util.hash_pandas_object(df[c], index=False).sum()
        except TypeError:                  # 해시가 안 되는 형이면 비교에서 제외
            continue
        fingerprint.setdefault(f"{df[c].dtype}|{key}", []).append(c)

    duplicate_of: dict[str, str] = {}
    for group in fingerprint.values():
        for i, a in enumerate(group):
            if a in duplicate_of:
                continue
            for b in group[i + 1:]:
                if b not in duplicate_of and df[a].equals(df[b]):
                    duplicate_of[b] = a
    if duplicate_of:
        sample = ", ".join(f"{b}={a}" for b, a in list(duplicate_of.items())[:4])
        notes.append(
            f"앞 변수와 데이터가 똑같은 변수 {len(duplicate_of)}개는 파생 복제본으로 보고 "
            f"뺐습니다 ({sample}…). 배너용 변수라면 '배너' 시트에서 쓰세요."
        )

    # ── 다중응답 묶음 찾기 ──
    ma_sets: dict[str, list[str]] = {}
    used_in_ma: set[str] = set()
    for prefix, members in _numeric_suffix_family(columns).items():
        members = [m for m in members if m not in duplicate_of and vl.get(m)]
        if len(members) < 2:
            continue
        first = tuple(sorted(vl[members[0]].items()))
        if any(tuple(sorted(vl[m].items())) != first for m in members):
            continue                       # 보기가 서로 다르면 한 문항이 아니다
        if not _is_category_coded_set(df, members):
            continue                       # 평가 배터리 → 각각 단수로 둔다
        ma_sets[prefix] = members
        used_in_ma.update(members)

    # ── 표목록 ──
    S = "평균,중위값,최소값,최대값"
    rows: list[list] = []
    skipped_text, skipped_id = [], []

    for col in columns:
        if col in duplicate_of or col in used_in_ma:
            continue
        labels = vl.get(col, {})
        title = cl.get(col) or col

        if labels:
            if len(labels) > ID_LABEL_LIMIT or df[col].dropna().nunique() == len(df):
                skipped_id.append(col)
                continue
            rows.append([title, "단수", col, "%+N", _scale_summary(labels),
                         "", 1, "Y", "행", "", ""])
        elif pd.api.types.is_numeric_dtype(df[col]):
            rows.append([title, "수치형", col, "", "", S, 2, "", "행", "", ""])
        else:
            skipped_text.append(col)

    # 다중응답 묶음은 원래 변수 순서에 맞춰 끼워 넣는다
    for prefix, members in ma_sets.items():
        base = cl.get(members[0]) or prefix
        title = prefix if prefix else base
        varlist = (f"{members[0]} to {members[-1]}"
                   if columns.index(members[-1]) - columns.index(members[0])
                   == len(members) - 1 else ", ".join(members))
        pos = min(len([r for r in rows
                       if columns.index(_txt(r[2]).split(",")[0].split(" to ")[0])
                       < columns.index(members[0])]), len(rows))
        rows.insert(pos, [title, "복수", varlist, "%+N", "", "", 1, "N", "행", "", ""])

    if ma_sets:
        notes.append(
            "다중응답으로 본 묶음: "
            + ", ".join(f"{p}({len(m)}개)" for p, m in ma_sets.items())
            + ". 변수마다 자기 코드값만 갖는 형태라서 묶었습니다."
        )
    if skipped_id:
        notes.append(
            f"보기가 너무 많거나 응답자마다 값이 다른 변수 {len(skipped_id)}개는 "
            f"명칭·ID 로 보고 뺐습니다 ({', '.join(skipped_id[:4])}…)."
        )
    if skipped_text:
        notes.append(
            f"문자 변수 {len(skipped_text)}개는 집계할 수 없어 뺐습니다 "
            f"({', '.join(skipped_text[:4])}…). 주관식이면 따로 정리하세요."
        )

    # ── 배너 후보 (사람이 정해야 하는 부분이라 '제안') ──
    banner_rows: list[list] = []
    for col in columns:
        if col in duplicate_of or col in used_in_ma:
            continue
        labels = vl.get(col, {})
        if labels and 2 <= len(labels) <= BANNER_MAX_CATEGORIES:
            banner_rows.append([DEFAULT_SET, "", col, ""])
        if len(banner_rows) >= BANNER_SUGGEST_LIMIT:
            break
    if banner_rows:
        notes.append(
            f"배너는 보기 {2}~{BANNER_MAX_CATEGORIES}개인 변수 "
            f"{len(banner_rows)}개를 후보로 넣어 뒀습니다. 실제로 쓸 것만 남기고 "
            "지우세요 — 배너 선택은 사람이 정해야 하는 부분입니다."
        )
    return rows, banner_rows, notes


def write_filled_form(df: pd.DataFrame, meta) -> tuple[bytes, list[str]]:
    """.sav 를 보고 자동으로 채운 양식을 만든다. (양식, 안내)"""
    from openpyxl import load_workbook

    rows, banner_rows, notes = suggest_form_rows(df, meta)
    wb = load_workbook(io.BytesIO(write_form_template(df, meta, example=False)))
    ws, wsb = wb[SHEET_TABLES], wb[SHEET_BANNER]

    for i, row in enumerate(rows):
        for j, v in enumerate(row, start=1):
            ws.cell(row=3 + i, column=j, value=v)
    for i, row in enumerate(banner_rows):
        for j, v in enumerate(row, start=1):
            wsb.cell(row=3 + i, column=j, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), notes
