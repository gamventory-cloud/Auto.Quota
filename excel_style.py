"""
╔══════════════════════════════════════════════════════════════════════════╗
║  파일명 : excel_style.py                                                  ║
║  위치   : 리포지토리 최상단  (utils.py 와 같은 폴더)                        ║
║                                                                          ║
║  원본 엑셀의 서식을 그대로 두고 값만 고쳐서 내보냅니다.                       ║
╚══════════════════════════════════════════════════════════════════════════╝

왜 이 파일이 필요한가
--------------------
pandas + xlsxwriter 로 다시 쓰면 **원본 서식이 전부 사라집니다**
(1행 색·굵게, 열 너비, 틀 고정, 표시 형식 등).
openpyxl 로 원본 통합문서를 열어 셀 값만 바꾸면 나머지는 손대지 않아도
그대로 남습니다. 특정 색을 코드에 박아넣지 않고 **원본에서 읽어 씁니다.**

핵심 동작
--------
  - Data/Label 시트 : 맨 위에 새 변수명 행을 끼워 넣고, 원래 머리행과
                      똑같은 서식을 복사해 붙입니다 (색·굵게·정렬·글꼴).
  - Code 시트       : 해당 셀의 값만 바꿉니다. 칠해진 배경은 그대로 유지됩니다.
  - 그 외 시트      : 건드리지 않습니다.
"""

from __future__ import annotations

import io
import re
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 원본에서 서식을 찾지 못했을 때만 쓰는 기본값 (연노랑 + 굵게)
DEFAULT_FILL = "FFFFCC"


def _has_fill(cell) -> bool:
    return bool(cell.fill and cell.fill.fill_type and cell.fill.fill_type != "none")


def describe_style(cell) -> dict:
    """셀 서식을 사람이 읽을 수 있게 요약 (화면에 '인식된 서식' 표시용)."""
    color = None
    if _has_fill(cell):
        rgb = getattr(cell.fill.start_color, "rgb", None)
        if isinstance(rgb, str):
            color = rgb[-6:]          # 'FFFFFFCC' -> 'FFFFCC'
    return {
        "채우기": color,
        "굵게": bool(cell.font and cell.font.bold),
        "글꼴": cell.font.name if cell.font else None,
        "크기": cell.font.size if cell.font else None,
        "가운데": (cell.alignment.horizontal == "center") if cell.alignment else False,
    }


def find_marked_style(ws, max_scan: int = 200):
    """시트에서 '칠해진' 셀을 찾아 그 서식을 돌려준다.

    Code 시트처럼 머리행이 따로 없고 문항 구분 행만 칠해진 경우에 쓴다.
    """
    for row in ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row)):
        for cell in row:
            if _has_fill(cell) and cell.value is not None:
                return cell
    return None


def _style_source(wb, sheet_names: list[str], header_row: int = 1):
    """새 행에 입힐 서식의 출처 셀을 고른다.

    1) 대상 시트의 머리행이 칠해져 있으면 그 셀
    2) 아니면 통합문서 어디든 칠해진 셀 (Code 시트의 문항 행 등)
    3) 둘 다 없으면 None -> 기본값 사용
    """
    for name in sheet_names:
        ws = wb[name]
        for c in range(1, min(ws.max_column, 20) + 1):
            cell = ws.cell(header_row, c)
            if _has_fill(cell):
                return cell
    for ws in wb.worksheets:
        cell = find_marked_style(ws)
        if cell is not None:
            return cell
    return None


def _apply_style(target, source) -> None:
    """source 셀의 서식을 target 에 복사. source 가 없으면 기본 서식."""
    if source is not None:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        return
    target.font = Font(name="맑은 고딕", size=9, bold=True)
    target.fill = PatternFill("solid", fgColor=DEFAULT_FILL)
    target.alignment = Alignment(horizontal="center", vertical="center")


def _shift_freeze(ws, rows: int = 1) -> None:
    """틀 고정을 아래로 밀어 새 머리행까지 고정되게 한다 ('A2' -> 'A3')."""
    if not ws.freeze_panes:
        return
    m = re.fullmatch(r"([A-Z]+)(\d+)", str(ws.freeze_panes))
    if m:
        ws.freeze_panes = f"{m.group(1)}{int(m.group(2)) + rows}"


def export_renamed_workbook(original_bytes: bytes,
                            new_header: list[str] | None,
                            rename_map: dict[str, str],
                            target_sheets: list[str],
                            code_sheet: str | None = None,
                            code_updates: dict[int, str] | None = None,
                            keep_old_header: bool = True) -> bytes:
    """원본 서식을 유지한 채 변수명을 반영한 통합문서를 만든다.

    new_header    : 열 순서대로 정렬된 새 변수명 목록 (열 개수가 맞을 때 우선 사용).
                    pandas 가 중복 열 이름을 'q1.1' 로 바꿔버리는 경우가 있어
                    이름 대조보다 위치 대조가 안전하다.
    rename_map    : {원래 열 이름: 새 이름} — 열 개수가 맞지 않을 때 사용.
    target_sheets : 머리행을 새로 끼워 넣을 시트 이름 목록.
    code_updates  : {행 인덱스(0부터): 새 이름} — Code 시트 A열에 덮어쓸 값.
    keep_old_header: True 면 원래 머리행을 두 번째 행으로 남긴다.
    """
    wb = load_workbook(io.BytesIO(original_bytes))
    targets = [s for s in target_sheets if s in wb.sheetnames and s != code_sheet]
    style_cell = _style_source(wb, targets)

    for name in targets:
        ws = wb[name]
        old_header = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else ""
                      for c in range(1, ws.max_column + 1)]

        if new_header and len(new_header) == ws.max_column:
            names = list(new_header)
        else:
            names = [rename_map.get(h, h) for h in old_header]

        if keep_old_header:
            ws.insert_rows(1)
            src_row = 2                      # 원래 머리행이 한 칸 내려간 위치
        else:
            src_row = 1

        for c, value in enumerate(names, start=1):
            cell = ws.cell(1, c)
            cell.value = value
            # 서식은 '원래 머리행' 것을 그대로 복사한다 (색을 코드에 박지 않음)
            source = ws.cell(src_row, c) if _has_fill(ws.cell(src_row, c)) else style_cell
            _apply_style(cell, source)

        if keep_old_header:
            ws.row_dimensions[1].height = ws.row_dimensions[2].height
            _shift_freeze(ws, 1)

    if code_sheet and code_sheet in wb.sheetnames and code_updates:
        ws = wb[code_sheet]
        for r_idx, new_name in code_updates.items():
            row = int(r_idx) + 1          # pandas(header=None) 0-based -> 엑셀 1-based
            if 1 <= row <= ws.max_row:
                ws.cell(row, 1).value = new_name   # 값만 교체, 배경색은 그대로

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def preview_header_style(original_bytes: bytes, sheet: str, header_row: int = 1) -> dict:
    """화면에 '원본에서 인식한 서식'을 보여주기 위한 요약."""
    wb = load_workbook(io.BytesIO(original_bytes), read_only=False)
    if sheet not in wb.sheetnames:
        return {}
    ws = wb[sheet]
    for c in range(1, min(ws.max_column, 20) + 1):
        cell = ws.cell(header_row, c)
        if _has_fill(cell) or (cell.font and cell.font.bold):
            info = describe_style(cell)
            info["출처"] = f"{sheet}!{get_column_letter(c)}{header_row}"
            return info
    marked = find_marked_style(ws)
    if marked is not None:
        info = describe_style(marked)
        info["출처"] = f"{sheet}!{marked.coordinate}"
        return info
    return {}
